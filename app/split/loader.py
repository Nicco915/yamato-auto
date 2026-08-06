# -*- coding: utf-8 -*-
"""读取 filled ContentsOfTheContainer Excel，返回 list[RawItem]。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from app.split.schemas import RawItem


def load_filled_excel(path: str | Path) -> list[RawItem]:
    """读取 filled ContentsOfTheContainer，只取有数据的行（跳过表头）。

    按表头列名定位（列顺序变化不影响），所需列：
      KANRI_NO、MINATO_MEI_KJ、CONTAINER_MEI、MAKER_MEI_KJ、
      SHOHIN_CD、净重、毛重、SOTOBAKO_D_HACCHU_SU、
      中文品名、D_HACCHU_SU、KAKAKUKEI、TSUKA_MEI（后 4 列供报关生成）

    表头在第 1 行，数据从第 2 行开始。跳过 KANRI_NO 为空的行。
    缺列时抛 ValueError。
    """
    REQUIRED = [
        "KANRI_NO", "MINATO_MEI_KJ", "CONTAINER_MEI", "MAKER_MEI_KJ",
        "SHOHIN_CD", "净重", "毛重", "SOTOBAKO_D_HACCHU_SU",
        "中文品名", "D_HACCHU_SU", "KAKAKUKEI", "TSUKA_MEI",
    ]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [n for n in REQUIRED if n not in header]
    if missing:
        wb.close()
        raise ValueError(f"filled Excel 缺少必需列: {missing}")
    col = {name: header.index(name) + 1 for name in REQUIRED}  # 1-based

    def num(v, as_int=False):
        if v is None:
            return None
        try:
            return int(v) if as_int else float(v)
        except (ValueError, TypeError):
            return None

    items: list[RawItem] = []
    for row_idx in range(2, ws.max_row + 1):
        kanri_val = ws.cell(row=row_idx, column=col["KANRI_NO"]).value
        if kanri_val is None:
            continue
        kanri_no = str(kanri_val).strip()
        if not kanri_no:
            continue

        def text(name):
            v = ws.cell(row=row_idx, column=col[name]).value
            return str(v).strip() if v is not None else ""

        items.append(RawItem(
            kanri_no=kanri_no,
            port=text("MINATO_MEI_KJ"),
            container_type=text("CONTAINER_MEI"),
            maker=text("MAKER_MEI_KJ"),
            sku=text("SHOHIN_CD"),
            net_weight=num(ws.cell(row=row_idx, column=col["净重"]).value),
            gross_weight=num(ws.cell(row=row_idx, column=col["毛重"]).value),
            pcs=num(ws.cell(row=row_idx, column=col["SOTOBAKO_D_HACCHU_SU"]).value, as_int=True),
            name_cn=text("中文品名"),
            qty_pieces=num(ws.cell(row=row_idx, column=col["D_HACCHU_SU"]).value, as_int=True),
            amount=num(ws.cell(row=row_idx, column=col["KAKAKUKEI"]).value),
            currency=text("TSUKA_MEI"),
        ))

    wb.close()
    return items