# -*- coding: utf-8 -*-
"""通道二：文本降维大模型通道（Excel / CSV）。

流程（第二阶段设计文档第 2 节）：
1. openpyxl 拆分合并单元格（把合并值填到每个单元格），不执行任何业务逻辑抓取；
2. sheet 级目标识别：只把箱单（Packing List）sheet 转成 Markdown，
   发票/合同/报关 sheet 不送 LLM（见 select_pl_sheets）；
3. 交给文本大模型做语义提取，强制 JSON 输出，解析失败最多重试 2 次。

旧版 .xls 用 pandas(xlrd) 读取；读不出的记入 unsupported。
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from . import llm_client
from .prompts import (
    SYSTEM_PROMPT,
    JsonParseError,
    build_retry_message,
    build_text_user_prompt,
    parse_payload,
)
from .schemas import ExtractedItem, apply_weight_basis
from .verify import verify_weight_basis

# 单个文件转成 Markdown 后的长度上限（控制 token 成本；
# 2026-07-27 由 24000 提高到 100000：正达 INV+PL 两 sheet 合计超限被截断，
# 导致 PL sheet 后半 29 个 SKU 丢失）
MAX_MARKDOWN_CHARS = 100000
# 单表最多保留的行列（防止巨型空表撑爆上下文）
MAX_ROWS = 400
MAX_COLS = 30

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls", ".csv"}
# 旧版 .xls 走 pandas/xlrd，不做合并单元格拆分
LEGACY_XLS_SUFFIXES = {".xls"}

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sheet 级目标识别（2026-07-27 人工核对引入）
# 箱单（Packing List）是提取的唯一目标 sheet；发票/合同/报关单 sheet 不送 LLM，
# 既省 token 又消除 PCS/单价等干扰数据。
# ---------------------------------------------------------------------------
# 明确的箱单 sheet 名（命中即只取这些 sheet）
PL_SHEET_PATTERNS = [r"(?i)\bPL\b", r"(?i)packing", "箱单", "装箱单"]
# 明确的非箱单 sheet 名（无箱单 sheet 时用于排除，兜底保留其余）
NON_PL_SHEET_PATTERNS = [
    r"(?i)\bINV\b", r"(?i)invoice", "发票",
    r"(?i)contract", "合同", "报关",
]


def select_pl_sheets(names: list[str]) -> list[str]:
    """从 sheet 名列表中选出箱单 sheet。

    规则：有明确箱单 sheet → 只取它们；
    没有 → 排除明确非箱单的，保留其余（如 Sheet1 这类无名 sheet）；
    全被排除 → 原样保留（宁多勿漏）。
    """
    pl = [n for n in names if any(re.search(p, n) for p in PL_SHEET_PATTERNS)]
    if pl:
        selected = pl
    else:
        rest = [n for n in names if not any(re.search(p, n) for p in NON_PL_SHEET_PATTERNS)]
        selected = rest or list(names)
    excluded = [n for n in names if n not in selected]
    logger.info("箱单 sheet 筛选 | 选中=%s | 排除=%s", selected, excluded or "（无）")
    return selected


class UnsupportedFileError(RuntimeError):
    """文件无法被本通道读取。"""


@dataclass
class ChannelResult:
    """单文件提取结果 + 统计信息。"""

    items: list[ExtractedItem] = field(default_factory=list)
    json_attempts: int = 0  # LLM 调用次数（含解析失败重试）
    json_parse_failures: int = 0  # JSON 解析失败次数
    error: str = ""
    notes: list[str] = field(default_factory=list)  # 确定性校验备注（verify.py）


# ---------------------------------------------------------------------------
# 第一步：纯格式降维（不调 LLM，可独立测试）
# ---------------------------------------------------------------------------

def _unmerge_sheet_to_grid(ws) -> list[list]:
    """把一个 worksheet 拆平为二维网格：合并单元格的值填充到所有覆盖单元格。"""
    merged_lookup: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        top_value = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_lookup[(r, c)] = top_value

    max_row = min(ws.max_row or 0, MAX_ROWS)
    max_col = min(ws.max_column or 0, MAX_COLS)
    grid: list[list] = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            if (r, c) in merged_lookup:
                row.append(merged_lookup[(r, c)])
            else:
                row.append(ws.cell(row=r, column=c).value)
        grid.append(row)
    return grid


def _grid_to_dataframe(grid: list[list]) -> pd.DataFrame:
    """网格 → DataFrame，裁掉全空的首尾行列。"""
    df = pd.DataFrame(grid)
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    # 去掉纯空白字符串行/列
    if not df.empty:
        non_blank = df.apply(lambda col: col.map(lambda v: str(v).strip() if v is not None else ""))
        df = df.loc[(non_blank != "").any(axis=1), (non_blank != "").any(axis=0)]
    return df.reset_index(drop=True)


def _df_to_markdown(df: pd.DataFrame) -> str:
    """DataFrame → Markdown 表格纯文本（无表头概念，全部按数据行处理）。"""
    if df.empty:
        return ""
    # A1：行内相邻同值长字符串压缩（在转为字符串前操作 DataFrame 原值）
    df = _compress_adjacent_dup_strings(df)
    # A2a：单射形态条码回填（在压缩后做，新增列会被 text_df.map 自动转字符串）
    df = _bind_orphan_barcode(df)
    text_df = df.map(lambda v: "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).replace("\n", " ").replace("|", "｜").strip())
    buf = io.StringIO()
    ncols = len(text_df.columns)
    buf.write("| " + " | ".join(f"c{i}" for i in range(ncols)) + " |\n")
    buf.write("| " + " | ".join("---" for _ in range(ncols)) + " |\n")
    for _, row in text_df.iterrows():
        buf.write("| " + " | ".join(row.tolist()) + " |\n")
    return buf.getvalue()


def xlsx_to_markdown(file_path: str) -> str:
    """xlsx/xlsm：openpyxl 拆合并单元格 → 箱单 sheet 转 Markdown。"""
    wb = load_workbook(file_path, data_only=True, read_only=False)
    selected = set(select_pl_sheets([ws.title for ws in wb.worksheets]))
    parts: list[str] = []
    for ws in wb.worksheets:
        if ws.title not in selected:
            continue
        grid = _unmerge_sheet_to_grid(ws)
        df = _grid_to_dataframe(grid)
        md = _df_to_markdown(df)
        if md:
            parts.append(f"### 工作表: {ws.title}\n{md}")
    return "\n\n".join(parts)


def legacy_xls_to_markdown(file_path: str) -> str:
    """旧版 .xls：pandas/xlrd 逐 sheet 读入 → Markdown（无法拆合并单元格，靠前向填充缓解）。"""
    try:
        sheets = pd.read_excel(file_path, sheet_name=None, header=None, engine="xlrd")
    except Exception as e:  # noqa: BLE001
        raise UnsupportedFileError(f"xlrd 无法读取 {file_path}: {e}") from e
    selected = set(select_pl_sheets(list(sheets.keys())))
    parts: list[str] = []
    for name, df in sheets.items():
        if name not in selected:
            continue
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        # 合并单元格在 xlrd 中表现为只有左上角有值，做有限的前向填充（仅行内向右填）
        df = df.ffill(axis=1)
        md = _df_to_markdown(df.reset_index(drop=True))
        if md:
            parts.append(f"### 工作表: {name}\n{md}")
    return "\n\n".join(parts)


def csv_to_markdown(file_path: str) -> str:
    """CSV：尝试常见中文编码读入 → Markdown。"""
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "shift_jis"):
        try:
            df = pd.read_csv(file_path, header=None, encoding=enc)
            return _df_to_markdown(df.head(MAX_ROWS))
        except Exception as e:  # noqa: BLE001
            last_err = e
    raise UnsupportedFileError(f"CSV 无法解码 {file_path}: {last_err}")


def excel_to_markdown(file_path: str) -> str:
    """对外公开：把 Excel/CSV 降维为 Markdown 纯文本（不调用 LLM）。"""
    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        md = xlsx_to_markdown(file_path)
    elif suffix in LEGACY_XLS_SUFFIXES:
        md = legacy_xls_to_markdown(file_path)
    elif suffix == ".csv":
        md = csv_to_markdown(file_path)
    else:
        raise UnsupportedFileError(f"不是表格文件: {file_path}")
    if not md.strip():
        raise UnsupportedFileError(f"文件无有效表格内容: {file_path}")
    if len(md) > MAX_MARKDOWN_CHARS:
        md = md[:MAX_MARKDOWN_CHARS] + "\n\n[... 内容过长已截断 ...]"
    return md


# ---------------------------------------------------------------------------
# 第二步：文本大模型语义提取
# ---------------------------------------------------------------------------

_NUM_RE = re.compile(r"^-?\d[\d,]*\.?\d*$")
_BARCODE_RE = re.compile(r"^\d{8,14}$")
# 条码标签前缀（如 "BARCODE:4936695359672"、"BARCODE  4936695359672"）
_BARCODE_LABEL_RE = re.compile(r"^\s*(?:BARCODE|BAR\s*CODE|BAR-NO|BAR\s*NO|条码)\s*:?\s*(\d{8,14})\s*$", re.IGNORECASE)
# 主行 / 合计行的关键词黑名单
_TOTAL_KEYWORDS = ("TOTAL", "合计", "小计", "SUBTOTAL", "SUB-TOTAL")


def _is_main_sku_row(row_values: list) -> bool:
    """判定一行是否为 SKU 主行：含 ≥1 个长文本 cell（≥6 且不含 TOTAL/合计/小计/SUBTOTAL）+ ≥2 个数值 cell。

    长文本定义：strip() 后长度 ≥ 6（避免被 `KGS`/`CTNS`/空品名等短串误判为品名）。
    数值定义：去掉 None/NaN 后，整型或浮点型 / 字符串形式的纯数字。
    """
    has_long_text = False
    num_count = 0
    for v in row_values:
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        if isinstance(v, str):
            s = v.strip()
            if not s:
                continue
            upper = s.upper()
            if any(kw.upper() in upper for kw in _TOTAL_KEYWORDS):
                # 含 TOTAL/合计/小计/SUBTOTAL 的 cell 直接否
                return False
            if len(s) >= 6:
                has_long_text = True
            # 短数值串（如 "120"、"3.5"）计入数值
            if _NUM_RE.match(s):
                num_count += 1
        elif isinstance(v, (int,)):
            # bool 是 int 的子类，要排除
            num_count += 1
        elif isinstance(v, float):
            num_count += 1
    return has_long_text and num_count >= 2


def _is_barcode_row(row_values: list) -> tuple[bool, str | None]:
    """判定一行是否为条码行；返回 (是否, 提取出的条码串)。

    优先级：
      1. cell 文本匹配 "BARCODE[: ](\\d{8,14})"
      2. cell 为纯 \\d{13}（13 位 EAN/JAN）独立成行
    """
    extracted: str | None = None
    for v in row_values:
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        if isinstance(v, str):
            s = v.strip()
            if not s:
                continue
            m = _BARCODE_LABEL_RE.match(s)
            if m:
                return True, m.group(1)
            if _BARCODE_RE.match(s):
                # 纯 13 位数字的 cell 暂存；最后统一看是不是"全行只有这一个数字 cell"
                extracted = extracted or s
    return (extracted is not None), extracted


def _bind_orphan_barcode(df: pd.DataFrame) -> pd.DataFrame:
    """单射形态条码回填（参见方案 §3.2 / 2026-08-11）。

    触发条件（全满足才启用）：
      1. 全表有且仅有 1 个 SKU 主行（含 ≥1 个长文本 cell + ≥2 个数值 cell）；
      2. 全表有且仅有 1 个条码行（BARCODE:xxxx 或纯 13 位数字），且不是主行；
      3. 条码行在主行下方（只向上绑定）。

    动作：把条码追加到主行末尾新 cell；条码行保留不删（宁冗勿漏）。

    不覆盖形态（不触发）：多主行 / 多条码 / 条码在主行上方 / 一条码多 SKU。
    """
    if df.empty or len(df) < 2:
        return df

    # 收集行级特征
    main_indices: list[int] = []
    barcode_indices: list[tuple[int, str]] = []  # (行索引, 条码数字)
    for idx, (_, row) in enumerate(df.iterrows()):
        vals = row.tolist()
        if _is_main_sku_row(vals):
            main_indices.append(idx)
        is_bc, bc = _is_barcode_row(vals)
        if is_bc and bc is not None:
            barcode_indices.append((idx, bc))

    if len(main_indices) != 1 or len(barcode_indices) != 1:
        return df
    main_idx = main_indices[0]
    bc_idx, bc_value = barcode_indices[0]
    if bc_idx == main_idx or bc_idx <= main_idx:
        # 条码行必须在主行下方；同一行 / 在主行上方 → 不触发
        return df

    # 复制避免原 DataFrame 被改写（pandas 链式赋值警告）
    df = df.copy()
    # 主行末尾追加一列：扩展到 df.shape[1]
    new_col_idx = df.shape[1]
    df.loc[main_idx, new_col_idx] = bc_value
    logger.info(
        "条码行回填 | sheet=%s 条码=%s 主行=%d 条码行=%d",
        "?", bc_value, main_idx, bc_idx,
    )
    return df


def _compress_adjacent_dup_strings(df: pd.DataFrame) -> pd.DataFrame:
    """A1 同行相邻同值长字符串压缩（参见方案 §3.1）。

    规则（仅对 DataFrame 行内字符串 cell）：
      - 只处理字符串 cell；数字（int / float）一律不动（N.W.=G.W. 是合法同值错位）
      - 只处理 strip() 后长度 ≥ 6 的串（KGS / CTNS / 净重 等短串保持列对齐）
      - 相邻同值 run：保留第一个，其余 cell 置空字符串

    DEBUG 日志：行内同值压缩 | sheet=%s 行=%d 值=%s 重复=%d
    """
    if df.empty:
        return df

    def _compress_row(row: pd.Series) -> pd.Series:
        prev_val: object = None
        prev_is_str = False
        prev_normalized: str = ""
        for col in df.columns:
            v = row[col]
            is_str = isinstance(v, str)
            if is_str:
                stripped = v.strip()
                if len(stripped) >= 6 and prev_is_str and stripped == prev_normalized:
                    # 同值 run 的非首 cell → 置空
                    row[col] = ""
                    logger.debug(
                        "行内同值压缩 | sheet=%s 行=%d 值=%s 重复=%d",
                        "?", int(row.name) if not isinstance(row.name, str) else row.name,
                        stripped, 2,
                    )
                    continue
                prev_is_str = True
                prev_normalized = stripped
            else:
                prev_is_str = False
                prev_normalized = ""
            prev_val = v
        return row

    return df.apply(_compress_row, axis=1)


def _drop_zero_rows(markdown_text: str) -> str:
    """纯 Python 预过滤全 0 占位行（不送 LLM）。

    判定：某 markdown 行含 8–14 位纯数字条码单元格，且该行其余所有数值
    单元格（QUANTITY/PACKAGES/净重/毛重/体积等）全为 0 → 该行为未发货占位行，
    删除不影响任何真实数据。无条码的行（表头、SUB TOTAL、说明文字）一律保留。
    （2026-07-27 正达案例：PL sheet 约 50 行 0 值占位行使 LLM 输出超 max_tokens
    被截断 / 生成超 120s 超时）
    """
    kept: list[str] = []
    dropped = 0
    lines = markdown_text.split("\n")
    for line in lines:
        if not line.startswith("|"):
            kept.append(line)
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        numerics = [c for c in cells if _NUM_RE.match(c)]
        has_barcode = any(_BARCODE_RE.match(c) for c in numerics)
        others = [c for c in numerics if not _BARCODE_RE.match(c)]
        if has_barcode and others and all(float(c.replace(",", "")) == 0 for c in others):
            dropped += 1
            continue  # 全 0 占位行，丢弃
        kept.append(line)
    if dropped:
        logger.info(
            "全 0 占位行预过滤 | 原行数=%d → 剩余行数=%d（过滤 %d 行）",
            len(lines), len(kept), dropped)
    else:
        logger.debug("全 0 占位行预过滤 | 原行数=%d，无需过滤", len(lines))
    return "\n".join(kept)


def sort_by_source_order(items: list, source_text: str) -> list:
    """按 sku_code 在源文本中首次出现的位置排序（稳定排序）。

    刚性保证输出顺序与单据行序一致（人工核对习惯）——LLM 的 JSON 数组
    顺序只是"恰好"按读入顺序，没有契约保证。找不到条码的项（sku_code
    为空或文本中不存在）保持原相对顺序排到末尾。
    （pdf_text_channel 亦复用本函数；vision 通道无文本层，无法定位，不排。）
    """
    def key(it) -> tuple:
        code = (getattr(it, "sku_code", None) or "").strip()
        pos = source_text.find(code) if code else -1
        return (pos < 0, pos)

    return sorted(items, key=key)


def extract_excel(file_path: str) -> ChannelResult:
    """提取单个 Excel/CSV 文件，返回结构化结果。"""
    result = ChannelResult()
    markdown_text = _drop_zero_rows(excel_to_markdown(file_path))  # 可能抛 UnsupportedFileError
    source_name = Path(file_path).name

    messages: list[dict] = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": build_text_user_prompt(markdown_text, source_name)},
    ]

    # JSON 解析失败最多重试 2 次（即最多 3 次调用）
    for attempt in range(3):
        result.json_attempts += 1
        raw = llm_client.extraction_chat_completion(
            messages, vision=False, source_file=source_name
        )
        try:
            payload = parse_payload(raw)
            for item in payload.items:
                item.source_file = file_path
            verified, notes = verify_weight_basis(payload.items, markdown_text)
            result.notes.extend(notes)
            result.items = apply_weight_basis(sort_by_source_order(verified, markdown_text))
            return result
        except JsonParseError as e:
            result.json_parse_failures += 1
            if attempt >= 2:
                # 重试耗尽：带堆栈记 ERROR 进 error.log，便于事后追查
                logger.exception(
                    "JSON 解析重试耗尽，最终失败 | %s | 共 %d 次尝试 | %s",
                    source_name, attempt + 1, str(e)[:300],
                )
                result.error = f"JSON 解析重试 2 次后仍失败: {e}"
                return result
            logger.warning(
                "JSON 解析失败，第 %d/2 次重试 | %s | 错误: %s",
                attempt + 1, source_name, str(e)[:200],
            )
            messages = messages + [
                {"role": "assistant", "content": raw},
                {"role": "user", "content": build_retry_message(raw, str(e))},
            ]
    return result
