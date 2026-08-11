"""Service 层：LangGraph 跑图逻辑集中在这里，路由层保持极薄。

【Celery 迁移预留接缝】
当前为同步实现（由路由层用 asyncio.to_thread 包裹调用，避免阻塞事件循环）。
日后迁移 Celery+Redis 时：
1. 把 run_until_interrupt() 整体下沉为 Celery task（worker 内直接调用即可，
   函数签名不变）；
2. 路由改为 task.delay(...) 返回 task_id，再加一个轮询接口查 AsyncResult；
3. resume_order() 因只做写 Excel+落库，耗时短，可保持同步接口不变。
"""
import contextlib
import json
import logging
import os
import sqlite3
import tempfile
import threading
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import ormsgpack
from langgraph.types import Command
from openpyxl import load_workbook
from sqlalchemy import select

from app.config import get_settings
from app.db.models import ReviewAudit
from app.db.session import get_session
from app.extraction.llm_client import usage_tracker
from app.extraction.session import SESSIONS_DIR
from app.graph import NODE2, NODE5, get_graph
from app.logging_config import logging_context

logger = logging.getLogger(__name__)


def _config(thread_id: str) -> dict:
    return {"configurable": {"thread_id": thread_id}}


def _safe_div(total, qty):
    """安全除法：返回 (结果, 公式字符串, 错误信息)。"""
    try:
        formula = f"{total} / {qty}"
        return total / qty, formula, None
    except (ZeroDivisionError, TypeError) as e:
        return None, f"{total} / {qty}", f"{type(e).__name__}: {e}"


# ---------------------------------------------------------------------------
# 节点级进度钩子（W4a）：graph.stream 的 updates 事件 → exec_progress 回调
# ---------------------------------------------------------------------------

def _make_progress_emitter(
    on_progress: Callable[[dict], None] | None,
    seed: dict[str, Any] | None = None,
) -> Callable[[dict], None]:
    """构造 stream 事件回调：从 updates 事件提取 node1/2/3/6 节点级进度。

    产出事件至少含 node/factory/done/total/message（type/tool/thread_id
    由调用方包装补充）。done/total 口径同 _summarize_snapshot：
    total = 装箱单工厂集 ∩ factory_filter；done = 已写回完成的工厂数
    （当前处理中的工厂不计入，node2 消息里以 done+1 作序号）。
    seed 用于 resume/rerun 场景预置跟踪器（resume 不再经过 node1，
    工厂全集/队列/当前工厂须从挂起现场取）。
    事件构造与回调整体 try/except：on_progress 抛异常绝不影响跑图。
    """
    from app.graph import NODE1, NODE2, NODE3, NODE4B, NODE6

    seed = seed or {}
    allow = set(seed["factory_filter"]) if seed.get("factory_filter") else None
    tracker: dict[str, Any] = {"total_set": set(), "pending": set(), "current": None}
    if seed.get("downstream_requirements"):
        total_set = set((seed.get("downstream_requirements") or {}).keys())
        if allow:
            total_set &= allow
        tracker["total_set"] = total_set
    tracker["pending"] = set(seed.get("pending_factories") or [])
    tracker["current"] = (seed.get("current_factory_data") or {}).get("factory_name")

    def emit(event: dict) -> None:
        if on_progress is None:
            return
        try:
            for node, value in event.items():
                if not isinstance(value, dict):
                    continue
                # 先更新跟踪器（updates 事件值即节点返回的 state 增量）
                if "downstream_requirements" in value:
                    total_set = set((value.get("downstream_requirements") or {}).keys())
                    if allow:
                        total_set &= allow
                    tracker["total_set"] = total_set
                if "pending_factories" in value:
                    tracker["pending"] = set(value.get("pending_factories") or [])
                current = (value.get("current_factory_data") or {}).get("factory_name")
                if current:
                    tracker["current"] = current

                total = len(tracker["total_set"])
                popped = len(tracker["total_set"] - tracker["pending"])
                factory = tracker["current"]
                if node == NODE1:
                    progress = {"node": node, "factory": None, "done": 0,
                                "total": total,
                                "message": f"装箱单解析完成，共 {total} 个工厂"}
                elif node == NODE2 and factory:
                    cur_data = value.get("current_factory_data") or {}
                    if cur_data.get("is_final_attempt"):
                        # W6a 暂缓二遍重试：不带序号（此时 pending 集合不变，
                        # done+1 序号口径会失真）
                        done = max(popped - 1, 0)
                        progress = {"node": node, "factory": factory, "done": done,
                                    "total": total,
                                    "message": f"开始重试 {factory}（暂缓重试）"}
                    else:
                        # 当前工厂刚出队列、在处理中，不计入 done；序号 = done+1
                        done = max(popped - 1, 0)
                        progress = {"node": node, "factory": factory, "done": done,
                                    "total": total,
                                    "message": f"开始处理 {factory}（{done + 1}/{total}）"}
                elif node == NODE3 and factory:
                    done = max(popped - 1, 0)
                    progress = {"node": node, "factory": factory, "done": done,
                                "total": total, "message": f"{factory} 提取完成"}
                elif node == NODE4B and factory:
                    # W6a 暂缓：done 不计、total 不变（暂缓厂未写回不算完成）
                    done = max(popped - 1, 0)
                    progress = {"node": node, "factory": factory, "done": done,
                                "total": total,
                                "message": f"{factory} 提取失败，已暂缓，将在其余工厂处理完后重试"}
                elif node == NODE6 and factory:
                    # 写回完成：当前工厂计入 done
                    progress = {"node": node, "factory": factory, "done": popped,
                                "total": total, "message": f"{factory} 写回完成"}
                else:
                    continue
                on_progress(progress)
        except Exception:  # noqa: BLE001 进度是辅助设施，抛异常绝不影响跑图
            logger.debug("on_progress 回调异常已忽略", exc_info=True)

    return emit


# ---------------------------------------------------------------------------
# 后台预提取（2026-07-28）：审核不阻塞提取
# ---------------------------------------------------------------------------
# 图首次 interrupt（工厂 A 挂起等审核）时，后台线程开始对剩余工厂逐个预提取。
# 每个工厂走：Node2 匹配文件夹 → target_identifier 识别正确单据 → Node3 调 LLM
# → 结果原子写入 sessions/{工厂}.json。图内 Node3 发现缓存命中时跳过 LLM。
# 后台线程串行（一次一个工厂），不并发打 API——只一个人用，不值得为并行 LLM
# 增加复杂度。提取并发只解决"审核时别闲着"的问题。
#
# 线程安全：预提取结果经原子 rename（tmp → target）落盘，图内 Node3 读缓存
# 时要么看到旧文件、要么看到完整新文件，不会读到半截 JSON。最坏情况是缓存
# 未命中（后台还没跑完），此时图 Node3 走正常提取流程，无数据丢失。

_PRE_EXTRACT_LOCK = threading.Lock()  # 保证同一批次只有一个预提取线程
_known_running: set[str] = set()      # 进程级已知在跑的批次


# ---------------------------------------------------------------------------
# 预提取进度落盘（2026-08-03）：对话页批次实时进度条的数据源
# ---------------------------------------------------------------------------
# 后台预提取线程每次状态变化即原子重写进度 JSON（tmp → rename，与 session
# 原子写模式一致），UI 批次端点原样带出给对话页轮询渲染——纯确定性数据，
# 不过 LLM。进度是辅助设施：写文件失败只记日志，绝不影响预提取本身。


def _write_batch_config(batch_id: str, config: dict[str, Any]) -> None:
    """写入批次配置文件 output/{batch_id}/batch_config.json。

    配置内容：thread_id、路径、创建时间、最后运行时间、运行次数。
    写文件失败只记日志，绝不影响主流程。
    """
    try:
        settings = get_settings()
        batch_dir = settings.batch_output_dir(batch_id)
        batch_dir.mkdir(parents=True, exist_ok=True)
        config_path = batch_dir / "batch_config.json"
        config_path.write_text(
            json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:  # noqa: BLE001
        logger.exception("[batch_config] 写入批次配置文件失败 batch_id=%s", batch_id)


def _read_batch_config(batch_id: str) -> dict[str, Any] | None:
    """读取批次配置文件，不存在或损坏时返回 None。"""
    try:
        settings = get_settings()
        config_path = settings.batch_output_dir(batch_id) / "batch_config.json"
        if not config_path.exists():
            return None
        return json.loads(config_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def update_batch_paths(
    batch_id: str,
    downstream_file_path: str | None = None,
    upstream_root: str | None = None,
    reset_checkpoint: bool = False,
) -> dict[str, Any]:
    """更新批次路径配置，可选重置 checkpoint。

    Args:
        batch_id: 批次号
        downstream_file_path: 新装箱单路径（None=不改）
        upstream_root: 新工厂文件夹路径（None=不改）
        reset_checkpoint: 是否重置 checkpoint（下次执行用新路径）

    Returns:
        {"updated": True/False, "config": {...}, "checkpoint_reset": True/False}

    Raises:
        ValueError: 批次不存在或路径无效
    """
    config = _read_batch_config(batch_id)
    if config is None:
        raise ValueError(f"批次不存在: {batch_id}")

    # 更新路径
    if downstream_file_path:
        d_path = Path(downstream_file_path).expanduser()
        if not d_path.is_file():
            raise ValueError(f"下游装箱单路径不存在或不是文件: {downstream_file_path}")
        config["downstream_file_path"] = str(d_path)
    if upstream_root:
        u_path = Path(upstream_root).expanduser()
        if not u_path.is_dir():
            raise ValueError(f"上游工厂文件夹路径不存在或不是目录: {upstream_root}")
        config["upstream_root"] = str(u_path)

    _write_batch_config(batch_id, config)

    checkpoint_reset = False
    if reset_checkpoint:
        from langgraph.graph import START
        graph = get_graph()
        cfg = _config(batch_id)
        update = {}
        if downstream_file_path:
            update["downstream_file_path"] = config["downstream_file_path"]
        if upstream_root:
            update["upstream_root"] = config["upstream_root"]
        graph.update_state(cfg, update, as_node=START)
        checkpoint_reset = True

    return {
        "updated": True,
        "config": config,
        "checkpoint_reset": checkpoint_reset,
    }


def rerun_batch(
    batch_id: str,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """完全重跑批次：先归档当前 output 产物到 _history/，再清空 sessions，再重置 checkpoint 重跑。

    重跑前把 output/{safe(batch_id)}/ 目录整体（containers/、declarations/、
    batch_state.json、batch_config.json）+ 顶层 {safe(batch_id)}_filled.xlsx
    一起搬到 output/_history/{safe(batch_id)}/r{N}_{ts}/，按 run_count 编号。

    然后清空 data/sessions/{safe(batch_id)}/*.json（不删目录）。

    这样每批次独立缓存；rerun 不污染历史产物，所有历史 run 可审计回看。
    """
    import shutil

    # 注：settings.history_output_dir 方法由 config.py 提供（如果还没有，
    # 请在 Settings 类加一个返回 output/_history/{safe(batch_id)} 的方法）。
    config = _read_batch_config(batch_id)
    if config is None:
        raise ValueError(f"批次不存在: {batch_id}")
    run_count = config.get("run_count", 1)

    settings = get_settings()
    batch_output_dir = settings.batch_output_dir(batch_id)

    # ① 归档 output（仅当上一轮产物存在）
    if batch_output_dir.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_run_dir = settings.history_output_dir(batch_id) / f"r{run_count}_{ts}"
        archive_run_dir.mkdir(parents=True, exist_ok=True)
        try:
            shutil.move(str(batch_output_dir), str(archive_run_dir / "run"))
            logger.info("[rerun] 已归档 output 到 %s", archive_run_dir)
        except (OSError, shutil.Error) as e:
            logger.error("[rerun] 归档失败，拒绝重跑: %s: %s", type(e).__name__, e)
            raise RuntimeError(f"output 归档失败，已拒绝 rerun: {e}") from e

    # ② 清空 sessions 目录下文件（不删目录）
    sessions_dir = SESSIONS_DIR / settings.safe_path_tag(batch_id)
    if sessions_dir.is_dir():
        for p in sessions_dir.glob("*.json"):
            try:
                p.unlink()
            except OSError as e:
                logger.warning("[rerun] 清理 session %s 失败（跳过）: %s: %s",
                               p.name, type(e).__name__, e)
        logger.info("[rerun] 已清空 sessions 目录: %s", sessions_dir)

    # ③ 原逻辑：重置 checkpoint、run_count + 1、_write_batch_config、调 run_until_interrupt
    from langgraph.graph import START
    graph = get_graph()
    cfg = _config(batch_id)
    graph.update_state(
        cfg,
        {
            "downstream_file_path": config["downstream_file_path"],
            "upstream_root": config["upstream_root"],
            "batch_id": batch_id,
        },
        as_node=START,
    )

    config["last_run_at"] = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    config["run_count"] = run_count + 1
    _write_batch_config(batch_id, config)

    return run_until_interrupt(
        thread_id=batch_id,
        downstream_file_path=config["downstream_file_path"],
        upstream_root=config["upstream_root"],
        on_progress=on_progress,
    )


def add_factories_to_batch(
    batch_id: str,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """补充工厂：重新解析装箱单，增量合并 pending_factories，继续执行。

    Args:
        batch_id: 批次号
        on_progress: 进度回调

    Returns:
        {"added": int, "factories": [...], "status": "..."}

    Raises:
        ValueError: 批次不存在或无新工厂
    """
    from app.nodes.parse_downstream import parse_downstream_file

    config = _read_batch_config(batch_id)
    if config is None:
        raise ValueError(f"批次不存在: {batch_id}")

    # 解析装箱单
    new_requirements, new_row_map = parse_downstream_file(config["downstream_file_path"])

    # 读取当前 checkpoint state
    graph = get_graph()
    cfg = _config(batch_id)
    snap = graph.get_state(cfg)
    if not snap.values:
        raise ValueError(f"批次 {batch_id} 无 checkpoint state")

    existing_requirements = snap.values.get("downstream_requirements", {})
    existing_row_map = snap.values.get("downstream_row_map", {})

    # 找新工厂
    new_factories = set(new_requirements.keys()) - set(existing_requirements.keys())
    if not new_factories:
        return {"added": 0, "factories": [], "message": "没有新工厂"}

    # 合并
    merged_requirements = {**existing_requirements, **{k: new_requirements[k] for k in new_factories}}
    merged_row_map = {**existing_row_map, **{k: new_row_map[k] for k in new_factories}}
    merged_pending = list(snap.values.get("pending_factories", [])) + list(new_factories)

    # 更新 state
    graph.update_state(
        cfg,
        {
            "downstream_requirements": merged_requirements,
            "downstream_row_map": merged_row_map,
            "pending_factories": merged_pending,
        },
        as_node=NODE2,
    )

    # 继续执行
    for event in graph.stream(None, cfg, stream_mode="updates"):
        pass  # 进度回调可选

    final = graph.get_state(cfg)
    return {
        "added": len(new_factories),
        "factories": sorted(new_factories),
        "status": final.values.get("validation_status", "unknown"),
    }


def _write_batch_state(
    batch_id: str,
    status: str,
    state: dict[str, Any] | None = None,
    error: str | None = None,
) -> None:
    """写入批次状态文件 output/{batch_id}/batch_state.json（关键状态转换时调用）。

    写入时机：批次启动（running）、Node5 挂起（pending_review）、
    Node7 完成（completed）、异常（error）。
    写文件失败只记日志，绝不影响跑图。
    """
    try:
        settings = get_settings()
        batch_dir = settings.batch_output_dir(batch_id)
        batch_dir.mkdir(parents=True, exist_ok=True)
        state_path = batch_dir / "batch_state.json"

        # 读取已有状态（合并 timestamps）
        existing: dict[str, Any] = {}
        if state_path.exists():
            try:
                existing = json.loads(state_path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                pass

        now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()

        new_state: dict[str, Any] = {
            "batch_id": batch_id,
            "status": status,
            "updated_at": now,
        }

        # timestamps：started_at 仅在首次写入或 running 时设置
        if "started_at" not in existing or status == "running":
            new_state["started_at"] = now
        else:
            new_state["started_at"] = existing.get("started_at")
        new_state["completed_at"] = now if status == "completed" else existing.get("completed_at")

        # 工厂信息（从 state 提取）
        if state:
            total_set = set((state.get("downstream_requirements") or {}).keys())
            pending_set = set(state.get("pending_factories") or [])
            new_state["factories"] = {
                "total": sorted(total_set),
                "done": sorted(total_set - pending_set),
                "pending": sorted(pending_set),
                "current": (state.get("current_factory_data") or {}).get("factory_name"),
                "deferred": [
                    d.get("factory_name", "") for d in (state.get("deferred_factories") or [])
                ],
            }
            # 输出路径
            output: dict[str, str] = {}
            if state.get("final_output_path"):
                output["containers"] = state["final_output_path"]
            output["declarations_dir"] = str(settings.batch_declarations_dir(batch_id))
            new_state["output"] = output
        else:
            new_state["factories"] = existing.get("factories", {})
            new_state["output"] = existing.get("output", {})

        if error:
            new_state["error"] = error

        state_path.write_text(
            json.dumps(new_state, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:  # noqa: BLE001 状态文件是辅助设施，写失败不影响主流程
        logger.exception("[batch_state] 写入批次状态文件失败 batch_id=%s", batch_id)


def _preextract_progress_path(thread_id: str) -> Path:
    """预提取进度文件路径：SESSIONS_DIR/_preextract_progress_{安全tag}.json。"""
    return SESSIONS_DIR / f"_preextract_progress_{get_settings().safe_path_tag(thread_id)}.json"


class _PreExtractProgress:
    """预提取进度落盘器：线程内维护状态，每次变化原子重写 JSON。

    文件结构：
      {"thread_id": ..., "updated_at": ISO时间,
       "factories": [{"factory": ..., "status": "pending|running|cached|done|failed",
                      "error": str|null, "ts": ISO时间}, ...]}
    """

    def __init__(self, thread_id: str, factories: list[str]) -> None:
        self._path = _preextract_progress_path(thread_id)
        self._state: dict[str, Any] = {
            "thread_id": thread_id,
            "updated_at": self._now(),
            "factories": [
                {"factory": f, "status": "pending", "error": None, "ts": None}
                for f in factories
            ],
        }
        self._by_name = {f["factory"]: f for f in self._state["factories"]}
        self._flush()

    @staticmethod
    def _now() -> str:
        return datetime.now().astimezone().isoformat(timespec="seconds")

    def update(self, factory: str, status: str, error: str | None = None) -> None:
        """状态流转：pending → running → cached/done/failed；每次变化即重写。"""
        entry = self._by_name.get(factory)
        if entry is None:
            return
        entry["status"] = status
        entry["error"] = error
        entry["ts"] = self._now()
        self._flush()

    def _flush(self) -> None:
        """原子重写进度文件（tmp → rename）；失败只记日志，绝不抛出。"""
        try:
            self._state["updated_at"] = self._now()
            SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
            tmp_fd, tmp_path = tempfile.mkstemp(
                suffix=".tmp", prefix=self._path.stem + ".", dir=str(SESSIONS_DIR))
            try:
                os.write(tmp_fd, json.dumps(self._state, ensure_ascii=False,
                                            indent=1).encode("utf-8"))
            finally:
                os.close(tmp_fd)
            os.replace(tmp_path, self._path)  # POSIX 原子 rename
        except OSError as e:
            logger.warning("[预提取] 进度文件写入失败（不影响预提取）：%s: %s",
                           type(e).__name__, e)


def load_pre_extraction_progress(thread_id: str) -> dict[str, Any] | None:
    """读预提取进度文件（批次端点用）：不存在/损坏返回 None（静默，绝不 500）。"""
    path = _preextract_progress_path(thread_id)
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001 损坏的进度文件按无进度处理
        return None
    if not isinstance(data, dict) or not isinstance(data.get("factories"), list):
        return None
    return data


def _start_pre_extraction(thread_id: str) -> None:
    """为当前批次启动后台预提取 daemon 线程（幂等：同一批次只跑一个）。"""
    graph = get_graph()
    cfg = _config(thread_id)
    snap = graph.get_state(cfg)
    values = snap.values or {}
    pending = list(values.get("pending_factories") or [])
    if not pending:
        return

    upstream_root = values.get("upstream_root") or get_settings().upstream_root
    requirements = values.get("downstream_requirements") or {}

    if thread_id in _known_running:
        return
    _known_running.add(thread_id)

    def _pre_extract():
        try:
            _pre_extract_factories(pending, upstream_root, requirements,
                                   thread_id=thread_id)
        finally:
            _known_running.discard(thread_id)

    t = threading.Thread(target=_pre_extract, daemon=True,
                         name=f"pre-extract-{thread_id}")
    t.start()
    logger.info("[预提取] 批次 %s：后台线程已启动，%d 个工厂待预提取",
                thread_id, len(pending))


def _pre_extract_factories(
    factories: list[str],
    upstream_root: str,
    requirements: dict[str, list[str]],
    thread_id: str | None = None,
) -> None:
    """后台线程主体：逐个工厂串行提取，结果存入 session JSON。

    thread_id 给出时同步落盘预提取进度（对话页批次进度条数据源）：
    启动全部 pending → 每个工厂 running → 缓存跳过 cached / 成功 done /
    异常 failed（error 记异常类型+摘要，截 200 字符）。
    """
    from app.factory_match import load_alias_map, match_factory_folder
    from app.nodes.extraction_node import _run_factory_session, _try_load_cached_session

    root_path = Path(upstream_root).expanduser()
    if not root_path.is_dir():
        logger.warning("[预提取] 上游目录 %s 不存在，跳过", upstream_root)
        return

    folders = [d.name for d in root_path.iterdir() if d.is_dir()]
    alias_map = load_alias_map()
    cutoff = get_settings().fuzzy_match_score_cutoff
    progress = _PreExtractProgress(thread_id, factories) if thread_id else None

    for factory in factories:
        # 缓存已存在且新鲜（路径证据仍在当前批次 upstream_root 之下）则跳过；
        # 新鲜度校验是 rerun_with_paths 改路径重跑场景的兜底：旧根目录留下
        # 的缓存视为陈旧，照常重提，不误用上一批次数据
        if _try_load_cached_session(thread_id, factory, str(root_path)) is not None:
            logger.info("[预提取] 工厂「%s」：缓存已存在，跳过", factory)
            if progress:
                progress.update(factory, "cached")
            continue

        # Node2 逻辑：匹配文件夹
        folder_name, score, method = match_factory_folder(
            factory, folders, alias_map, cutoff=cutoff)
        if not folder_name:
            logger.warning("[预提取] 工厂「%s」：未匹配到文件夹，跳过", factory)
            if progress:
                progress.update(factory, "failed", "未匹配到工厂文件夹")
            continue
        folder_path = str(root_path / folder_name)
        expected_skus = requirements.get(factory, [])

        if progress:
            progress.update(factory, "running")
        try:
            _run_factory_session(thread_id, folder_path, factory, expected_skus)
            logger.info("[预提取] 工厂「%s」：完成（%s，得分 %.1f）",
                        factory, method, score)
            if progress:
                progress.update(factory, "done")
        except Exception as e:
            logger.exception("[预提取] 工厂「%s」：异常 %s，跳过", factory, e)
            if progress:
                progress.update(factory, "failed",
                                f"{type(e).__name__}: {e}"[:200])


# ---------------------------------------------------------------------------
# 每批次独立 session 缓存目录（2026-08-10）
# ---------------------------------------------------------------------------
# 背景：早期实现把 sessions/*.json 全部平铺在 SESSIONS_DIR 顶层，按工厂
# 名命名、无批次维度。新批次启动时 Node3 缓存命中和后台预提取会沿用
# 上一批次的结果，导致错配。新设计：每批次独立目录
# data/sessions/{safe(batch_id)}/{factory}.json，跨批次物理隔离、零冲突。
# 新批次启动时幂等创建目录，目录已存在则复用——把里面的 *.json 留作下次
# run 增量（rerun 显式清空是另一条路径，由 rerun_batch 负责）。


def _ensure_batch_session_dir(thread_id: str) -> Path:
    """批次启动时确保 data/sessions/{safe(batch_id)}/ 目录存在。

    每批次独立缓存目录。新批次启动会幂等创建，不做归档。
    目录不存在则 mkdir，目录已存在则直接复用，文件留作下次 run 增量。
    返回该目录路径供调用方使用。
    """
    settings = get_settings()
    batch_dir = SESSIONS_DIR / settings.safe_path_tag(thread_id)
    batch_dir.mkdir(parents=True, exist_ok=True)
    return batch_dir


def run_until_interrupt(
    thread_id: str,
    downstream_file_path: str | None = None,
    upstream_root: str | None = None,
    factory_filter: list[str] | None = None,
    factory_alias_overrides: dict[str, str] | None = None,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """启动流程，执行 Node1-4 直到 Node5 的 interrupt 挂起（或全程无挂起跑完）。

    factory_alias_overrides：批次级「仅本次生效」工厂名对照（W5），
    非空才写进 initial_state（Node2 匹配的最高优先档，不落盘）。
    on_progress：节点级进度回调（W4a），收到至少含
    node/factory/done/total/message 的事件 dict；抛异常不影响跑图。

    返回：
      - {"status": "pending_human_review", "review_data": payload}
      - {"status": "completed", "final_state": ...}
    """
    # L2 日志关联：绑定批次号，图内节点日志（context 拷贝传播）自动携带；
    # logging_context 退出时恢复前值（嵌套调用安全，不抹外层绑定）
    with logging_context(thread_id=thread_id):
        graph = get_graph()
        initial_state: dict[str, Any] = {}
        if downstream_file_path:
            initial_state["downstream_file_path"] = downstream_file_path
        if upstream_root:
            initial_state["upstream_root"] = upstream_root
        if factory_filter:
            initial_state["factory_filter"] = factory_filter
        if factory_alias_overrides:
            initial_state["factory_alias_overrides"] = factory_alias_overrides
        initial_state["batch_id"] = thread_id
        _ensure_batch_session_dir(thread_id)
        snap = graph.get_state(_config(thread_id))
        _write_batch_state(thread_id, "running", state=initial_state)

        emit = _make_progress_emitter(on_progress, seed=initial_state)
        try:
            for event in graph.stream(initial_state, _config(thread_id), stream_mode="updates"):
                emit(event)
                if "__interrupt__" in event:
                    payload = event["__interrupt__"][0].value
                    # 首次 interrupt 后启动后台预提取：
                    # 工厂 A 挂起等审核 → 后台线程开始提取工厂 B、C、D…
                    _start_pre_extraction(thread_id)
                    # 从 checkpoint 读取当前 state 写入批次状态
                    current = graph.get_state(_config(thread_id))
                    _write_batch_state(thread_id, "pending_review", state=current.values)
                    return {
                        "status": "pending_human_review",
                        "thread_id": thread_id,
                        "review_data": payload,
                    }
        except Exception as e:  # noqa: BLE001
            _write_batch_state(thread_id, "error", error=str(e))
            raise

        final = graph.get_state(_config(thread_id))
        _write_batch_state(thread_id, "completed", state=final.values)
        return {
            "status": "completed",
            "thread_id": thread_id,
            "final_output_path": final.values.get("final_output_path"),
        }


def resume_order(thread_id: str, resume_data: dict,
                 on_progress: Callable[[dict], None] | None = None) -> dict[str, Any]:
    """用人工反馈数据唤醒挂起的图，继续执行 Node6/7（写 Excel + 落库）。

    resume_data 结构见 nodes/human_review.py 的 docstring。
    on_progress：节点级进度回调（W4a），跟踪器以挂起现场为种子
    （resume 不再经过 node1，工厂全集/队列从 state 预置）。

    审计：stream 前 _prepare_audit 抓取原始 payload 做 diff 快照，
    返回前 _write_audit 落 review_audits（失败只警告，绝不阻塞 resume）。
    """
    graph = get_graph()
    state = graph.get_state(_config(thread_id))
    if not state.next:
        raise ValueError("该任务没有处于等待审核状态，或已完成。")

    # L2 日志关联：resume 与首次运行是不同的请求/context，需重新绑定；
    # 工厂名从挂起现场取（首次运行的绑定已随 logging_context 退出恢复），
    # 保证审核后续节点（Node5 收尾/Node6 写回）日志仍带工厂名
    factory = (state.values.get("current_factory_data") or {}).get("factory_name")
    with logging_context(thread_id=thread_id, factory=factory):
        prepared = _prepare_audit(thread_id, resume_data)

        # 恢复执行，直到下一个 interrupt（多工厂循环时）或 END
        emit = _make_progress_emitter(on_progress, seed=state.values)
        try:
            for event in graph.stream(
                Command(resume=resume_data), _config(thread_id), stream_mode="updates"
            ):
                emit(event)
                if "__interrupt__" in event:
                    payload = event["__interrupt__"][0].value
                    # 每次 interrupt 后继续后台预提取剩余工厂
                    # （工厂 B 审核中 → 后台提取 C、D…）
                    _start_pre_extraction(thread_id)
                    current = graph.get_state(_config(thread_id))
                    _write_batch_state(thread_id, "pending_review", state=current.values)
                    result = {
                        "status": "pending_human_review",
                        "thread_id": thread_id,
                        "review_data": payload,
                    }
                    _write_audit(prepared, result.get("status"))
                    return result
        except Exception as e:  # noqa: BLE001
            _write_batch_state(thread_id, "error", error=str(e))
            raise

        final = graph.get_state(_config(thread_id))
        _write_batch_state(thread_id, "completed", state=final.values)
        result = {
            "status": "success",
            "message": "数据已成功落库并写入下游表格",
            "final_validation_status": final.values.get("validation_status"),
            "final_output_path": final.values.get("final_output_path"),
        }
        _write_audit(prepared, result.get("status"))
        return result


def rerun_with_paths(
    thread_id: str,
    upstream_root: str | None = None,
    downstream_file_path: str | None = None,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """对话改路径后的当前批次重跑：带新路径从 Node1 重新执行到 Node5 挂起。

    机制（langgraph 1.2.9 实测）：挂起线程有未完成的 interrupt 任务，
    Command(goto=...) 输入会被旧任务卡死；必须先 update_state(as_node=START)
    写入新路径并作废旧现场（下一个 checkpoint 从 Node1 重新开始），
    再 invoke(None) 触发 Node1→Node5 全链重跑，Node5 产生新 interrupt payload。

    仅当 thread 处于挂起等待状态时允许重跑；已完成的批次抛错
    （路径修改已写 .env，对后续批次生效）。
    """
    from langgraph.graph import START

    graph = get_graph()
    cfg = _config(thread_id)
    snap = graph.get_state(cfg)
    if not snap.values:
        raise ValueError(f"thread {thread_id} 不存在")
    if not snap.next:
        raise ValueError(
            f"thread {thread_id} 已完成，无法重跑（路径修改已对后续批次生效）"
        )

    update: dict[str, Any] = {}
    if upstream_root:
        update["upstream_root"] = upstream_root
    if downstream_file_path:
        update["downstream_file_path"] = downstream_file_path

    # L2 日志关联：重跑也是一次完整跑图，绑定批次号（工厂名由 Node2 重绑）
    with logging_context(thread_id=thread_id):
        graph.update_state(cfg, update, as_node=START)
        emit = _make_progress_emitter(on_progress, seed=snap.values)
        for event in graph.stream(None, cfg, stream_mode="updates"):
            emit(event)
            if "__interrupt__" in event:
                _start_pre_extraction(thread_id)
                return {
                    "status": "pending_human_review",
                    "thread_id": thread_id,
                    "review_data": event["__interrupt__"][0].value,
                }

        final = graph.get_state(cfg)
        return {
            "status": "completed",
            "thread_id": thread_id,
            "final_output_path": final.values.get("final_output_path"),
        }


def retry_factory_extraction(
    thread_id: str,
    folder: str | None = None,
    save: bool = False,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """单厂重试：只重跑当前挂起工厂的 Node3→Node5，已审核工厂不受影响。

    机制同 rerun_with_paths（langgraph 1.2.9 坑）：update_state(as_node=NODE2)
    作废 Node5 interrupt 现场，invoke 从 Node3 继续。Node2 仅作锚点不会重跑，
    current_factory_data 仍在 state 中供 Node3 使用。

    对照注入（W6b）：
    - folder 非 None 时，先经 factory_match.validate_subfolder 校验
      （upstream_root 下现存一级子目录，ValueError 直接上抛），
      再写入 current_factory_data.folder_path 与批次级
      factory_alias_overrides[factory]=folder——仅本次批次生效，不落盘。
    - save=True 且 folder 非 None：save_alias_entries 永久落 alias_map.json，
      后续批次自动生效；返回 dict 带 "alias_saved"。save 异常不阻断重试
      主流程——捕获记 warning，返回里带 "alias_save_error"。
    - save=True 但 folder 为 None：ValueError（无对照可保存，防误用）。
    """
    graph = get_graph()
    cfg = _config(thread_id)
    snap = graph.get_state(cfg)
    if not snap.values:
        raise ValueError(f"thread {thread_id} 不存在")
    if NODE5 not in (snap.next or ()):
        raise ValueError(f"thread {thread_id} 当前未挂起待审核，无法单厂重试")

    cur = snap.values.get("current_factory_data") or {}
    factory = cur.get("factory_name")
    if not factory:
        raise ValueError(f"thread {thread_id} 缺少当前工厂数据，无法单厂重试")

    if save and folder is None:
        raise ValueError("save=true 必须同时提供 folder（无对照可保存）")

    # 对照注入 / 自动重匹配：先尝试确定 folder_path
    validated_path: Path | None = None
    matched_folder_name: str | None = None

    if folder is not None:
        from app.factory_match import validate_subfolder
        upstream_root = (
            snap.values.get("upstream_root") or get_settings().upstream_root)
        validated_path = validate_subfolder(upstream_root, folder)
        matched_folder_name = folder
    elif not cur.get("folder_path"):
        # 该工厂首次 folder_router 未匹配到文件夹（folder_path 为空），
        # retry 时自动重跑一次文件夹匹配，避免 update_state(as_node=NODE2)
        # 跳过 Node2 后永远 no_folder_matched。
        from app.factory_match import load_alias_map, match_factory_folder
        upstream_root = (
            snap.values.get("upstream_root") or get_settings().upstream_root)
        root_path = Path(upstream_root).expanduser()
        if root_path.is_dir():
            folders = [d.name for d in root_path.iterdir() if d.is_dir()]
            alias_map = load_alias_map()
            cutoff = get_settings().fuzzy_match_score_cutoff
            folder_name, score, method = match_factory_folder(
                factory, folders, alias_map, cutoff=cutoff)
            if folder_name:
                validated_path = root_path / folder_name
                matched_folder_name = folder_name
                logger.info(
                    "retry 自动重匹配到文件夹: %s -> %s (%s, %.1f)",
                    factory, folder_name, method, score)

    # L2 日志关联：单厂重试也是一次跑图，绑定批次号（工厂名由 Node3 重绑）
    with logging_context(thread_id=thread_id):
        # 置强制重提标志：Node3 跳过会话缓存重新走提取流程；
        # Node3 所有返回分支自清 force_reextract=False，不会残留到后续工厂。
        # 不清 extracted_items——Node3 各分支都会覆写 cur["extracted_items"]
        update: dict[str, Any] = {"force_reextract": True}
        if validated_path is not None:
            # 对照注入：folder_path 直接指向校验/重匹配后的目录；
            # 批次级 factory_alias_overrides 同步写入，仅本次生效不落盘
            cur2 = dict(cur)
            cur2["folder_path"] = str(validated_path)
            update["current_factory_data"] = cur2
            overrides = dict(snap.values.get("factory_alias_overrides") or {})
            overrides[factory] = matched_folder_name
            update["factory_alias_overrides"] = overrides
        graph.update_state(cfg, update, as_node=NODE2)

        alias_saved: dict | None = None
        alias_save_error: str | None = None
        if save and validated_path is not None:
            from app.factory_match import save_alias_entries
            try:
                # 永久对照落盘（.bak 备份 + 原子写），后续批次自动生效
                alias_saved = save_alias_entries({factory: folder})
            except Exception as e:  # noqa: BLE001 落盘失败不阻断重试主流程
                logger.warning(
                    "对照永久保存失败（重试仍继续）: %s -> %s (%s)",
                    factory, folder, e)
                alias_save_error = f"{type(e).__name__}: {e}"

        emit = _make_progress_emitter(on_progress, seed=snap.values)
        for event in graph.stream(None, cfg, stream_mode="updates"):
            emit(event)
            if "__interrupt__" in event:
                _start_pre_extraction(thread_id)
                result: dict[str, Any] = {
                    "status": "pending_human_review",
                    "thread_id": thread_id,
                    "factory": factory,
                    "review_data": event["__interrupt__"][0].value,
                }
                if alias_saved is not None:
                    result["alias_saved"] = alias_saved
                if alias_save_error is not None:
                    result["alias_save_error"] = alias_save_error
                return result

        final = graph.get_state(cfg)
        result = {
            "status": "completed",
            "thread_id": thread_id,
            "factory": factory,
            "final_output_path": final.values.get("final_output_path"),
        }
        if alias_saved is not None:
            result["alias_saved"] = alias_saved
        if alias_save_error is not None:
            result["alias_save_error"] = alias_save_error
        return result


def force_extract_file(
    thread_id: str,
    file_path: str,
    pages: list[int] | None = None,
    force_vision: bool = False,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """指定单个文件强制重新提取（跳过自动目标识别，可指定页码/强制视觉）。

    适用场景：
    - 自动识别选错了箱单（比如挑了报关汇总版）
    - 箱单是扫描件/图片，自动识别扫不出
    - 一份 PDF 混了多份单据，操作员只要其中几页
    - 文本层识别把数字读错，要换看图方式重新识别

    机制：
    1. 校验批次挂起（NODE5 in snap.next）
    2. 校验 file_path 在批次 upstream_root 之下（防路径穿越）
    3. 校验 pages（每个 1..≤500、去重升序、长度 ≤ 12）
    4. 从批次专属 session 目录加载 FactorySession
    5. 调 session.force_extract(session, file_path, pages=pages)
    6. 原子写回 session JSON（与 Node3 写策略一致）
    7. graph.update_state(as_node=NODE2) 不带 force_reextract=True
       ——要 Node3 命中缓存、读到刚写入的 force_extract 结果
    8. graph.stream 直到 __interrupt__，返回新 review_data
    """
    from app.extraction.session import (
        FactorySession,
        batch_session_dir,
        batch_session_path,
        force_extract as session_force_extract,
    )

    graph = get_graph()
    cfg = _config(thread_id)
    snap = graph.get_state(cfg)
    if not snap.values:
        raise ValueError(f"thread {thread_id} 不存在")
    if NODE5 not in (snap.next or ()):
        raise ValueError(f"thread {thread_id} 当前未挂起待审核，无法指定文件提取")

    cur = snap.values.get("current_factory_data") or {}
    factory = cur.get("factory_name")
    if not factory:
        raise ValueError(f"thread {thread_id} 缺少当前工厂数据，无法指定文件提取")

    # 页码 sanity：1-based、去重升序、范围 1..500、长度 ≤ 12
    clean_pages: list[int] | None = None
    if pages is not None:
        if not isinstance(pages, list) or not all(isinstance(p, int) for p in pages):
            raise ValueError("pages 必须是整数列表")
        clean_pages = sorted(set(pages))
        if not clean_pages:
            clean_pages = None
        else:
            if clean_pages[0] < 1 or clean_pages[-1] > 500:
                raise ValueError(f"pages 必须在 1..500 范围内，实际={clean_pages}")
            if len(clean_pages) > 12:
                raise ValueError(
                    f"pages 长度 {len(clean_pages)} 超过上限 12，请缩小范围"
                )

    # 路径白名单：必须在本批次 upstream_root 之下，且是真实文件
    upstream_root = (
        snap.values.get("upstream_root") or get_settings().upstream_root)
    file_path_obj = Path(file_path).expanduser().resolve()
    if not file_path_obj.is_file():
        raise ValueError(f"文件不存在: {file_path}")
    upstream_root_obj = Path(upstream_root).expanduser().resolve()
    try:
        file_path_obj.relative_to(upstream_root_obj)
    except ValueError as e:
        raise ValueError(
            f"文件不在本批次上游工厂文件夹之下，禁止操作 | "
            f"file={file_path_obj} | upstream_root={upstream_root_obj}"
        ) from e

    # 加载批次专属 session JSON（不存在时报清晰错误，避免 FactorySession.load
    # 静默返回空 session 覆盖既有成果）
    sess_path = batch_session_path(thread_id, factory)
    if not sess_path.is_file():
        raise ValueError(
            f"该工厂尚无提取会话（{sess_path}），请先用 retry_factory 重跑整厂后再指定文件"
        )
    session_obj = FactorySession.load(str(sess_path))

    # 执行强制提取（force_vision 在 pages 有值时自动为 True）
    with logging_context(thread_id=thread_id):
        process_result = session_force_extract(
            session_obj, str(file_path_obj), pages=clean_pages,
            force_vision=force_vision,
        )
        if process_result.action == "channel_error":
            raise ValueError(process_result.message or "提取失败")

        # 原子写回 session JSON（与 extraction_node.py:141-148 一致）
        import os, tempfile as _tempfile
        sess_dir = batch_session_dir(thread_id)
        sess_dir.mkdir(parents=True, exist_ok=True)
        data = session_obj.to_dict()
        fd, tmp_path = _tempfile.mkstemp(
            dir=str(sess_dir), prefix=f".{factory}.", suffix=".json.tmp",
        )
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                import json as _json
                _json.dump(data, f, ensure_ascii=False, indent=2)
            os.replace(tmp_path, str(sess_path))
        except Exception:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            raise

        # 不置 force_reextract=True：要让 Node3 命中缓存读回刚才写的内容
        graph.update_state(cfg, {}, as_node=NODE2)

        emit = _make_progress_emitter(on_progress, seed=snap.values)
        for event in graph.stream(None, cfg, stream_mode="updates"):
            emit(event)
            if "__interrupt__" in event:
                _start_pre_extraction(thread_id)
                return {
                    "status": "pending_human_review",
                    "thread_id": thread_id,
                    "factory": factory,
                    "file_path": str(file_path_obj),
                    "pages": clean_pages,
                    "review_data": event["__interrupt__"][0].value,
                    "message": process_result.message,
                }

        final = graph.get_state(cfg)
        return {
            "status": "completed",
            "thread_id": thread_id,
            "factory": factory,
            "file_path": str(file_path_obj),
            "pages": clean_pages,
            "final_output_path": final.values.get("final_output_path"),
            "message": process_result.message,
        }


def get_order_state(thread_id: str) -> dict[str, Any]:
    """查询指定 thread_id 的当前状态（前端轮询/调试入口）。"""
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    return {
        "thread_id": thread_id,
        "exists": bool(snap.values),
        "next_nodes": list(snap.next),
        "values": snap.values,
    }


def get_review_payload(thread_id: str) -> dict[str, Any] | None:
    """从 checkpoint 读取当前挂起的 interrupt payload（审核界面刷新后恢复现场用）。

    注意：payload 不在 state.values 里，而是在 tasks[].interrupts 中。
    未挂起或 thread_id 不存在时返回 None。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    for task in snap.tasks:
        if task.interrupts:
            return task.interrupts[0].value
    return None


def _rebuild_items_from_output_excel(state: dict[str, Any], factory_name: str) -> list[dict]:
    """从最终输出 Excel 按工厂+SKU 聚合重建 calculated_items（兜底）。

    读取 state.values 中的 final_output_path，按 MAKER_MEI_KJ 过滤指定工厂，
    按 SHOHIN_CD 聚合 D_HACCHU_SU / 净重 / 毛重，并反算单件重量。
    读取失败时只记 warning，返回空列表，不阻塞 reopen。
    """
    path = state.get("final_output_path")
    if not path:
        return []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "reopen Excel 兜底读取失败: %s: %s", type(e).__name__, e
        )
        return []

    settings = get_settings()
    header = [c.value for c in ws[1]]

    def _idx(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    col_factory = _idx(settings.col_factory)
    col_sku = _idx(settings.col_sku)
    col_qty = _idx("D_HACCHU_SU")
    col_net = _idx(settings.col_net)
    col_gross = _idx(settings.col_gross)
    col_name_kj = _idx("SHOHIN_MEI_KJ")
    col_name_e = _idx("SHOHIN_MEI_E")
    col_name_cn = _idx(settings.col_name_cn)

    if col_factory is None or col_sku is None or col_qty is None \
            or col_net is None or col_gross is None:
        logger.warning("reopen Excel 兜底缺少必要列")
        return []

    aggregated: dict[str, dict] = {}
    max_idx = max(col_factory, col_sku, col_qty, col_net, col_gross)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max_idx:
            continue
        if row[col_factory] != factory_name:
            continue
        sku = str(row[col_sku] or "")
        if not sku:
            continue
        try:
            qty = float(row[col_qty] or 0)
            net = float(row[col_net] or 0)
            gross = float(row[col_gross] or 0)
        except (TypeError, ValueError):
            continue
        if sku not in aggregated:
            sku_name = None
            if col_name_kj is not None:
                sku_name = row[col_name_kj]
            if not sku_name and col_name_e is not None:
                sku_name = row[col_name_e]
            aggregated[sku] = {
                "sku": sku,
                "qty": qty,
                "net": net,
                "gross": gross,
                "sku_name": sku_name,
                "name_cn": row[col_name_cn] if col_name_cn is not None else None,
            }
        else:
            aggregated[sku]["qty"] += qty
            aggregated[sku]["net"] += net
            aggregated[sku]["gross"] += gross

    out: list[dict] = []
    for agg in aggregated.values():
        qty = agg["qty"]
        net = agg["net"]
        gross = agg["gross"]
        unit_net, net_formula, _ = _safe_div(net, qty)
        unit_gross, gross_formula, _ = _safe_div(gross, qty)
        out.append({
            "sku": agg["sku"],
            "extracted_data": {
                "total_quantity": qty,
                "total_net_weight": net,
                "total_gross_weight": gross,
                "weight_unit": "KG",
                "source_file": str(path),
                "sku_name": agg["sku_name"],
            },
            "calculation": {
                "net_formula": net_formula,
                "gross_formula": gross_formula,
                "calculated_unit_net": unit_net,
                "calculated_unit_gross": unit_gross,
            },
            "status": "Normal",
            "error_msg": None,
            "is_human_edited": False,
            "is_new_sku": False,
            "unexpected_sku": False,
            "db_record": {},
            "name_cn": agg["name_cn"],
        })
    return out


def reopen_factory_for_edit(thread_id: str, factory_name: str) -> dict[str, Any] | None:
    """重开已审核工厂：从 checkpoint state + ReviewAudit 反向构建可编辑 payload。

    适用：用户从批次详情页点「重新打开」已审核工厂，进入 reopen 模式 review。
    不调 graph.update_state / 不调 Command(resume=...)——纯只读，
    让 LangGraph state 保持原状（已审核工厂 state.next 为空，但
    current_factory_data 仍在 state.values 中可读）。
    找不到工厂或数据时返回 None。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    values = snap.values or {}
    cur = values.get("current_factory_data") or {}

    # 优先用当前工厂的 calculated_items；否则从已审核快照 factory_outputs 读取；
    # 都没有则尝试从最终输出 Excel 重建。
    state_factory = cur.get("factory_name")
    if state_factory == factory_name:
        factory_items = cur.get("calculated_items") or []
    else:
        factory_items = (values.get("factory_outputs") or {}).get(factory_name) or []
        if not factory_items:
            factory_items = _rebuild_items_from_output_excel(values, factory_name)

    factory_meta = cur if state_factory == factory_name else {
        "factory_name": factory_name,
        "folder_path": cur.get("folder_path"),
        "source_documents": cur.get("source_documents") or [],
    }

    items_payload = _build_reopen_items_payload(factory_items, thread_id, factory_name)

    payload = {
        "factory_name": factory_name,
        "folder_path": factory_meta.get("folder_path"),
        "source_documents": factory_meta.get("source_documents") or [],
        "missing_skus": cur.get("missing_skus") or [],
        "items": items_payload,
        "extraction_issues": cur.get("extraction_issues") or [],
        "extraction_coverage": cur.get("extraction_coverage") or {},
        "weight_diff_warn_ratio": get_settings().weight_diff_warn_ratio,
        "reopen_mode": True,  # 标记前端进入 reopen 模式（提交走 reopen 端点）
    }
    return payload


def _build_reopen_items_payload(items: list[dict], thread_id: str,
                                factory_name: str) -> list[dict]:
    """把 calculated_items 转成 review_data.items 形状 + ReviewAudit 反向填充。

    反向填充：
    - changes_json 中的 diff → 覆写 extracted_data 对应字段
    - new_skus_json → 补 name_cn / hs_code / inspection_required / name_en / name_jp
    """
    audit = _latest_approved_audit(thread_id, factory_name)
    changes = json.loads(audit.changes_json or "[]") if audit else []
    new_skus = json.loads(audit.new_skus_json or "[]") if audit else []

    by_sku_changes: dict[str, list[dict]] = {}
    for c in changes:
        by_sku_changes.setdefault(str(c.get("sku") or ""), []).append(c)
    by_sku_new = {str(n.get("sku") or ""): n for n in new_skus}

    out: list[dict] = []
    for item in items:
        sku = str(item.get("sku") or "")
        ed = dict(item.get("extracted_data") or {})
        # 反向覆写人工改过的字段
        for c in by_sku_changes.get(sku, []):
            field = c.get("field")
            if field:
                ed[field] = c.get("new")
        # 反向补新 SKU 的人工补录字段
        ns = by_sku_new.get(sku)
        rebuilt = dict(item)
        rebuilt["extracted_data"] = ed
        rebuilt["is_human_edited"] = bool(by_sku_changes.get(sku)) or bool(ns)
        if ns:
            for f in ("name_cn", "name_en", "name_jp",
 "hs_code", "inspection_required"):
                v = ns.get(f)
                if v is not None:
                    rebuilt[f] = v
        out.append(rebuilt)
    return out


def _latest_approved_audit(thread_id: str, factory_name: str) -> ReviewAudit | None:
    """查本批次本工厂最近一条已批准的 ReviewAudit。无返回 None。"""
    try:
        with get_session() as db:
            return db.scalar(
                select(ReviewAudit)
                .where(
                    ReviewAudit.thread_id == thread_id,
                    ReviewAudit.factory_name == factory_name,
                    ReviewAudit.approved.is_(True),
                )
                .order_by(ReviewAudit.ts.desc())
                .limit(1)
            )
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "查 ReviewAudit 失败（reopen 不阻塞）| thread=%s | factory=%s | %s",
            thread_id, factory_name, e,
        )
        return None


def apply_reopen_payload(thread_id: str, factory_name: str,
                         resume_data: dict) -> dict[str, Any]:
    """把 reopen 模式的编辑结果写回 Excel + master.db，不污染 LangGraph state。

    通过构造临时 state dict 复用现有 writer._write_excel / _upsert_db 逻辑，
    不重构 writer.py。调 _write_audit 留痕（同 resume_order 审计契约）。
    """
    from app.nodes import writer as writer_mod  # 延迟导入避免循环

    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    values = snap.values or {}
    cur = values.get("current_factory_data") or {}

    # 目标工厂不一定是当前工厂；允许 reopen 已审核过的其他工厂
    if not values.get("downstream_file_path"):
        raise ValueError(f"批次 {thread_id} 缺少 downstream_file_path，无法 reopen")

    # 构造临时 state（最小字段集供 _write_excel / _upsert_db 使用）
    items = (resume_data or {}).get("items") or []

    # 取原始 items 做 diff：当前工厂直接取；已审核工厂从 factory_outputs 或 Excel 重建
    if cur.get("factory_name") == factory_name:
        original_items = cur.get("calculated_items") or []
    else:
        original_items = (values.get("factory_outputs") or {}).get(factory_name) or []
        if not original_items:
            original_items = _rebuild_items_from_output_excel(values, factory_name)

    fake_state: dict[str, Any] = {
        "batch_id": values.get("batch_id") or thread_id,
        "downstream_file_path": values["downstream_file_path"],
        "downstream_row_map": values.get("downstream_row_map") or {},
        "validation_status": "Approved",  # reopen 视为通过
        "current_factory_data": {
            **cur,
            "factory_name": factory_name,
            "calculated_items": items,
        },
    }

    with logging_context(thread_id=thread_id, factory=factory_name):
        out_path = writer_mod._ensure_output_copy(fake_state)
        written = writer_mod._write_excel(fake_state, out_path)
        inserted, updated = writer_mod._upsert_db(fake_state)

    # 审计留痕：复用 _prepare_audit + _write_audit（仅 approved=True 路径）
    try:
        prepared = {
            "thread_id": thread_id,
            "factory_name": factory_name,
            "approved": bool((resume_data or {}).get("approved", True)),
            "edited_count": sum(
                1 for i in items
                if i.get("is_human_edited") or i.get("is_new_sku")
            ),
            "changes": _prepare_audit_changes_from_items(items, original_items),
            "new_skus": _prepare_audit_new_skus(items),
        }
        _write_audit(prepared, "reopen")
    except Exception as e:  # noqa: BLE001 同 _write_audit 兜底
        logger.warning("⚠️ reopen 审计落库失败: %s: %s", type(e).__name__, e)

    logger.info(
        "[reopen] thread=%s factory=%s written=%d insert=%d update=%d",
        thread_id, factory_name, written, inserted, updated,
    )
    return {
        "status": "success",
        "message": (
            f"工厂「{factory_name}」重新提交完成："
            f"写入 {written} 行 Excel / INSERT {inserted} / UPDATE {updated}"
        ),
        "thread_id": thread_id,
        "factory_name": factory_name,
        "final_output_path": str(out_path),
        "written": written,
        "inserted": inserted,
        "updated": updated,
    }


def _prepare_audit_changes_from_items(new_items: list[dict],
                                     old_items: list[dict]) -> list[dict]:
    """对比 new_items 与 old_items，提取 diff（_prepare_audit 简化版）。"""
    old_by_sku = {str(i.get("sku") or ""): i for i in old_items}
    changes: list[dict] = []
    for item in new_items:
        sku = str(item.get("sku") or "")
        old = old_by_sku.get(sku) or {}
        old_ext = old.get("extracted_data") or {}
        new_ext = item.get("extracted_data") or {}
        for f in ("total_quantity", "total_net_weight", "total_gross_weight"):
            new_v = new_ext.get(f)
            if new_v is not None and new_v != old_ext.get(f):
                changes.append({"sku": sku, "field": f,
                                "old": old_ext.get(f), "new": new_v})
    return changes


def _prepare_audit_new_skus(items: list[dict]) -> list[dict]:
    """提取 new_skus_json 字段。"""
    out: list[dict] = []
    for item in items:
        if not item.get("is_new_sku"):
            continue
        out.append({
            "sku": item.get("sku"),
            "name_cn": item.get("name_cn"),
            "name_en": item.get("name_en"),
            "name_jp": item.get("name_jp"),
            "hs_code": item.get("hs_code"),
            "inspection_required": item.get("inspection_required"),
        })
    return out


# ---------------------------------------------------------------------------
# 批次管理 API（工作台/批次详情页）：checkpoint 只读枚举 + 状态推导
# ---------------------------------------------------------------------------


def _open_checkpoint_ro() -> sqlite3.Connection:
    """打开 checkpoints.db 的只读连接（URI mode=ro，Windows 兼容）。

    绝不复用 graph 单例内部的 saver 连接（graph.py 的连接无锁，
    跨线程并发读写会互相干扰）；这里每次新建独立只读连接。
    全新部署 db 文件不存在（或文件在但表未建）时先用 SqliteSaver.setup()
    建库建表——mode=ro 打不开不存在的文件、SELECT 打不了无表的库
    （首发批次 500 的坑）。
    """
    path = Path(get_settings().checkpoint_db_abs).resolve()
    conn = None
    if path.exists():
        conn = sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True)
        has_table = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='checkpoints'"
        ).fetchone()
        if has_table:
            return conn
        conn.close()
    # 文件不存在或无表：初始化 schema 后重开只读连接
    from langgraph.checkpoint.sqlite import SqliteSaver

    path.parent.mkdir(parents=True, exist_ok=True)
    rw = sqlite3.connect(str(path))
    try:
        SqliteSaver(rw).setup()
    finally:
        rw.close()
    return sqlite3.connect(f"file:{path.as_posix()}?mode=ro", uri=True)


def _list_thread_ids(conn: sqlite3.Connection) -> list[str]:
    """枚举全部批次 thread_id，按最早 checkpoint 倒序（最新批次在前）。"""
    rows = conn.execute(
        "SELECT thread_id FROM checkpoints "
        "GROUP BY thread_id ORDER BY MIN(checkpoint_id) DESC"
    ).fetchall()
    return [r[0] for r in rows]


def _thread_created_ts(conn: sqlite3.Connection, thread_id: str) -> str | None:
    """取该 thread 最早 checkpoint 的创建时间（blob 内 msgpack 的 ts 字段）。

    checkpoint 表的 metadata 列没有时间戳，创建时间只能从
    checkpoint blob（ormsgpack 序列化）的 "ts" 键解出；解包失败返回 None。
    """
    row = conn.execute(
        "SELECT checkpoint FROM checkpoints "
        "WHERE thread_id = ? ORDER BY checkpoint_id ASC LIMIT 1",
        (thread_id,),
    ).fetchone()
    if not row:
        return None
    try:
        return ormsgpack.unpackb(row[0])["ts"]
    except Exception:  # noqa: BLE001 序列化格式漂移不应拖垮列表
        return None


def _summarize_snapshot(thread_id: str, snap, created_ts: str | None) -> dict[str, Any]:
    """从 StateSnapshot 推导批次摘要：三态 status + 进度 + 当前工厂。"""
    values = snap.values or {}

    # ---- 状态三态 ----
    if not values:
        status = "unknown"
    elif any(t.interrupts for t in snap.tasks):
        status = "pending_review"   # 有待审核 interrupt
    elif snap.next:
        status = "running"          # 无 interrupt 但 next 非空（LLM 提取中）
    else:
        status = "completed"

    # ---- 进度：分母 = downstream_requirements 键集（有 filter 时取交集）----
    total_set = set((values.get("downstream_requirements") or {}).keys())
    factory_filter = values.get("factory_filter")
    if factory_filter:
        total_set &= set(factory_filter)
    pending = set(values.get("pending_factories") or [])
    current = (values.get("current_factory_data") or {}).get("factory_name")
    # 当前工厂已被 pop 出 pending；图还在跑（next 非空）时不算完成
    done = len(total_set - pending) - (1 if current and snap.next else 0)
    done = max(done, 0)

    return {
        "thread_id": thread_id,
        "status": status,
        "progress": {"done": done, "total": len(total_set), "current_factory": current},
        "current_factory": current,
        "created_at": created_ts,
        "updated_at": snap.created_at,
    }


def get_batch_summary(thread_id: str) -> dict[str, Any]:
    """单批次轻量摘要：三态 status + 进度 + 当前工厂（复用 _summarize_snapshot）。

    调度 Agent 开场提示等只需状态/工厂的轻量场景用；比 get_batch_detail
    轻——不查 audit、不加载工厂会话。thread 不存在时返回 status="unknown"、
    current_factory=None 的摘要（不抛异常，调用方自行判空）。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    return _summarize_snapshot(thread_id, snap, None)


def get_batch_upstream_root(thread_id: str) -> str | None:
    """该批次 checkpoint state 里记录的 upstream_root；state 无此键时返回 None。

    供审核页白名单做批次级二级兜底（W1）：改全局路径后，旧批次的单据
    仍应按其创建时（已经过 is_dir 校验）的 root 放行。绝不接受请求参数，
    唯一来源是 checkpoint state。
    注意：state 无 upstream_root 表示该批次走 .env 缺省——由全局白名单
    （自动跟随 settings）覆盖，这里绝不回退 settings 当前值，否则调用方
    会把"当时的 settings"缓存成长期有效的批次 root，改路径后产生陈旧放行。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    root = (snap.values or {}).get("upstream_root")
    return str(root) if root else None


def list_batches() -> dict[str, Any]:
    """批次列表：只读连接枚举 thread，逐个 get_state 推导状态。

    单线程推导异常时降级为 status="error"（其余字段 None），不拖垮整表。
    """
    graph = get_graph()
    with contextlib.closing(_open_checkpoint_ro()) as conn:
        thread_ids = _list_thread_ids(conn)
        created_map = {tid: _thread_created_ts(conn, tid) for tid in thread_ids}

    batches: list[dict[str, Any]] = []
    for tid in thread_ids:
        try:
            snap = graph.get_state(_config(tid))
            batches.append(_summarize_snapshot(tid, snap, created_map[tid]))
        except Exception as e:  # noqa: BLE001 单个坏 thread 不影响整表
            batches.append({
                "thread_id": tid,
                "status": "error",
                "error": f"{type(e).__name__}: {e}",
                "progress": None,
                "current_factory": None,
                "created_at": created_map[tid],
                "updated_at": None,
            })
    return {"batches": batches}


def _load_factory_session(batch_id: str, factory: str) -> dict[str, Any] | None:
    """加载指定批次的工厂提取会话摘要。

    读取路径为 per-batch 目录 ``data/sessions/{safe(batch_id)}/{factory}.json``，
    不再使用全局旧缓存 ``data/sessions/{factory}.json``，避免把上一批次文件错误
    展示给当前批次。per-batch 会话文件不存在或损坏时返回 None。

    coverage 由 items 键集 vs expected_skus 重算（与 extraction/session.py
    的 coverage() 口径一致）。
    """
    path = SESSIONS_DIR / get_settings().safe_path_tag(batch_id) / f"{factory}.json"
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001 损坏的会话文件不阻塞批次详情
        return None

    items = data.get("items") or {}
    expected = set(data.get("expected_skus") or [])
    have = set(items.keys())
    if expected:
        coverage = {
            "extracted": len(have & expected),
            "expected": len(expected),
            "missing": sorted(expected - have),
            "extra": sorted(have - expected),
        }
    else:
        coverage = {"extracted": len(have), "expected": None, "missing": [], "extra": []}

    return {
        "factory": factory,
        "status": data.get("status"),
        "updated_at": data.get("updated_at"),
        "issues": data.get("issues") or [],
        "history": data.get("history") or [],
        "file_records": data.get("file_records") or {},
        "targets": data.get("targets") or [],
        "deferred": data.get("deferred") or [],
        "expected_skus": sorted(expected),
        "items": items,
        "coverage": coverage,
    }


def _usage_with_scope() -> dict[str, Any]:
    """全局 LLM 用量摘要 + scope 标注（进程内累计，无 thread 标签）。"""
    summary = usage_tracker.summary()
    summary["scope"] = "process_lifetime"
    summary["note"] = "进程内累计，重启清零；无 thread 标签，为全局用量"
    return summary


def get_batch_detail(thread_id: str) -> dict[str, Any]:
    """批次详情：state 摘要 + factories[]（角色+会话摘要）+ audit[] + usage。

    thread 不存在时抛 ValueError（路由层转 404）。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    if not snap.values:
        raise ValueError(f"thread {thread_id} 不存在")

    with contextlib.closing(_open_checkpoint_ro()) as conn:
        created_ts = _thread_created_ts(conn, thread_id)

    detail = _summarize_snapshot(thread_id, snap, created_ts)
    values = snap.values

    total_set = set((values.get("downstream_requirements") or {}).keys())
    factory_filter = values.get("factory_filter")
    if factory_filter:
        total_set &= set(factory_filter)
    pending = set(values.get("pending_factories") or [])
    current = (values.get("current_factory_data") or {}).get("factory_name")

    factories = []
    for name in sorted(total_set):
        if name == current and snap.next:
            role = "current"
        elif name in pending:
            role = "pending"
        else:
            role = "done"
        factories.append({
            "factory": name,
            "role": role,
            "session": _load_factory_session(thread_id, name),
        })

    with get_session() as session:
        audit_rows = session.scalars(
            select(ReviewAudit)
            .where(ReviewAudit.thread_id == thread_id)
            .order_by(ReviewAudit.audit_id)
        ).all()
        audit = [{
            "ts": r.ts.isoformat() if r.ts else None,
            "factory_name": r.factory_name,
            "approved": r.approved,
            "edited_count": r.edited_count,
            "changes": json.loads(r.changes_json or "[]"),
            "new_skus": json.loads(r.new_skus_json or "[]"),
            "result_status": r.result_status,
        } for r in audit_rows]

    detail.update({
        "downstream_file_path": values.get("downstream_file_path"),
        "upstream_root": values.get("upstream_root"),
        "final_output_path": values.get("final_output_path"),
        "validation_status": values.get("validation_status"),
        "factories": factories,
        "audit": audit,
        "usage": _usage_with_scope(),
    })
    return detail


def delete_batch(thread_id: str) -> dict[str, Any]:
    """删除过往批次：清掉 checkpoints.db 里该 thread_id 的全部状态（checkpoints + writes）。

    状态语义（与 _summarize_snapshot 一致）：
      - snap.values 为空 → ValueError（路由转 404）；
      - 有 task.interrupts → pending_review（允许删，废弃场景）；
      - next 非空且无 interrupt → running → RuntimeError（路由转 409）；
      - next 为空 → completed（允许删）。

    删除用独立 rw 连接（不碰 graph 单例 saver 连接，WAL 下并发安全）；
    只动 checkpoints.db——sessions/*.json、主数据、输出文件一律不碰。
    删除成功后再向 review_audits 插一条 batch_deleted 留痕行
    （既有审计记录一律保留；留痕失败只警告，不反过来搞挂已完成的删除）。
    """
    graph = get_graph()
    snap = graph.get_state(_config(thread_id))
    if not snap.values:
        raise ValueError(f"thread {thread_id} 不存在")
    if not any(t.interrupts for t in snap.tasks) and snap.next:
        raise RuntimeError(f"批次正在运行，禁止删除: {thread_id}")

    # L2 日志关联：删除动作及其审计留痕日志携带批次号
    with logging_context(thread_id=thread_id):
        path = Path(get_settings().checkpoint_db_abs).resolve()
        conn = sqlite3.connect(str(path))
        try:
            cur = conn.execute("DELETE FROM writes WHERE thread_id = ?", (thread_id,))
            writes_removed = cur.rowcount
            cur = conn.execute("DELETE FROM checkpoints WHERE thread_id = ?", (thread_id,))
            checkpoints_removed = cur.rowcount
            conn.commit()
        finally:
            conn.close()

        # 审计留痕（顺序：先删成功再留痕；失败只警告，不阻塞返回）
        try:
            with get_session() as session:
                session.add(ReviewAudit(
                    thread_id=thread_id,
                    factory_name=None,
                    approved=False,
                    edited_count=0,
                    changes_json="[]",
                    new_skus_json="[]",
                    result_status="batch_deleted",
                ))
                session.commit()
        except Exception as e:  # noqa: BLE001 与 _write_audit 同哲学：留痕失败不阻塞
            logger.warning("⚠️⚠️ [审计落库失败] thread=%s "
                           "批次已删除，但 batch_deleted 留痕写入失败：%s: %s",
                           thread_id, type(e).__name__, e)

        return {
            "deleted": thread_id,
            "checkpoints_removed": checkpoints_removed,
            "writes_removed": writes_removed,
        }


def batch_delete_batches(thread_ids: list[str]) -> dict[str, Any]:
    """批量删除批次：逐个调用 delete_batch，部分失败不影响其他。

    返回 {deleted: [str], failed: [{id, reason}]}。
    """
    deleted: list[str] = []
    failed: list[dict[str, str]] = []
    for tid in thread_ids:
        try:
            delete_batch(tid)
            deleted.append(tid)
        except (ValueError, RuntimeError) as e:
            failed.append({"id": tid, "reason": str(e)})
    return {"deleted": deleted, "failed": failed}


# ---------------------------------------------------------------------------
# 重复处理预检（W4b）：sessions 完成态 + 审核落库记录，四档判定工厂是否已处理
# ---------------------------------------------------------------------------

# 会话完成态（与 extraction/session.py 状态机口径一致）
_SESSION_DONE_STATUSES = ("complete_auto", "complete_manual")
# 会话进行态：提过但没提完，不算已处理，也不算"从没碰过"
_SESSION_PARTIAL_STATUSES = ("collecting", "waiting_pl")


def _session_status_light(
    batch_id: str,
    factory: str,
) -> tuple[str | None, str | None]:
    """轻量读 sessions/{safe(batch_id)}/{factory}.json，只取 (status, updated_at)。

    仅查本批次目录。每个批次独立缓存，跨批次不回看历史。文件不存在/JSON 损坏
    一律静默回落，返回 (None, None)。
    """
    settings = get_settings()
    path = SESSIONS_DIR / settings.safe_path_tag(batch_id) / f"{factory}.json"
    if not path.is_file():
        return None, None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return None, None
    return data.get("status"), data.get("updated_at")


def check_processed_factories(
    thread_id: str,                                  # ← 新增必填
    downstream_file_path: str | None = None,
    factory_names: list[str] | None = None,
) -> dict[str, Any]:
    """四档预检（仅查本批次）。

    工厂集合：给出 factory_names 时直接用之（不解析文件，调用方已知名单）；
    缺省时解析装箱单（parse_requirements），解析失败抛 ValueError（路由层
    转 422）。

    level 四档：
      - audited：本批次 review_audits 有 approved=true
      - session_complete：本批次 sessions/{batch_id_safe}/{factory}.json 为 complete_*
      - partial：collecting / waiting_pl
      - none：无任何记录
    单工厂查询异常降级 level="none"，不拖垮整表。
    """
    if factory_names is not None:
        names = list(dict.fromkeys(factory_names))  # 保序去重
    else:
        from app.nodes.parse_downstream import parse_requirements

        downstream = downstream_file_path or get_settings().downstream_file_path
        try:
            requirements, _ = parse_requirements(str(Path(downstream).expanduser()))
        except Exception as e:  # noqa: BLE001 统一转 ValueError，调用方转 422
            raise ValueError(
                f"装箱单解析失败，无法预检重复处理: {type(e).__name__}: {e}"
            ) from e
        names = list(requirements.keys())

    # 审核落库记录：一次性查全部工厂的 approved 记录（ts 倒序），逐厂取首条
    last_audit: dict[str, dict[str, Any]] = {}
    if names:
        try:
            with get_session() as session:
                rows = session.scalars(
                    select(ReviewAudit)
                    .where(ReviewAudit.factory_name.in_(names),
                           ReviewAudit.thread_id == thread_id,             # ← 新增：限本批次
                           ReviewAudit.approved.is_(True))
                    .order_by(ReviewAudit.ts.desc())
                ).all()
            for r in rows:
                if r.factory_name in last_audit:
                    continue
                last_audit[r.factory_name] = {
                    "ts": r.ts.isoformat() if r.ts else None,
                    "thread_id": r.thread_id,
                    "result_status": r.result_status,
                }
        except Exception as e:  # noqa: BLE001 审计查询失败按无审核记录处理
            logger.warning("⚠️ [预检] review_audits 查询失败，按无审核记录处理：%s: %s",
                           type(e).__name__, e)

    factories: list[dict[str, Any]] = []
    processed_count = 0
    for name in names:
        try:
            status, updated_at = _session_status_light(thread_id, name)
            audit = last_audit.get(name)
            if audit is not None:
                level = "audited"
            elif status in _SESSION_DONE_STATUSES:
                level = "session_complete"
            elif status in _SESSION_PARTIAL_STATUSES:
                level = "partial"
            else:
                level = "none"
            processed = level in ("audited", "session_complete")
            if processed:
                processed_count += 1
            factories.append({
                "factory": name,
                "processed": processed,
                "session_status": status,
                "session_updated_at": updated_at,
                "last_audit": audit,
                "level": level,
            })
        except Exception as e:  # noqa: BLE001 单厂异常降级 none，不拖垮整表
            logger.warning("⚠️ [预检] 工厂 %s 查询异常，按未处理降级：%s: %s",
                           name, type(e).__name__, e)
            factories.append({
                "factory": name,
                "processed": False,
                "session_status": None,
                "session_updated_at": None,
                "last_audit": None,
                "level": "none",
            })

    return {
        "factories": factories,
        "processed_count": processed_count,
        "total_count": len(factories),
    }


def create_batch(
    thread_id: str,
    downstream_file_path: str | None = None,
    upstream_root: str | None = None,
    factory_filter: list[str] | None = None,
    factory_alias_overrides: dict[str, str] | None = None,
    skip_processed: bool = False,
    on_progress: Callable[[dict], None] | None = None,
) -> dict[str, Any]:
    """发起新批次：thread_id 查重 + 路径存在性校验后复用 run_until_interrupt。

    - thread_id 空白 → ValueError；
    - checkpoints 已有同名 thread → FileExistsError（路由层转 409）；
    - 路径缺省取 settings；downstream 必须是文件、upstream 必须是目录
      （Path(p).expanduser() 兼容 Windows 盘符路径），否则 ValueError
      且消息指明哪个路径不存在（路由层转 422）；
    - factory_filter 只处理指定工厂（调试/冒烟/跳过已处理用），透传
      run_until_interrupt，None=全部工厂；
    - factory_alias_overrides 批次级「仅本次生效」工厂名对照（W5），
      非空才写入 state（Node2 最高优先匹配档，不落盘）；
    - skip_processed=True 时先跑 check_processed_factories 预检（W4b），
      取「未处理工厂」差集作 factory_filter（与显式 factory_filter 互斥，
      factory_filter 优先）；差集为空时不跑图，直接返回 skipped_all
      （规避 Node2 空队列假审核包）。每次调用实时重算差集，不沿用任何
      预览结论。
    """
    thread_id = (thread_id or "").strip()
    if not thread_id:
        raise ValueError("thread_id 不能为空")

    with contextlib.closing(_open_checkpoint_ro()) as conn:
        row = conn.execute(
            "SELECT 1 FROM checkpoints WHERE thread_id = ? LIMIT 1",
            (thread_id,),
        ).fetchone()
    if row:
        raise FileExistsError(f"批次已存在，请换一个 thread_id: {thread_id}")

    settings = get_settings()
    downstream = downstream_file_path or settings.downstream_file_path
    upstream = upstream_root or settings.upstream_root

    d_path = Path(downstream).expanduser()
    if not d_path.is_file():
        raise ValueError(f"下游装箱单路径不存在或不是文件: {downstream}")
    u_path = Path(upstream).expanduser()
    if not u_path.is_dir():
        raise ValueError(f"上游工厂文件夹路径不存在或不是目录: {upstream}")
    # 加固：上游目录必须能匹配到至少 1 个工厂（用 W5 prescan 干跑一遍），
    # 否则下游 12 个工厂会全部 no_folder_matched，浪费算力且用户看到一堆 0 SKU 错误。
    # 默认生产开启；测试场景（空 tmpdir）可设 FACTORY_DIR_STRICT=0 跳过。
    if os.environ.get("FACTORY_DIR_STRICT", "1") == "1":
        try:
            _scan = prescan_factory_aliases(
                downstream_file_path=str(d_path), upstream_root=str(u_path))
            _resolved = _scan.get("resolved") or {}
            _candidates = _scan.get("candidates") or {}
            _unmatched = _scan.get("unmatched") or []
            _total = len(_resolved) + len(_candidates) + len(_unmatched)
            if _total > 0 and not _resolved:
                raise ValueError(
                    f"上游路径下未匹配到任何工厂文件夹: {upstream}"
                    f"（检查路径是否少带 /工厂 等子目录名？例如 83/工厂 而不是 83）")
        except ValueError:
            raise
        except Exception:
            # 装箱单解析失败 / W5 异常 → 软警告，不阻断（交给 dispatcher preview 兜底）
            pass

    # W4b：skip_processed 实时重算差集（不沿用预览结论，防 TTL 内竞态）；
    # 显式 factory_filter 优先（互斥语义）
    effective_filter = factory_filter
    skipped: list[str] = []
    if skip_processed and not factory_filter:
        precheck = check_processed_factories(
            thread_id=thread_id,                              # ← 新增
            downstream_file_path=str(d_path),
        )
        skipped = [f["factory"] for f in precheck["factories"] if f["processed"]]
        remaining = [f["factory"] for f in precheck["factories"]
                     if not f["processed"]]
        if not remaining:
            logger.info("[W4b] 批次 %s 全部工厂均已处理，跳过建图", thread_id)
            return {
                "status": "skipped_all",
                "message": "全部工厂均已处理，无可提取工厂",
                "processed": skipped,
            }
        effective_filter = remaining

    result = run_until_interrupt(thread_id, str(d_path), str(u_path),
                                 factory_filter=effective_filter,
                                 factory_alias_overrides=factory_alias_overrides,
                                 on_progress=on_progress)
    if skipped:
        result["skipped_processed"] = skipped

    # 写入批次配置
    now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    batch_config = {
        "thread_id": thread_id,
        "downstream_file_path": str(d_path),
        "upstream_root": str(u_path),
        "created_at": now,
        "last_run_at": now,
        "run_count": 1,
    }
    _write_batch_config(thread_id, batch_config)
    return result


# ---------------------------------------------------------------------------
# 工厂名对照预扫（W5）：发起批次前一次性把装箱单工厂分三档，供确认门展示
# ---------------------------------------------------------------------------

# 「确定命中」分档线：alias/alias_ci/exact 记 100 直接确定；
# fuzzy 命中须 >= 85 才算确定，否则降级为「低置信推荐」
_PRESCAN_CERTAIN_SCORE = 85.0


def prescan_factory_aliases(
    downstream_file_path: str | None = None,
    upstream_root: str | None = None,
    factory_filter: list[str] | None = None,
) -> dict[str, Any]:
    """工厂名对照预扫：解析装箱单工厂集合，逐个匹配上游一级子目录，分三档返回。

    返回：
      - resolved:   {工厂: {"folder", "score", "method"}} 确定命中
                    （alias/alias_ci/exact，或 fuzzy 得分 >= 85）；
      - candidates: {工厂: [{"folder", "score", "signals"}]} 低置信推荐
                    （fuzzy 40~85 或包含信号），需人工确认；
      - unmatched:  [工厂, ...] 无任何候选，需人工指定；
      - warnings:   [str, ...] 非致命问题。

    缺省参数取 settings。装箱单解析失败 / 目录不可读只进 warnings 不抛出；
    目录不存在时 warnings 且全部工厂进 unmatched。
    """
    from app.factory_match import (
        load_alias_map, match_factory_folder, recommend_candidates)
    from app.nodes.parse_downstream import parse_requirements

    settings = get_settings()
    downstream = downstream_file_path or settings.downstream_file_path
    upstream = upstream_root or settings.upstream_root

    result: dict[str, Any] = {
        "resolved": {}, "candidates": {}, "unmatched": [], "warnings": []}

    try:
        requirements, _ = parse_requirements(str(Path(downstream).expanduser()))
    except Exception as e:  # noqa: BLE001 预扫是辅助设施，解析失败不抛出
        result["warnings"].append(
            f"装箱单解析失败，无法预扫工厂对照: {type(e).__name__}: {e}")
        return result

    factories = list(requirements.keys())
    if factory_filter:
        allow = set(factory_filter)
        factories = [f for f in factories if f in allow]

    u_path = Path(upstream).expanduser()
    if not u_path.is_dir():
        result["warnings"].append(f"上游工厂文件夹不存在或不是目录: {upstream}")
        result["unmatched"] = factories
        return result
    try:
        folders = [d.name for d in u_path.iterdir() if d.is_dir()]
    except OSError as e:
        result["warnings"].append(f"上游工厂文件夹不可读: {e}")
        result["unmatched"] = factories
        return result

    alias_map = load_alias_map()
    cutoff = settings.fuzzy_match_score_cutoff
    for factory in factories:
        folder, score, method = match_factory_folder(
            factory, folders, alias_map, cutoff=cutoff)
        if folder and (method != "fuzzy" or score >= _PRESCAN_CERTAIN_SCORE):
            result["resolved"][factory] = {
                "folder": folder, "score": score, "method": method}
            continue
        candidates = recommend_candidates(factory, folders, cutoff=cutoff)
        if candidates:
            result["candidates"][factory] = candidates
        else:
            result["unmatched"].append(factory)
    return result


# ---------------------------------------------------------------------------
# 审核审计：resume 前抓取原始 payload 做 diff 快照，成功后落 review_audits
# ---------------------------------------------------------------------------

# 人工可修改的底层数值字段（None 与数值严格区分：None 表示未提交/未改动）
_AUDIT_NUMERIC_FIELDS = ("total_quantity", "total_net_weight", "total_gross_weight")


def _prepare_audit(thread_id: str, resume_data: dict) -> dict[str, Any] | None:
    """resume 前准备审计快照：diff 提交 items vs 原始 payload items。

    - changes：数值字段（_AUDIT_NUMERIC_FIELDS）old→new 对照，
      None 与数值严格区分（提交值 None 视为未改动，不计 diff）；
    - new_skus：原始 payload 中 is_new_sku 项的人工补录字段
      （name_cn/hs_code/inspection_required/name_en/name_jp）；
    - 原始 payload 拿不到（未挂起等）返回 None。
    """
    payload = get_review_payload(thread_id)
    if payload is None:
        return None

    orig_items = {i.get("sku"): i for i in payload.get("items") or []}
    changes: list[dict[str, Any]] = []
    new_skus: list[dict[str, Any]] = []
    edited_count = 0

    for item in (resume_data or {}).get("items") or []:
        sku = item.get("sku")
        orig = orig_items.get(sku) or {}
        orig_ext = orig.get("extracted_data") or {}
        new_ext = item.get("extracted_data") or {}

        sku_changed = False
        for f in _AUDIT_NUMERIC_FIELDS:
            new_v = new_ext.get(f)
            if new_v is not None and new_v != orig_ext.get(f):
                sku_changed = True
                # 扁平结构（冻结契约）：batch.html 审计区按 c.field/c.old/c.new 渲染
                changes.append({"sku": sku, "field": f, "old": orig_ext.get(f), "new": new_v})
        if sku_changed:
            edited_count += 1

        if orig.get("is_new_sku"):
            new_skus.append({
                "sku": sku,
                "name_cn": item.get("name_cn"),
                "hs_code": item.get("hs_code"),
                "inspection_required": item.get("inspection_required"),
                "name_en": item.get("name_en"),
                "name_jp": item.get("name_jp"),
            })

    return {
        "thread_id": thread_id,
        "factory_name": payload.get("factory_name"),
        "approved": bool((resume_data or {}).get("approved", False)),
        "edited_count": edited_count,
        "changes": changes,
        "new_skus": new_skus,
    }


def _write_audit(prepared: dict[str, Any] | None, result_status: str | None) -> None:
    """把审计快照写入 review_audits。

    审计是辅助设施，不能反过来搞挂已成功的 resume：本函数 try/except
    包死，任何失败只记警告日志，绝不向外抛出。
    """
    if not prepared:
        return
    try:
        with get_session() as session:
            session.add(ReviewAudit(
                thread_id=prepared["thread_id"],
                factory_name=prepared.get("factory_name"),
                approved=prepared.get("approved", False),
                edited_count=prepared.get("edited_count", 0),
                changes_json=json.dumps(prepared.get("changes") or [], ensure_ascii=False),
                new_skus_json=json.dumps(prepared.get("new_skus") or [], ensure_ascii=False),
                result_status=result_status,
            ))
            session.commit()
    except Exception as e:  # noqa: BLE001 故意包死，见 docstring
        logger.warning("⚠️⚠️ [审计落库失败] thread=%s "
                       "resume 已成功，但 review_audits 写入失败：%s: %s",
                       prepared.get('thread_id'), type(e).__name__, e)
