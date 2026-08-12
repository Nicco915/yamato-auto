# -*- coding: utf-8 -*-
"""提取层回归基线快照生成器（批次 2 / Step 2.1）。

设计目标
--------
为雅玛多单证项目生成提取层的回归基线快照，用于 Step 2.1 之后的
A1+A2a+A3+A4+C2 改动后做 diff 对比。本脚本只跑「**确定性管线**」部分，
不调真实 LLM——基线快照记录每个 fixture 文件经过确定性处理后的结构特征：

1. markdown 文本（excel 转出的 PL sheet 纯文本，确定性）
2. markdown 行数（同值压缩前/后，用于 A1 验证 _df_to_markdown 行为）
3. 候选画像（target_identifier.scan_file 的输出，确定性）
4. _bind_orphan_barcode 触发后的最终列数（A2a 验证项）
5. _drop_zero_rows 过滤前后行数（_merge 接收 mock 结果的覆盖率前置）

不包含 LLM 输出（items）——LLM 行为由 Step 2.3 端到端验证覆盖
（仅达安+兆丰+其他工厂抽样）。

测试红线（2026-08-11 事故教训）
------------------------------
本脚本启动时强制要求 ``YAMATO_TEST_MODE=1``，并强制使用 ``YAMATO_SESSIONS_DIR``
指向临时目录，绝不读真实 ``data/sessions/``。fixture 临时复制到隔离工作目录
（避免污染原 fixture 的访问时间 / mtime）。
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
from pathlib import Path

# ----- 测试红线守卫：必须在 import app.* 之前 -----
assert os.environ.get("YAMATO_TEST_MODE") == "1", (
    "extraction_snapshot 仅用于测试，必须设 YAMATO_TEST_MODE=1 后再跑"
)

# 允许从任意 cwd 直接运行本文件
APP_ROOT = Path(__file__).resolve().parents[2]
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

# 现在才能 import app.* —— session.py 模块级会读 YAMATO_SESSIONS_DIR
# （已由调用方注入临时目录）
from app.extraction.excel_channel import (  # noqa: E402
    EXCEL_SUFFIXES,
    _bind_orphan_barcode,
    _drop_zero_rows,
    _df_to_markdown,
    _grid_to_dataframe,
    _unmerge_sheet_to_grid,
    legacy_xls_to_markdown,
    select_pl_sheets,
    xlsx_to_markdown,
)
from app.extraction.target_identifier import scan_file  # noqa: E402

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")


# ---------------------------------------------------------------------------
# 确定性管线：每个 fixture 文件 → 结构特征 dict
# ---------------------------------------------------------------------------

def _profile_excel(file_path: Path, work_dir: Path) -> dict:
    """对 Excel/CSV 跑确定性管线，产出结构特征（不调 LLM）。

    字段：
    - factory_slug         : 工厂目录名（fixture 一级目录）
    - source_filename      : 文件名（含扩展名）
    - source_path          : 隔离副本的绝对路径
    - suffixes             : 后缀列表（小写）
    - selected_sheets      : 选中的 PL sheet 列表（确定性）
    - markdown_raw         : 完整 markdown 文本（确定性）
    - markdown_raw_lines   : markdown 行数（确定性）
    - markdown_after_zero_filter : _drop_zero_rows 后的 markdown
    - markdown_after_zero_filter_lines : 同上，行数
    - markdown_after_bind   : _bind_orphan_barcode 后重渲染的 markdown
    - markdown_after_bind_lines : 同上，行数
    - last_col_after_bind    : _bind_orphan_barcode 后的最大列数（A2a 验证项）
    - profile               : scan_file 画像（确定性）
    """
    suffix = file_path.suffix.lower()
    selected = []
    raw_md = ""
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), data_only=True, read_only=False)
        selected = select_pl_sheets([ws.title for ws in wb.worksheets])
        raw_md = xlsx_to_markdown(str(file_path))
    elif suffix == ".xls":
        import pandas as pd
        sheets = pd.read_excel(str(file_path), sheet_name=None, header=None, engine="xlrd")
        selected = select_pl_sheets(list(sheets.keys()))
        raw_md = legacy_xls_to_markdown(str(file_path))
    elif suffix == ".csv":
        import pandas as pd
        # CSV 没有 sheet 概念；选中的就是 "csv"
        selected = ["csv"]
        df = pd.read_csv(str(file_path), header=None)
        raw_md = _df_to_markdown(df.head(400))

    md_after_zero = _drop_zero_rows(raw_md)

    # _bind_orphan_barcode 在 _df_to_markdown 之前应用，
    # 需要重建 DataFrame 才能拿到准确列数；只在 xlsx/xls 上做
    bound_col_count = None
    bound_md = md_after_zero
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(str(file_path), data_only=True, read_only=False)
        parts: list[str] = []
        max_cols_after_bind = 0
        for ws in wb.worksheets:
            if ws.title not in selected:
                continue
            grid = _unmerge_sheet_to_grid(ws)
            df = _grid_to_dataframe(grid)
            df2 = _bind_orphan_barcode(df)
            max_cols_after_bind = max(max_cols_after_bind, df2.shape[1])
            md = _df_to_markdown(df2)
            if md:
                parts.append(f"### 工作表: {ws.title}\n{md}")
        bound_md = _drop_zero_rows("\n\n".join(parts))
    elif suffix == ".xls":
        import pandas as pd
        sheets = pd.read_excel(str(file_path), sheet_name=None, header=None, engine="xlrd")
        parts = []
        max_cols_after_bind = 0
        for name, df in sheets.items():
            if name not in selected:
                continue
            df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
            df = df.ffill(axis=1)
            df2 = _bind_orphan_barcode(df.reset_index(drop=True))
            max_cols_after_bind = max(max_cols_after_bind, df2.shape[1])
            md = _df_to_markdown(df2)
            if md:
                parts.append(f"### 工作表: {name}\n{md}")
        bound_md = _drop_zero_rows("\n\n".join(parts))
    elif suffix == ".csv":
        # CSV 不应用 _bind_orphan_barcode（只对 Excel sheet 启用）
        max_cols_after_bind = None

    # 候选画像（确定性）
    prof = scan_file(str(file_path))
    profile_dict = {
        "path": prof.path,
        "channel": prof.channel,
        "barcodes": sorted(prof.barcodes),
        "has_net": prof.has_net,
        "has_gross": prof.has_gross,
        "has_qty": prof.has_qty,
        "is_candidate": prof.is_candidate,
        "error": prof.error,
    }

    return {
        "factory_slug": file_path.parent.name,
        "source_filename": file_path.name,
        "source_path": str(file_path),
        "suffix": suffix,
        "selected_sheets": selected,
        "markdown": raw_md,
        "markdown_lines": len(raw_md.split("\n")),
        "markdown_after_zero_filter": md_after_zero,
        "markdown_after_zero_filter_lines": len(md_after_zero.split("\n")),
        "markdown_after_bind": bound_md,
        "markdown_after_bind_lines": len(bound_md.split("\n")),
        "last_col_after_bind": max_cols_after_bind,
        "profile": profile_dict,
    }


def _collect_excel_files(fixture_root: Path) -> list[Path]:
    """递归收集 fixture 目录下所有 Excel/CSV 文件（含 .xls）。"""
    files: list[Path] = []
    for p in sorted(fixture_root.rglob("*")):
        if not p.is_file():
            continue
        if p.name.startswith(".") or p.name in {".DS_Store", "Thumbs.db"}:
            continue
        if p.suffix.lower() in EXCEL_SUFFIXES:
            files.append(p)
    return files


def _copy_to_workdir(src: Path, work_dir: Path) -> Path:
    """复制 fixture 文件到隔离工作目录（保 mtime 不污染原文件访问时间）。"""
    target = work_dir / src.name
    # 读字节 + 写字节，避免 shutil.copy2 改 mtime
    target.write_bytes(src.read_bytes())
    return target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--fixtures", nargs="+", required=True,
        help="一个或多个 fixture 根目录（递归遍历 xls/xlsx/xlsm/csv）",
    )
    parser.add_argument(
        "--out", required=True,
        help="基线快照输出根目录（按 factory_slug/ 分子目录）",
    )
    args = parser.parse_args(argv)

    fixtures = [Path(f).resolve() for f in args.fixtures]
    out_root = Path(args.out).resolve()

    # 临时工作目录（fixture 隔离副本）
    work_root = Path(tempfile.mkdtemp(prefix="yamato_extraction_snapshot_"))
    logger.info("临时工作目录：%s", work_root)

    summary: dict[str, list[str]] = {}
    for fix_root in fixtures:
        if not fix_root.exists():
            logger.warning("fixture 目录不存在：%s", fix_root)
            continue
        factory_slug = fix_root.name
        out_dir = out_root / factory_slug
        out_dir.mkdir(parents=True, exist_ok=True)
        summary[factory_slug] = []

        # 每工厂一个独立子工作目录（避免同 factory 内重名）
        factory_work = work_root / factory_slug
        factory_work.mkdir(parents=True, exist_ok=True)

        for src in _collect_excel_files(fix_root):
            iso_copy = _copy_to_workdir(src, factory_work)
            data = _profile_excel(iso_copy, factory_work)
            snap_path = out_dir / f"{iso_copy.stem}.json"
            snap_path.write_text(
                json.dumps(data, ensure_ascii=False, indent=1),
                encoding="utf-8",
            )
            logger.info(
                "生成快照 | factory=%s | file=%s | md_lines=%d | after_bind=%d | cols=%s",
                factory_slug, iso_copy.name,
                data["markdown_lines"],
                data["markdown_after_bind_lines"],
                data["last_col_after_bind"],
            )
            summary[factory_slug].append(str(snap_path))

    print()
    print(json.dumps(summary, ensure_ascii=False, indent=1))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())