"""Node 6: 写入与循环（Writer Node）——双写机制（《第三阶段.md》改造点 C）。

1. 写业务表：把人工确认后的精确数据写回下游 Excel 副本
   （先复制到 output/ 目录，绝不覆盖原件）；
   写入规则（2026-07-28 用户定）：
   - 原文件没有 中文品名/净重/毛重 三列：首次写入时在 SHOHIN_MEI_E 后插入
     （与既有填好文件布局一致：第 32/33/34 列），已存在则跳过；
   - 净重 = 单件净重 × SOTOBAKO_D_HACCHU_SU，毛重同理，2 位小数；
   - 中文品名 = 主数据 name_cn（新 SKU 经 Node5 人工补录）；
   - 写入单元格格式：字号 9 + 四周细边框（2026-08-04 用户定），
     字体族/加粗保留单元格原有设置；
   - 表格格式严格不变：全程 openpyxl，禁止 pandas 写入。
2. 写数据库（Upsert）：
   - 新 SKU：INSERT 人工补录的多语言品名/HS 编码/单件重量；
   - 老 SKU：人工微调过重量时 UPDATE 刷新历史重量。

写盘保护（2026-08-11）：落盘前做可写探测 + 捕获 PermissionError/OSError(EACCES/EBUSY)，
把英文 traceback 转成中文提示。Windows 下 Excel/LibreOffice 独占锁文件时，
裸 PermissionError 会让用户不知所措；探测可以在做完一堆耗时工作之前先发现问题。
"""
import errno
import logging
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from sqlalchemy import select

from app.config import get_settings
from app.db.models import Factory, FactorySKU
from app.db.session import get_session
from app.logging_config import bind_factory_from_state
from app.state import AgentState

logger = logging.getLogger(__name__)

# Windows 共享冲突（Excel/LibreOffice 独占打开产生的 EBUSY）也归为「文件被占用」一类提示
_FILE_LOCK_ERRNOS = frozenset({errno.EACCES, errno.EBUSY, errno.EPERM, errno.ETXTBSY})

# 待添加的三列（插入到 SHOHIN_MEI_E 之后，与既有填好文件布局一致）
NEW_COL_NAMES = ("中文品名", "净重", "毛重")
INSERT_AFTER_COL = "SHOHIN_MEI_E"

# 写入单元格格式（2026-08-04 用户定）：字号 9 + 四周细边框
WRITE_FONT_SIZE = 9
_THIN_SIDE = Side(style="thin")
WRITE_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE, bottom=_THIN_SIDE)


def _apply_write_format(cell) -> None:
    """写入单元格格式：字号 9 + 四周细边框；字体族/加粗/颜色保留原设置。"""
    font = copy(cell.font)   # StyleProxy → Font，可改后再赋值
    font.sz = WRITE_FONT_SIZE
    cell.font = font
    cell.border = WRITE_BORDER


def _ensure_output_copy(state: AgentState) -> Path:
    """确保 output/{batch_id}/containers/ 下存在原件副本；已存在则复用。"""
    settings = get_settings()
    batch_id = state.get("batch_id") or "unknown"
    out_dir = settings.batch_containers_dir(batch_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    src = Path(state["downstream_file_path"])
    dst = out_dir / f"{src.stem}_filled{src.suffix}"
    if not dst.exists():
        try:
            shutil.copy2(src, dst)
        except OSError as e:
            # 原件被占用时也走中文提示；优先指向原文件路径（用户日常认知的是原件）
            if _is_file_lock_error(e):
                raise RuntimeError(_format_file_busy_msg(src)) from e
            raise
        logger.info("[Node6] 已复制原件到 %s（绝不覆盖原件）", dst)
    return dst


def _probe_writable(path: Path) -> None:
    """以追加模式短暂打开再关闭，验证可写。绝不以「w」模式探测（会清空文件）。

    Excel/LibreOffice 在 Windows 下打开 xlsx 会独占锁，此时以「a」模式打开仍会
    抛 PermissionError(EBUSY)；文件不存在则跳过探测，由后续真实写入走相同异常路径。

    注意：只有 PermissionError 或 _FILE_LOCK_ERRNOS 中的错误才会被翻译成
    「文件被占用」提示；IsADirectoryError/FileNotFoundError/EIO 等仍原样抛出，
    避免误导用户。
    """
    if not path.exists():
        return
    try:
        with path.open("a"):
            pass
    except OSError as e:
        if _is_file_lock_error(e):
            raise RuntimeError(_format_file_busy_msg(path)) from e
        raise


def _is_file_lock_error(exc: OSError) -> bool:
    """判断异常是否属于「文件被占用 / 权限拒绝」类错误。"""
    errno_no = getattr(exc, "errno", None)
    return errno_no in _FILE_LOCK_ERRNOS or isinstance(exc, PermissionError)


def _format_file_busy_msg(path: Path) -> str:
    """把 PermissionError/OSError(EACCES/EBUSY) 翻译成可读的中文提示。

    Windows 下 EBUSY/EACCES 几乎都意味着 Excel/LibreOffice 独占锁；
    macOS/Linux 下若出现则多是目录权限或罕见共享冲突。
    """
    return (
        f"输出文件正被 Excel 或其他程序占用，请先关闭该文件后重试。"
        f"文件路径：{path}"
    )


def _ensure_three_columns(ws) -> None:
    """表头缺 中文品名/净重/毛重 时在 SHOHIN_MEI_E 后插入；已存在则跳过。

    部分缺失属于异常布局（非原文件也非既有填好文件），零容错直接报错交人工。
    """
    header = [c.value for c in ws[1]]
    missing = [n for n in NEW_COL_NAMES if n not in header]
    if not missing:
        return
    if len(missing) != len(NEW_COL_NAMES):
        raise ValueError(f"下游表列布局异常：三列部分缺失 {missing}，请人工确认")
    anchor = header.index(INSERT_AFTER_COL) + 1  # 1 基列号
    ws.insert_cols(anchor + 1, len(NEW_COL_NAMES))
    # 表头样式照搬锚点列（字体/边框/填充），保证格式严格不变
    for j, name in enumerate(NEW_COL_NAMES):
        cell = ws.cell(row=1, column=anchor + 1 + j, value=name)
        cell._style = copy(ws.cell(row=1, column=anchor)._style)
    logger.info("[Node6] 原文件无 %s 三列，已在 %s 后插入",
                list(NEW_COL_NAMES), INSERT_AFTER_COL)


def _write_excel(state: AgentState, out_path: Path) -> int:
    """按行号精准写回 中文品名/净重/毛重 单元格，返回写入的行数。"""
    settings = get_settings()
    cur = state.get("current_factory_data") or {}
    factory = cur.get("factory_name")
    row_map = (state.get("downstream_row_map") or {}).get(factory) or {}

    wb = load_workbook(out_path)
    ws = wb[wb.sheetnames[0]]
    _ensure_three_columns(ws)

    # 插入列后按表头名重新定位（列位可能已右移）
    header = [c.value for c in ws[1]]
    col_net = header.index(settings.col_net) + 1
    col_gross = header.index(settings.col_gross) + 1
    col_cn = header.index(settings.col_name_cn) + 1
    col_qty = header.index(settings.col_qty) + 1

    written = 0
    for item in cur.get("calculated_items") or []:
        calc = item.get("calculation") or {}
        unit_net = calc.get("calculated_unit_net")
        unit_gross = calc.get("calculated_unit_gross")
        name_cn = item.get("name_cn") or (item.get("db_record") or {}).get("name_cn")
        if unit_net is None and unit_gross is None and not name_cn:
            continue  # Error 项无有效单重，留给人工线下处理
        for excel_row in row_map.get(str(item.get("sku")), []):
            qty = ws.cell(row=excel_row, column=col_qty).value
            qty = float(qty) if isinstance(qty, (int, float)) else 0.0
            if name_cn:
                cell = ws.cell(row=excel_row, column=col_cn, value=name_cn)
                _apply_write_format(cell)
            if unit_net is not None:
                cell = ws.cell(row=excel_row, column=col_net, value=round(unit_net * qty, 2))
                _apply_write_format(cell)
            if unit_gross is not None:
                cell = ws.cell(row=excel_row, column=col_gross, value=round(unit_gross * qty, 2))
                _apply_write_format(cell)
            written += 1
    # 落盘前可写探测：让 Excel 独占锁导致的失败尽早冒泡，避免做完耗时写操作才报错
    _probe_writable(out_path)
    try:
        wb.save(out_path)
    except (PermissionError, OSError) as e:
        if _is_file_lock_error(e):
            raise RuntimeError(_format_file_busy_msg(out_path)) from e
        raise
    return written


def _upsert_db(state: AgentState) -> tuple[int, int]:
    """主数据落库，返回 (插入数, 更新数)。"""
    cur = state.get("current_factory_data") or {}
    factory_name = cur.get("factory_name") or "未知工厂"
    inserted = updated = 0

    with get_session() as session:
        factory = session.scalar(
            select(Factory).where(Factory.factory_name == factory_name)
        )
        if factory is None:
            factory = Factory(factory_name=factory_name)
            session.add(factory)
            session.flush()

        for item in cur.get("calculated_items") or []:
            sku = str(item.get("sku") or "")
            if not sku:
                continue
            calc = item.get("calculation") or {}
            unit_net = calc.get("calculated_unit_net")
            unit_gross = calc.get("calculated_unit_gross")

            record = session.scalar(
                select(FactorySKU).where(
                    FactorySKU.factory_id == factory.factory_id,
                    FactorySKU.sku_code == sku,
                )
            )
            if record is None:
                # 新 SKU：INSERT（含人工补录的合规字段）
                record = FactorySKU(
                    factory_id=factory.factory_id,
                    sku_code=sku,
                    name_cn=item.get("name_cn"),
                    name_en=item.get("name_en"),
                    name_jp=item.get("name_jp"),
                    hs_code=item.get("hs_code"),
                    inspection_required=bool(item.get("inspection_required", False)),
                    unit_net_weight=unit_net,
                    unit_gross_weight=unit_gross,
                )
                session.add(record)
                inserted += 1
            elif item.get("is_human_edited"):
                # 老 SKU 且人工微调过：UPDATE 刷新重量与合规字段
                if unit_net is not None:
                    record.unit_net_weight = unit_net
                if unit_gross is not None:
                    record.unit_gross_weight = unit_gross
                for f in ("name_cn", "name_en", "name_jp", "hs_code"):
                    if item.get(f) is not None:
                        setattr(record, f, item[f])
                if item.get("inspection_required") is not None:
                    record.inspection_required = bool(item["inspection_required"])
                updated += 1
        session.commit()
    return inserted, updated


def writer(state: AgentState) -> dict:
    # L2 日志关联：从 state 重绑当前工厂名（多工厂 resume 链上 submit 拷贝的
    # context 残留上一工厂）；节点独立 context 保证不外泄，无需清理
    bind_factory_from_state(state)
    cur = state.get("current_factory_data") or {}
    factory = cur.get("factory_name")

    if state.get("validation_status") != "Approved":
        logger.warning("[Node6] 工厂「%s」审核未通过（%s），跳过写入",
                       factory, state.get('validation_status'))
        return {}

    out_path = _ensure_output_copy(state)
    written = _write_excel(state, out_path)
    inserted, updated = _upsert_db(state)

    logger.info("[Node6] 工厂「%s」：写入 %d 行 Excel；"
                "落库 INSERT %d / UPDATE %d", factory, written, inserted, updated)

    factory_outputs = state.get("factory_outputs") or {}
    # 存完整 current_factory_data 快照（reopen 时全文恢复，与首次审核字段一致）；
    # 旧格式列表（仅 calculated_items）由 reopen 层兼容读取
    factory_outputs[factory] = dict(cur)
    return {"final_output_path": str(out_path), "factory_outputs": factory_outputs}
