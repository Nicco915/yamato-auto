# -*- coding: utf-8 -*-
"""报关单生成器单测（pytest）。

真实文件：ContentsOfTheContainer_202624_青島XD_20260708.xlsx
真实映射/品名组：master.db（37 条 product_mappings + 3 组品名组）
人工样本：96/报关单/报关*.xls（港口级 USD 总额对账）
"""

from __future__ import annotations

import glob
import os
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path

import pytest
import xlrd

from app.declare.aggregator import (
    aggregate_ticket,
    normal_row_sort_key,
    rows_for_ticket,
)
from app.declare.mapping import build_mapping_index
from app.declare.naming import (
    PORT_MAP,
    declaration_filename,
    format_onboard,
    ticket_letter,
    ticket_title,
)
from app.declare.template_filler import fill_declaration
from app.split.engine import propose
from app.split.loader import load_filled_excel
from app.split.normalize import classify_sj_factories, normalize_maker

# ---- Constants ----

FIXTURE = "/Users/nz/Downloads/yamato/96/ContentsOfTheContainer_202624_青島XD_20260708.xlsx"
SAMPLES_DIR = "/Users/nz/downloads/yamato/96/报关单"
DB_PATH = "/Users/nz/Downloads/yamato/app/app/data/master.db"
TEMPLATE = "/Users/nz/Downloads/yamato/app/app/templates/declaration_template.xlsx"

# 主仓库模版路径的兜底（worktree 内运行时使用相对副本）
_WORKTREE_TEMPLATE = str(
    Path(__file__).resolve().parents[1] / "templates" / "declaration_template.xlsx"
)

FALLBACK_SJ = ["青島貝来", "Ｃ．正達工芸品"]
NORMALIZE_MAP = {
    "青島貝来国際貿易有限公司": "青島貝来",
    "上海億鑽五金工具有限公司（青島）": "上海億鑽五金工具（青島）",
}

TOKYO_USD_KNOWN = 186264.641  # 人工核对的东京 USD 总额

_REQUIRES = [FIXTURE, SAMPLES_DIR, DB_PATH]


def _db_rows(sql: str, args=()):
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        return [dict(r) for r in con.execute(sql, args)]
    finally:
        con.close()


# ---- Fixtures ----

@pytest.fixture(scope="module")
def raw_items():
    if not Path(FIXTURE).exists():
        pytest.skip(f"Real data file not found: {FIXTURE}")
    raw = load_filled_excel(FIXTURE)
    for r in raw:
        r.maker = normalize_maker(r.maker, NORMALIZE_MAP)
    return raw


@pytest.fixture(scope="module")
def sj_map(raw_items):
    return classify_sj_factories(raw_items, {}, FALLBACK_SJ)


@pytest.fixture(scope="module")
def proposal(raw_items, sj_map):
    return propose(raw_items, sj_map)


@pytest.fixture(scope="module")
def groups():
    """品名组配置（DB → dict，aggregator 不碰 DB）。"""
    if not Path(DB_PATH).exists():
        pytest.skip(f"master.db not found: {DB_PATH}")
    out = []
    for g in _db_rows("SELECT * FROM product_groups"):
        members = _db_rows(
            "SELECT * FROM product_group_members WHERE group_id=? "
            "ORDER BY display_order",
            (g["id"],),
        )
        out.append({**g, "members": members})
    return out


@pytest.fixture(scope="module")
def mapping_index():
    if not Path(DB_PATH).exists():
        pytest.skip(f"master.db not found: {DB_PATH}")
    return build_mapping_index(_db_rows("SELECT * FROM product_mappings"))


@pytest.fixture(scope="module")
def ticket_results(proposal, raw_items, sj_map, groups, mapping_index):
    """每票 -> (ticket, 源行, AggregateResult)。"""
    results = []
    for pg in proposal.ports:
        for t in pg.groups:
            rows = rows_for_ticket(t, raw_items, sj_map)
            res = aggregate_ticket(rows, groups, mapping_index)
            results.append((t, rows, res))
    return results


# ---- Helpers ----

def _sample_port_usd(cn_port: str) -> float:
    """读人工样本：某港口全部报关文件的 USD 汇总行 F 值求和。

    汇总行特征：E='USD' 且 A（序号）、B（品名）为空（明细行 A/B 非空）。
    兼容单行式（东京A 的 C/D/E/F/G/H 合一）与两行式（JPY 行 + USD 行）。
    """
    total = 0.0
    files = sorted(glob.glob(os.path.join(SAMPLES_DIR, f"报关{cn_port}*.xls")))
    assert files, f"未找到港口 {cn_port} 的样本文件"
    for f in files:
        sh = xlrd.open_workbook(f).sheet_by_index(0)
        for r in range(sh.nrows):
            a = sh.cell_value(r, 0)
            b = sh.cell_value(r, 1)
            e = sh.cell_value(r, 4)
            fv = sh.cell_value(r, 5)
            if e == "USD" and a in ("", None) and b in ("", None) and fv not in ("", None):
                total += fv
    return total


# ---- Tests ----

class TestTicketConservation:
    """票级守恒：明细合计 == 票内源行合计（容差 0.01）。"""

    def test_amount_net_gross_cartons_pieces(self, ticket_results, groups):
        group_members = {g["source_name_cn"]: len(g["members"]) for g in groups}
        for t, rows, res in ticket_results:
            # 金额：set_split 末行残差兜底 / box_share 等分，均严格守恒
            src_amount = sum(r.amount or 0.0 for r in rows)
            det_amount = sum(x.amount or 0.0 for x in res.rows)
            assert abs(det_amount - src_amount) < 0.01, (
                f"{t.ticket_no} 金额守恒失败: {det_amount} vs {src_amount}"
            )
            # 净重/毛重/箱数：组行仅首行计一次（其余 None 不参与求和）
            for label, src_f, det_f in [
                ("净重", lambda r: r.net_weight, lambda x: x.net),
                ("毛重", lambda r: r.gross_weight, lambda x: x.gross),
                ("箱数", lambda r: r.pcs, lambda x: x.cartons),
            ]:
                src = sum(src_f(r) or 0.0 for r in rows)
                det = sum(det_f(x) or 0.0 for x in res.rows)
                assert abs(det - src) < 0.01, (
                    f"{t.ticket_no} {label}守恒失败: {det} vs {src}"
                )
            # 件数：组行每个组件都计套数/件数 → 每组源行多计 (成员数-1) 倍
            src_pieces = sum(r.qty_pieces or 0 for r in rows)
            extra = sum(
                (r.qty_pieces or 0) * (group_members[r.name_cn] - 1)
                for r in rows
                if r.name_cn in group_members
            )
            det_pieces = sum(x.pieces or 0.0 for x in res.rows)
            assert abs(det_pieces - (src_pieces + extra)) < 0.01, (
                f"{t.ticket_no} 件数守恒失败: {det_pieces} vs {src_pieces + extra}"
            )

    def test_no_mapping_warnings(self, ticket_results):
        """本批次 33 个中文品名（含组组件名）全部命中 37 条映射。"""
        for t, rows, res in ticket_results:
            assert res.warnings == [], f"{t.ticket_no} 映射警告: {res.warnings}"


class TestPortReconciliation:
    """港口级对账：生成的 USD 总额与人工样本一致（容差 0.01）。"""

    def test_port_usd_totals(self, ticket_results):
        port_usd: dict[str, float] = defaultdict(float)
        for t, rows, res in ticket_results:
            port_usd[t.port] += sum(x.amount or 0.0 for x in res.rows)

        assert abs(port_usd["東京港"] - TOKYO_USD_KNOWN) < 0.01, (
            f"東京港 {port_usd['東京港']} != 已知 {TOKYO_USD_KNOWN}"
        )
        if not Path(SAMPLES_DIR).exists():
            pytest.skip(f"样本目录不存在: {SAMPLES_DIR}")
        for port, info in PORT_MAP.items():
            expected = _sample_port_usd(info["cn"])
            assert abs(port_usd[port] - expected) < 0.01, (
                f"{port} 生成 {port_usd[port]} != 样本 {expected}"
            )


class TestSetSplit:
    """6件套 set_split：1 行 → 6 组件行。"""

    def test_six_piece_set(self, ticket_results, groups):
        g6 = next(g for g in groups if g["name"] == "6件套")
        members = [m["product_name_cn"] for m in g6["members"]]  # 已按 display_order
        target = None
        for t, rows, res in ticket_results:
            if any(r.name_cn == "6件套" for r in rows):
                target = (t, rows, res)
                break
        assert target is not None, "没有任何票包含 6件套"
        t, rows, res = target

        block = [x for x in res.rows if x.name_cn in members]
        assert len(block) == 6, f"{t.ticket_no} 6件套应拆为 6 行，实际 {len(block)}"
        assert [x.name_cn for x in block] == members, "组件行顺序应为 display_order"
        # 组件行在明细最前面
        assert [x.name_cn for x in res.rows[:6]] == members

        sets = sum(r.qty_pieces or 0 for r in rows if r.name_cn == "6件套")
        src = {
            "cartons": sum(r.pcs or 0 for r in rows if r.name_cn == "6件套"),
            "net": sum(r.net_weight or 0.0 for r in rows if r.name_cn == "6件套"),
            "gross": sum(r.gross_weight or 0.0 for r in rows if r.name_cn == "6件套"),
        }
        prices = {m["product_name_cn"]: m["split_price"] for m in g6["members"]}
        for i, x in enumerate(block):
            assert x.pieces == sets, f"{x.name_cn} pieces 应等于套数 {sets}"
            # 本批次 6件套单价数据自洽，末行残差恰等于 split_price×套数
            assert abs(x.amount - prices[x.name_cn] * sets) < 0.01, (
                f"{x.name_cn} amount {x.amount} != {prices[x.name_cn]}×{sets}"
            )
            if i == 0:
                assert x.cartons == src["cartons"]
                assert abs(x.net - src["net"]) < 0.01
                assert abs(x.gross - src["gross"]) < 0.01
            else:
                assert x.cartons is None and x.net is None and x.gross is None, (
                    f"{x.name_cn} 非首组件行的箱数/净重/毛重应留空"
                )
        # 组套小计（供模版 F16/G16）
        assert res.set_subtotal is not None
        sub_amount, sub_net = res.set_subtotal
        src_amount = sum(r.amount or 0.0 for r in rows if r.name_cn == "6件套")
        assert abs(sub_amount - src_amount) < 0.01
        assert abs(sub_net - src["net"]) < 0.01


class TestBoxShare:
    """烟灰缸+支架 box_share：1 行 → 2 行，金额等分，净重组件行留空。"""

    def test_ashtray_split(self, ticket_results):
        target = None
        for t, rows, res in ticket_results:
            if any(r.name_cn == "烟灰缸" for r in rows):
                target = (t, rows, res)
                break
        assert target is not None, "没有任何票包含 烟灰缸"
        t, rows, res = target

        block = [x for x in res.rows if x.name_cn in ("烟灰缸", "烟灰缸支架")]
        assert len(block) == 2, f"{t.ticket_no} 烟灰缸应拆为 2 行，实际 {len(block)}"
        assert block[0].name_cn == "烟灰缸"
        assert block[1].name_cn == "烟灰缸支架"

        src = [r for r in rows if r.name_cn == "烟灰缸"]
        total_amount = sum(r.amount or 0.0 for r in src)
        total_pieces = sum(r.qty_pieces or 0 for r in src)
        # 金额各半
        assert abs(block[0].amount - total_amount / 2) < 0.01
        assert abs(block[1].amount - total_amount / 2) < 0.01
        # 每行 pieces=源行件数
        assert block[0].pieces == total_pieces
        assert block[1].pieces == total_pieces
        # 箱数/净重/毛重仅首行，第二行留空
        assert block[0].cartons is not None
        assert block[0].net is not None
        assert block[0].gross is not None
        assert block[1].cartons is None
        assert block[1].net is None
        assert block[1].gross is None


class TestSorting:
    """排序：组组件在前固定序；普通行拼音序，ASCII 开头品名在汉字前。"""

    def test_normal_rows_pinyin_order(self, ticket_results, groups):
        group_sources = {g["source_name_cn"] for g in groups}
        member_names = {
            m["product_name_cn"] for g in groups for m in g["members"]
        }
        checked = 0
        for t, rows, res in ticket_results:
            normal_names = [
                r.name_cn for r in rows if r.name_cn not in group_sources
            ]
            if not normal_names:
                continue
            checked += 1
            n_normal = len(set(normal_names))
            tail = res.rows[len(res.rows) - n_normal:]
            tail_names = [x.name_cn for x in tail]
            # 尾部恰好是普通行（多重集一致），且按排序键升序
            assert Counter(tail_names) == Counter(set(normal_names)), (
                f"{t.ticket_no} 普通行集合不符: {tail_names}"
            )
            keys = [normal_row_sort_key(n) for n in tail_names]
            assert keys == sorted(keys), f"{t.ticket_no} 普通行未按拼音序: {tail_names}"
        assert checked > 0

    def test_ascii_names_before_hanzi(self, ticket_results, groups):
        """'44木架' 类数字开头品名排在汉字开头品名之前。"""
        group_sources = {g["source_name_cn"] for g in groups}
        found = False
        for t, rows, res in ticket_results:
            normal_names = sorted(
                {r.name_cn for r in rows if r.name_cn not in group_sources}
            )
            if "44木架" not in normal_names:
                continue
            hanzi = [n for n in normal_names if n and ord(n[0]) >= 128]
            if not hanzi:
                continue
            found = True
            tail_names = [x.name_cn for x in res.rows[len(res.rows) - len(normal_names):]]
            idx_44 = tail_names.index("44木架")
            first_hanzi_idx = min(tail_names.index(n) for n in hanzi)
            assert idx_44 < first_hanzi_idx, (
                f"{t.ticket_no} '44木架'(第{idx_44}普通行) 应排在汉字品名"
                f"(第{first_hanzi_idx}普通行) 之前: {tail_names}"
            )
        assert found, "没有找到含 44木架 与其他汉字品名共存的票"


class TestNaming:
    def test_ticket_letter(self):
        assert ticket_letter(0) == "A"
        assert ticket_letter(1) == "B"
        assert ticket_letter(25) == "Z"
        with pytest.raises(ValueError):
            ticket_letter(26)
        with pytest.raises(ValueError):
            ticket_letter(-1)

    def test_ticket_title(self):
        assert ticket_title("東京港", 0) == "东京A票"
        assert ticket_title("名古屋港", 2) == "名古屋C票"

    def test_declaration_filename(self):
        assert declaration_filename("神戸港", 1) == "报关神户B.xlsx"
        assert declaration_filename("横浜港", 0) == "报关横滨A.xlsx"

    def test_format_onboard(self):
        assert format_onboard("20260725") == "2026.7.25"
        assert format_onboard("20261105") == "2026.11.5"
        with pytest.raises(ValueError):
            format_onboard("2026-07")

    def test_unknown_port(self):
        with pytest.raises(ValueError):
            ticket_title("大阪港", 0)


class TestFullCoverage:
    """全量行归属：所有源行被且仅被一票覆盖。"""

    def test_every_row_exactly_once(self, ticket_results, raw_items):
        cnt: Counter[int] = Counter()
        for t, rows, res in ticket_results:
            for r in rows:
                cnt[id(r)] += 1
        missing = [r for r in raw_items if cnt[id(r)] == 0]
        dup = [k for k, v in cnt.items() if v > 1]
        assert not missing, (
            f"{len(missing)} 行未被任何票覆盖: "
            f"{[(r.kanri_no, r.maker, r.sku) for r in missing[:5]]}"
        )
        assert not dup, f"{len(dup)} 行被多票重复覆盖"


class TestTemplateFiller:
    """模版填充冒烟测试：票头、明细、两行式汇总。"""

    def test_fill_and_readback(self, ticket_results, tmp_path):
        template = TEMPLATE if Path(TEMPLATE).exists() else _WORKTREE_TEMPLATE
        if not Path(template).exists():
            pytest.skip(f"模版不存在: {template}")
        t, rows, res = ticket_results[0]
        out = tmp_path / "out.xlsx"
        fill_declaration(
            template, str(out),
            ticket_name=ticket_title(t.port, 0),
            invoice_no="YILT656",
            onboard=format_onboard("20260725"),
            port_to_en=PORT_MAP[t.port]["en"],
            set_subtotal=res.set_subtotal,
            rows=res.rows,
        )
        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["J1"].value == ticket_title(t.port, 0)
        assert ws["A2"].value == "INVOICE NO:YILT656"
        assert ws["G12"].value == "2026.7.25"
        assert ws["H13"].value == "QINGDAO"
        assert ws["H14"].value == PORT_MAP[t.port]["en"]
        if res.set_subtotal is not None:
            assert abs(ws["F16"].value - res.set_subtotal[0]) < 0.01
            assert abs(ws["G16"].value - res.set_subtotal[1]) < 0.01

        # 明细从 18 行起
        first = res.rows[0]
        assert ws.cell(row=18, column=1).value == 1
        assert ws.cell(row=18, column=2).value == first.name_cn
        n = len(res.rows)
        # 明细后空 1 行 → JPY 行 / USD 行
        jpy_row = 18 + n + 1
        usd_row = jpy_row + 1
        assert ws.cell(row=jpy_row, column=1).value is None  # 空行确认
        assert ws.cell(row=18 + n, column=1).value is None
        assert ws.cell(row=jpy_row, column=5).value == "JPY"
        assert ws.cell(row=usd_row, column=5).value == "USD"
        total_amount = sum(x.amount or 0.0 for x in res.rows)
        total_cartons = sum(x.cartons or 0.0 for x in res.rows)
        assert abs(ws.cell(row=usd_row, column=6).value - total_amount) < 0.01
        assert abs(ws.cell(row=jpy_row, column=3).value - total_cartons) < 0.01
        wb.close()
