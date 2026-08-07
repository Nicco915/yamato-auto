# -*- coding: utf-8 -*-
"""报关单生成服务——按已确认的分票方案批量生成报关 Excel。

流程（全部为纯 Python 计算，零 LLM）：
1. 读 DB：confirmed Declaration（按港口分组、票内按 ticket_no 排序）、
   product_mappings → 映射索引、product_groups + members → 品名组配置；
2. 读 split 图 state 拿 source_file_path → load_filled_excel → 归一化 + 商检判定；
3. 每票 rows_for_ticket → aggregate_ticket → fill_declaration 写文件；
4. 输出 output/{batch_id}/declarations/{split_thread_id}/，重跑先清空旧 xlsx（幂等）。
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from app.config import get_settings
from app.db.models import (
    Declaration,
    ProductGroup,
    ProductGroupMember,
    ProductMapping,
)
from app.db.session import get_session
from app.declare.aggregator import aggregate_ticket, rows_for_ticket
from app.declare.mapping import build_mapping_index
from app.factory_match import load_excel_normalize_map, load_inspection_factories
from app.declare.naming import (
    PORT_MAP,
    declaration_filename,
    format_onboard,
    ticket_title,
)
from app.declare.template_filler import fill_declaration
from app.split.loader import load_filled_excel
from app.split.normalize import classify_sj_factories, normalize_maker
from app.split.schemas import Ticket, TicketItem

logger = logging.getLogger(__name__)

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1] / "templates" / "declaration_template.xlsx"
)


def declarations_dir(split_thread_id: str) -> Path:
    """该分票任务的报关单输出目录（output/{batch_id}/declarations/{split_thread_id}/）。

    batch_id 从 split_thread_id 反推：去掉 "split-" 前缀（约定 split-{父批次}）。
    """
    batch_id = split_thread_id.removeprefix("split-")
    return get_settings().batch_declarations_dir(batch_id) / split_thread_id


def _source_file_from_graph(split_thread_id: str) -> str:
    """从 split 图 checkpoint state 拿 source_file_path。"""
    from app.split.graph import get_split_graph  # 延迟导入，避免 split↔declare 循环

    snap = get_split_graph().get_state(
        {"configurable": {"thread_id": split_thread_id}}
    )
    source = (snap.values or {}).get("source_file_path", "")
    if not source:
        raise ValueError(f"分票任务 {split_thread_id} 无 source_file_path（图未启动？）")
    return source


def _read_etd(source_path: str) -> str:
    """从 filled Excel 直接按列名读 ETD_YMD_T 第一行非空值（全批次同船期）。"""
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = [
            str(c.value).strip() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        if "ETD_YMD_T" not in header:
            raise ValueError("filled Excel 缺少 ETD_YMD_T 列，无法确定开船日期")
        idx = header.index("ETD_YMD_T")
        for row in ws.iter_rows(min_row=2):
            v = row[idx].value
            if v is not None and str(v).strip():
                return str(v).strip()
    finally:
        wb.close()
    raise ValueError("filled Excel 的 ETD_YMD_T 列全为空，无法确定开船日期")


def generate_declarations(split_thread_id: str, invoice_number: str) -> dict:
    """按已确认的分票方案生成全部报关单。

    Args:
        split_thread_id: 分票图 thread_id（split- 前缀）。
        invoice_number: 人工输入的发票号码段，如 '656'。

    Returns:
        {"generated": [文件名...], "count": int, "warnings": [...], "out_dir": str}

    Raises:
        ValueError: 无 confirmed Declaration（方案未确认）等输入问题。
    """
    # ---- 1. 读 confirmed Declaration，按港口分组、票内按 ticket_no 排序 ----
    with get_session() as sess:
        decls = (
            sess.query(Declaration)
            .filter(
                Declaration.split_thread_id == split_thread_id,
                Declaration.status == "confirmed",
            )
            .order_by(Declaration.port, Declaration.ticket_no)
            .all()
        )
        if not decls:
            raise ValueError("分票方案尚未确认")

        # product_mappings → 映射索引（expire_on_commit=False，ORM 可直接用）
        mapping_index = build_mapping_index(sess.query(ProductMapping).all())

        # product_groups + members → aggregator 需要的组配置 dict
        groups: list[dict] = []
        for g in sess.query(ProductGroup).all():
            members = (
                sess.query(ProductGroupMember)
                .filter(ProductGroupMember.group_id == g.id)
                .order_by(ProductGroupMember.display_order)
                .all()
            )
            groups.append({
                "name": g.name,
                "group_type": g.group_type,
                "source_name_cn": g.source_name_cn,
                "members": [
                    {
                        "product_name_cn": m.product_name_cn,
                        "display_order": m.display_order,
                        "split_price": m.split_price,
                        "split_net_weight": m.split_net_weight,
                    }
                    for m in members
                ],
            })

    # ---- 2. 读 filled Excel + 归一化 + 商检判定 + ETD ----
    source_file = _source_file_from_graph(split_thread_id)
    raw_items = load_filled_excel(source_file)
    if not raw_items:
        raise ValueError(f"filled Excel 无数据行: {source_file}")
    # 归一化映射与商检名单：DB 优先，config 兜底
    normalize_map = load_excel_normalize_map()
    for r in raw_items:
        r.maker = normalize_maker(r.maker, normalize_map)
    sj_map = classify_sj_factories(raw_items, {}, load_inspection_factories())
    onboard = format_onboard(_read_etd(source_file))

    # ---- 3. 输出目录：幂等——先清空旧 xlsx 再生成 ----
    out_dir = declarations_dir(split_thread_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    for old in out_dir.glob("*.xlsx"):
        try:
            old.unlink()
        except OSError as e:
            logger.warning("清理旧报关单文件失败 %s: %s", old.name, e)

    # ---- 4. 逐票生成（港口内按 ticket_no 顺序编字母 A/B/...） ----
    generated: list[str] = []
    all_warnings: list[str] = []

    by_port: dict[str, list[Declaration]] = {}
    for d in decls:
        by_port.setdefault(d.port, []).append(d)

    for port, port_decls in by_port.items():
        if port not in PORT_MAP:
            raise ValueError(f"未知港口：{port!r}")
        invoice_no = f"YIL{PORT_MAP[port]['inv']}{invoice_number}"
        for idx, d in enumerate(port_decls):
            ticket = Ticket(
                ticket_no=d.ticket_no,
                port=d.port,
                container_type=d.container_type,
                items=[TicketItem(**it) for it in d.items],
                sj_factories=[
                    f.get("factory_name", "") if isinstance(f, dict) else str(f)
                    for f in (d.sj_factories or [])
                ],
            )
            rows = rows_for_ticket(ticket, raw_items, sj_map)
            res = aggregate_ticket(rows, groups, mapping_index)

            fname = declaration_filename(port, idx)
            fill_declaration(
                str(TEMPLATE_PATH),
                str(out_dir / fname),
                ticket_name=ticket_title(port, idx),
                invoice_no=invoice_no,
                onboard=onboard,
                port_to_en=PORT_MAP[port]["en"],
                set_subtotal=res.set_subtotal,
                rows=res.rows,
            )
            generated.append(fname)
            for w in res.warnings:
                all_warnings.append(f"{d.ticket_no}: {w}")

    logger.info(
        "generate_declarations: split_thread_id=%s 生成 %d 票 → %s（%d 条警告）",
        split_thread_id, len(generated), out_dir, len(all_warnings),
    )
    return {
        "generated": generated,
        "count": len(generated),
        "warnings": all_warnings,
        "out_dir": str(out_dir),
    }
