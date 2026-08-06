# -*- coding: utf-8 -*-
"""产品映射 Excel 导入（供一次性脚本与 /api/v1/mappings 上传端点共用）。

Excel 结构（表头第 1 行）：产品 | 税号 | 供应商 | 商检 | 产品组一 | 自定义七
- 按 (产品, 供应商) 去重；幂等 upsert
- 商检列：'商检' → True，空/'(空白)' → False
"""
from __future__ import annotations

import io
from pathlib import Path

import openpyxl

from app.db.models import ProductMapping


def _norm(value):
    """单元格归一：None/空白/'(空白)' → None，其余 strip 后转 str。"""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "(空白)":
        return None
    return text


def parse_mapping_rows(source: str | Path | bytes | io.BytesIO) -> list[dict]:
    """从 xlsx 的 Sheet1（或首个 sheet）解析映射行，按 (产品, 供应商) 去重。

    source 可为路径或字节流（API 上传）。
    """
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    seen: dict[tuple[str, str | None], dict] = {}
    for row in rows[1:]:  # 跳过表头
        product, hs_code, supplier, inspection, name_en, unit_code = (
            list(row) + [None] * 6
        )[:6]
        product = _norm(product)
        if not product:
            continue
        supplier = _norm(supplier)
        key = (product, supplier)
        if key in seen:
            continue
        seen[key] = {
            "product_name_cn": product,
            "hs_code": _norm(hs_code),
            "supplier_name": supplier,
            "inspection_required": _norm(inspection) == "商检",
            "name_en": _norm(name_en),
            "unit_code": _norm(unit_code),
        }
    return list(seen.values())


def upsert_mappings(session, rows: list[dict]) -> tuple[int, int]:
    """按 (product_name_cn, supplier_name) upsert。返回 (新增数, 更新数)。"""
    created = updated = 0
    for data in rows:
        existing = (
            session.query(ProductMapping)
            .filter(
                ProductMapping.product_name_cn == data["product_name_cn"],
                ProductMapping.supplier_name.is_(None)
                if data["supplier_name"] is None
                else ProductMapping.supplier_name == data["supplier_name"],
            )
            .first()
        )
        if existing:
            existing.hs_code = data["hs_code"]
            existing.inspection_required = data["inspection_required"]
            existing.name_en = data["name_en"]
            existing.unit_code = data["unit_code"]
            updated += 1
        else:
            session.add(ProductMapping(**data))
            created += 1
    return created, updated
