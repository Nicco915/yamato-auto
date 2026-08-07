# -*- coding: utf-8 -*-
"""报关单模版填充——复制模版后写入票头、明细、汇总，纯 openpyxl 操作。"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Side

from app.declare.aggregator import DetailRow

DETAIL_START_ROW = 18  # 明细从第 18 行起（第 17 行为表头）

# 写入单元格格式：四周细边框（2026-08-04 用户定，与 writer.py 的 WRITE_BORDER 一致）
_THIN_SIDE = Side(style="thin")
WRITE_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE, bottom=_THIN_SIDE)


def _apply_write_border(cell) -> None:
    """写入单元格加四周细边框；字体/填充保留原设置。"""
    cell.border = WRITE_BORDER


def fill_declaration(
    template_path: str,
    out_path: str,
    *,
    ticket_name: str,        # '东京A票'
    invoice_no: str,         # 'YILT656'（调用方拼好）
    onboard: str,            # '2026.7.25'
    port_to_en: str,         # 'TOKYO'
    set_subtotal: tuple[float, float] | None,  # (金额, 净重) 或 None
    rows: list[DetailRow],
) -> None:
    """填充报关单模版并另存为 out_path。

    版式（对照人工样本）：
      J1=票名、A2='INVOICE NO:xxx'、G12=开船日、H13='QINGDAO'、H14=目的港英文
      F16/G16=组套小计（仅 set_subtotal 非 None 时填）
      18 行起逐行明细：A=seq B=品名 C=箱数 D=件数 E=币制 F=金额 G=净重 H=毛重
                       I='商检'（inspection 时） J=计量单位代码；None 的格子留空
      明细后空 1 行，再两行式汇总：
        JPY 行：C=总箱数 D=总件数 E='JPY' G=总净重 H=总毛重
        USD 行：E='USD' F=总金额
    """
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    ws["J1"] = ticket_name
    ws["A2"] = f"INVOICE NO:{invoice_no}"
    ws["G12"] = onboard
    ws["H13"] = "QINGDAO"
    ws["H14"] = port_to_en

    if set_subtotal is not None:
        ws["F16"] = set_subtotal[0]
        ws["G16"] = set_subtotal[1]

    # TODO: 明细行/汇总行的字体可后续复制第 17 行表头样式；边框已统一细线
    r = DETAIL_START_ROW
    for row in rows:
        c = ws.cell(row=r, column=1, value=row.seq);        _apply_write_border(c)
        c = ws.cell(row=r, column=2, value=row.name_cn);    _apply_write_border(c)
        if row.cartons is not None:
            c = ws.cell(row=r, column=3, value=row.cartons); _apply_write_border(c)
        if row.pieces is not None:
            c = ws.cell(row=r, column=4, value=row.pieces);  _apply_write_border(c)
        c = ws.cell(row=r, column=5, value=row.currency or "USD"); _apply_write_border(c)
        if row.amount is not None:
            c = ws.cell(row=r, column=6, value=row.amount); _apply_write_border(c)
        if row.net is not None:
            c = ws.cell(row=r, column=7, value=row.net);    _apply_write_border(c)
        if row.gross is not None:
            c = ws.cell(row=r, column=8, value=row.gross);  _apply_write_border(c)
        if row.inspection:
            c = ws.cell(row=r, column=9, value="商检");      _apply_write_border(c)
        if row.unit_code:
            c = ws.cell(row=r, column=10, value=row.unit_code); _apply_write_border(c)
        r += 1

    # 空 1 行后写两行式汇总；总箱数/净重/毛重 = 各行非 None 值求和
    r += 1
    total_cartons = sum(x.cartons for x in rows if x.cartons is not None)
    total_pieces = sum(x.pieces for x in rows if x.pieces is not None)
    total_amount = sum(x.amount for x in rows if x.amount is not None)
    total_net = sum(x.net for x in rows if x.net is not None)
    total_gross = sum(x.gross for x in rows if x.gross is not None)

    # JPY 行（无金额列）
    c = ws.cell(row=r, column=3, value=total_cartons); _apply_write_border(c)
    c = ws.cell(row=r, column=4, value=total_pieces);  _apply_write_border(c)
    c = ws.cell(row=r, column=5, value="JPY");         _apply_write_border(c)
    c = ws.cell(row=r, column=7, value=total_net);     _apply_write_border(c)
    c = ws.cell(row=r, column=8, value=total_gross);   _apply_write_border(c)
    r += 1
    # USD 行（仅金额）
    c = ws.cell(row=r, column=5, value="USD");         _apply_write_border(c)
    c = ws.cell(row=r, column=6, value=total_amount);  _apply_write_border(c)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
