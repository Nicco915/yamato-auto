# -*- coding: utf-8 -*-
"""「跳过本工厂」功能回归测试（Skipped 状态 + factory_skipped 审计 + 详情页徽章）。

覆盖：
1. Node5 单测级：resume {"skipped": true} 不做 items 合并，
   calculated_items 原样保留，validation_status=Skipped（monkeypatch interrupt）；
2. Node5 回归：不带 skipped 的正常 approve 路径不受影响；
3. 端到端（mock 提取跑真实图）：两工厂批次，A approve、B 跳过 →
   批次 success；Excel 只写 A 的行（B 行 净重/毛重 留空）；
   factory_outputs 只有 A；review_audits 有 B 的 factory_skipped 行
   （approved=false），A 为正常 approved 行；
4. get_batch_detail：B 的 role=="skipped"（灰色徽章数据源），A 为 done。

隔离（血泪红线 2026-08-11，与 tests/logging_context_test.py 同模式）：
checkpoint/master db、output、sessions 全部指向临时目录——import 全部
app 模块后调 validation/_test_isolation.isolate_to_tmp（llm_client 的
load_dotenv override 已在 import 时执行完毕，此刻重设 env + 清缓存才有效，
且带真实库路径断言守卫）。全程内存级跑图，不起子进程。

用法（在 app/ 目录下）：
  python3 tests/factory_skip_test.py
  PYTHONPATH=. python3 -m pytest tests/factory_skip_test.py -q
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

# ---- env 前置（EXTRACTION_MOCK 需在 import app 之前；db 路径在 import 后隔离）----
os.environ["EXTRACTION_MOCK"] = "1"                      # 提取走 mock，不调 LLM
os.environ["DISPATCHER_MOCK"] = "1"

APP_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_ROOT))
sys.path.insert(0, str(APP_ROOT / "validation"))

from openpyxl import Workbook, load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.api import service  # noqa: E402
from app.db.models import ReviewAudit  # noqa: E402
from app.db.session import get_session  # noqa: E402
from app.nodes import extraction_node as en  # noqa: E402
from app.nodes import human_review as hr  # noqa: E402

from _test_isolation import isolate_to_tmp  # noqa: E402

# 隔离必须在 import app 模块之后（load_dotenv override 红线）；
# graph/engine 都是惰性单例，首次调用才按隔离后的 settings 建临时库
TMP = isolate_to_tmp("yamato_factory_skip_test_")

# ---- 测试夹具：两个工厂的最小下游装箱单 + 上游空文件夹 ----
F_A, F_B = "跳过厂甲", "跳过厂乙"
SKU_A, SKU_B = "4900000000001", "4900000000002"
THREAD = "SKIP-E2E"


def _make_fixtures() -> tuple[str, str]:
    """生成最小下游 xlsx（两个工厂各 1 SKU）与上游工厂文件夹，返回路径。"""
    xlsx = TMP / "downstream.xlsx"
    wb = Workbook()
    ws = wb.active
    # Node6 首次写入时会在 SHOHIN_MEI_E 后插入 中文品名/净重/毛重 三列
    ws.append(["MAKER_MEI_KJ", "SHOHIN_CD", "SHOHIN_MEI_E", "SOTOBAKO_D_HACCHU_SU"])
    ws.append([F_A, SKU_A, "ITEM-A", 10])
    ws.append([F_B, SKU_B, "ITEM-B", 20])
    wb.save(xlsx)

    upstream = TMP / "upstream"
    (upstream / F_A).mkdir(parents=True)
    (upstream / F_B).mkdir(parents=True)
    return str(xlsx), str(upstream)


def _audit_rows(thread_id: str) -> list[ReviewAudit]:
    with get_session() as session:
        return list(session.scalars(
            select(ReviewAudit)
            .where(ReviewAudit.thread_id == thread_id)
            .order_by(ReviewAudit.audit_id)
        ).all())


# ---------------------------------------------------------------------------
# 1. Node5 单测级：skipped 不做 items 合并
# ---------------------------------------------------------------------------

def test_human_review_skipped_no_merge():
    """resume {"skipped": true}：不合并 items，calculated_items 原样保留。"""
    original_items = [
        {"sku": SKU_A,
         "extracted_data": {"total_quantity": 50, "total_net_weight": 250.0,
                            "total_gross_weight": 265.0},
         "calculation": {"calculated_unit_net": 5.0, "calculated_unit_gross": 5.3},
         "status": "Normal"},
    ]
    state = {"current_factory_data": {"factory_name": "单测跳过厂",
                                      "calculated_items": original_items}}

    orig_interrupt = hr.interrupt
    # 即便人工提交里带了 items，跳过分支也必须原样忽略（不做任何合并）
    hr.interrupt = lambda payload: {  # noqa: E731
        "approved": False, "skipped": True,
        "items": [{"sku": "9999999999999",
                   "extracted_data": {"total_quantity": 1}}],
    }
    try:
        out = hr.human_review(state)
    finally:
        hr.interrupt = orig_interrupt

    assert out["validation_status"] == "Skipped", out
    merged = out["current_factory_data"]["calculated_items"]
    assert merged == original_items, "跳过分支不得改动 calculated_items"
    assert len(merged) == 1 and merged[0]["sku"] == SKU_A
    print("[断言通过] Node5 skipped：items 不合并，calculated_items 原样保留")


def test_human_review_normal_approve_regression():
    """回归：不带 skipped 的正常 approve 仍走合并并置 Approved。"""
    original_items = [
        {"sku": SKU_A,
         "extracted_data": {"total_quantity": 50, "total_net_weight": 250.0,
                            "total_gross_weight": 265.0},
         "status": "Normal"},
    ]
    state = {"current_factory_data": {"factory_name": "单测批准厂",
                                      "calculated_items": original_items}}

    orig_interrupt = hr.interrupt
    hr.interrupt = lambda payload: {"approved": True, "items": []}  # noqa: E731
    try:
        out = hr.human_review(state)
    finally:
        hr.interrupt = orig_interrupt

    assert out["validation_status"] == "Approved", out
    # items=[] 时人工未返回的 SKU 原样保留（既有合并语义）
    assert out["current_factory_data"]["calculated_items"] == original_items
    print("[断言通过] Node5 回归：approve 路径不受 skipped 分支影响")


# ---------------------------------------------------------------------------
# 2. 端到端：A approve + B 跳过（mock 提取跑真实图 Node1→Node7）
# ---------------------------------------------------------------------------

def test_e2e_approve_a_skip_b(monkeypatch):
    # 全量 pytest 下，先收集的模块可能在本文件设置 EXTRACTION_MOCK 之前就
    # import 了 extraction_node（_session_mod 已绑定真实提取线，走真实提取会
    # 拿到空结果 → 占位数据，与本测试前提不符）。monkeypatch 强制回 mock，
    # 测试结束自动恢复，不影响同进程其他测试
    monkeypatch.setattr(en, "_session_mod", None)
    monkeypatch.setattr(en, "_session_import_error", "EXTRACTION_MOCK=1（测试强制）")

    xlsx, upstream = _make_fixtures()

    # ---- 第一程：run 到 A 挂起 ----
    r1 = service.run_until_interrupt(
        THREAD, downstream_file_path=xlsx, upstream_root=upstream)
    assert r1["status"] == "pending_human_review", f"未挂起: {r1}"
    assert r1["review_data"]["factory_name"] == F_A

    # ---- 第二程：A approve → B 挂起 ----
    r2 = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r2["status"] == "pending_human_review", f"未二次挂起: {r2}"
    assert r2["review_data"]["factory_name"] == F_B

    # ---- 第三程：B 跳过 → 批次跑完 ----
    r3 = service.resume_order(
        THREAD, {"approved": False, "skipped": True, "items": []})
    assert r3["status"] == "success", f"批次未正常完成: {r3}"
    assert r3["final_validation_status"] == "Skipped", r3

    # factory_outputs 只有 A（Skipped 不进快照——后续「补充工厂」依此判定）
    state = service.get_order_state(THREAD)
    outputs = state["values"].get("factory_outputs") or {}
    assert F_A in outputs, f"A 应进 factory_outputs: {list(outputs)}"
    assert F_B not in outputs, f"B 跳过不得进 factory_outputs: {list(outputs)}"
    print("[断言通过] 端到端：批次 success，factory_outputs 只含 A")

    # Excel 只写 A 的行：B 行的 净重/毛重 留空（列由 A 的首次写入插入）
    out_path = r3["final_output_path"]
    assert out_path, r3
    wb = load_workbook(out_path)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    col_sku = header.index("SHOHIN_CD") + 1
    col_net = header.index("净重") + 1
    col_gross = header.index("毛重") + 1
    row_of = {}
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_sku).value
        if v is not None:
            row_of[str(v)] = row
    row_a, row_b = row_of[SKU_A], row_of[SKU_B]
    # mock 提取：单件净重 5.0 / 毛重 5.3，A 行 SOTOBAKO=10 → 50.0 / 53.0
    assert ws.cell(row=row_a, column=col_net).value == 50.0, \
        f"A 行净重应为 50.0: {ws.cell(row=row_a, column=col_net).value}"
    assert ws.cell(row=row_a, column=col_gross).value == 53.0
    assert ws.cell(row=row_b, column=col_net).value in (None, ""), \
        f"B 行净重必须留空: {ws.cell(row=row_b, column=col_net).value}"
    assert ws.cell(row=row_b, column=col_gross).value in (None, "")
    print("[断言通过] Excel：A 行已写入 50.0/53.0，B 行留空")

    # 审计留痕：A 为正常 approved 行；B 为 factory_skipped 且 approved=false
    audits = _audit_rows(THREAD)
    by_factory: dict[str, list[ReviewAudit]] = {}
    for r in audits:
        by_factory.setdefault(r.factory_name, []).append(r)
    a_rows = by_factory.get(F_A) or []
    b_rows = by_factory.get(F_B) or []
    assert a_rows, f"缺 A 的审计行: {audits}"
    assert a_rows[-1].approved is True
    assert a_rows[-1].result_status == "pending_human_review", \
        f"A 审计状态应保持原语义: {a_rows[-1].result_status}"
    assert b_rows, f"缺 B 的审计行: {audits}"
    assert b_rows[-1].approved is False, "跳过行 approved 必须为 false"
    assert b_rows[-1].result_status == "factory_skipped", \
        f"B 审计状态应为 factory_skipped: {b_rows[-1].result_status}"
    assert b_rows[-1].edited_count == 0
    print("[断言通过] 审计：B 落 factory_skipped(approved=false)，A 正常 approved")


def test_batch_detail_skipped_role():
    """get_batch_detail：B role=skipped（灰色徽章），A role=done。"""
    detail = service.get_batch_detail(THREAD)
    roles = {f["factory"]: f["role"] for f in detail["factories"]}
    assert roles.get(F_A) == "done", roles
    assert roles.get(F_B) == "skipped", roles
    # audit 区也带 result_status，供前端渲染「跳过」文案
    statuses = {(a["factory_name"], a["result_status"]) for a in detail["audit"]}
    assert (F_B, "factory_skipped") in statuses, statuses
    print("[断言通过] get_batch_detail：B=skipped / A=done，audit 带 factory_skipped")


def main():
    import pytest  # 脚本模式下提供 MonkeyPatch（pytest 运行时是注入的 fixture）
    test_human_review_skipped_no_merge()
    test_human_review_normal_approve_regression()
    with pytest.MonkeyPatch.context() as mp:
        test_e2e_approve_a_skip_b(mp)
    test_batch_detail_skipped_role()
    print("\nfactory_skip_test: PASS")


if __name__ == "__main__":
    main()
