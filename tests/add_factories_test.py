# -*- coding: utf-8 -*-
"""批次「补充工厂」功能测试（未处理工厂增量补入 + 两种起始状态续跑）。

覆盖：
1. 三工厂批次：A approve、B 跳过、C approve → completed；add_factories_to_batch
   → B 被拉回 pending 并续跑到 Node5 挂起；approve B → 再次 completed，
   Excel 含三家、factory_outputs 含三家、审计 B 先 factory_skipped 后 approved；
2. 装箱单尾部追加新工厂 D：两工厂批次跑完后改装箱单加一行 D →
   add_factories → D 进入处理并挂起；approve D → Excel D 行正常写入
   （验证输出副本行同步，不写幽灵数据）；
3. 无待补充：全部 approve 后调用 → added==0 + message，state 不动；
4. 已 approve 工厂不被重复处理（factory_outputs 键不重复、Excel 行不重复写）；
5. pending_review 状态调用：挂起中补充 → Node5 重挂起、当前工厂 payload 不丢，
   pending 已合并，审核链走完四家全部落 factory_outputs。

续跑语义实测结论（langgraph 1.2.9，本文件 e2e 即为验证）：
- completed（next 为空）：update_state(as_node=NODE6) 后条件边 _route_after_writer
  重新路由 → next=NODE2，stream(None) 正常续跑到 Node5 挂起；
- pending_review：update_state 无论 as_node 是什么都会销毁 Node5 的 interrupt
  任务（get_review_payload 返回 None），且销毁后直接 resume 会让当前工厂
  静默跳过写入（Node5 不再执行，validation_status 残留 Pending）——因此
  service 采用 as_node=NODE4 + stream(None) 立即重挂起，Node5 用未变的
  current_factory_data 重建等价 payload 再次 interrupt。

隔离（血泪红线 2026-08-11，与 tests/factory_skip_test.py 同模式）：
checkpoint/master db、output、sessions 全部指向临时目录——import 全部
app 模块后调 validation/_test_isolation.isolate_to_tmp。全程内存级跑图。

用法（在 app/ 目录下）：
  python3 tests/add_factories_test.py
  PYTHONPATH=. python3 -m pytest tests/add_factories_test.py -q
"""
from __future__ import annotations

import os
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---- env 前置（EXTRACTION_MOCK 需在 import app 之前；db 路径在 import 后隔离）----
os.environ["EXTRACTION_MOCK"] = "1"                      # 提取走 mock，不调 LLM
os.environ["DISPATCHER_MOCK"] = "1"

APP_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_ROOT))
sys.path.insert(0, str(APP_ROOT / "validation"))

from openpyxl import Workbook, load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.api import service  # noqa: E402
from app.db.models import ReviewAudit  # noqa: E402
from app.db.session import get_session  # noqa: E402
from app.nodes import extraction_node as en  # noqa: E402

from _test_isolation import isolate_to_tmp  # noqa: E402

# 隔离必须在 import app 模块之后（load_dotenv override 红线）；
# graph/engine 都是惰性单例，首次调用才按隔离后的 settings 建临时库
TMP = isolate_to_tmp("yamato_add_factories_test_")

# mock 提取口径：单件净重 5.0 / 毛重 5.3（见 extraction_node._mock_items）
MOCK_UNIT_NET, MOCK_UNIT_GROSS = 5.0, 5.3
HEADER = ["MAKER_MEI_KJ", "SHOHIN_CD", "SHOHIN_MEI_E", "SOTOBAKO_D_HACCHU_SU"]


def _write_manifest(path: Path, rows: list[tuple[str, str, str, int]]) -> str:
    """生成最小下游装箱单 xlsx，返回路径字符串。"""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for factory, sku, name, qty in rows:
        ws.append([factory, sku, name, qty])
    wb.save(path)
    return str(path)


def _make_upstream(tag: str, factories: list[str]) -> str:
    root = TMP / f"upstream_{tag}"
    for f in factories:
        (root / f).mkdir(parents=True, exist_ok=True)
    return str(root)


def _write_batch_config(thread: str, xlsx: str, upstream: str) -> None:
    """run_until_interrupt 不写 batch_config.json（那是 create_batch 的职责），
    测试手工补一份——add_factories_to_batch 靠它定位装箱单。"""
    now = datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    service._write_batch_config(thread, {
        "thread_id": thread,
        "downstream_file_path": xlsx,
        "upstream_root": upstream,
        "created_at": now,
        "last_run_at": now,
        "run_count": 1,
    })


def _force_mock_extraction(monkeypatch) -> None:
    """全量 pytest 下，先收集的模块可能在本文件设置 EXTRACTION_MOCK 之前就
    import 了 extraction_node（_session_mod 已绑定真实提取线）。monkeypatch
    强制回 mock，测试结束自动恢复（与 factory_skip_test 同模式）。"""
    monkeypatch.setattr(en, "_session_mod", None)
    monkeypatch.setattr(en, "_session_import_error", "EXTRACTION_MOCK=1（测试强制）")


def _audit_rows(thread_id: str) -> list[ReviewAudit]:
    with get_session() as session:
        return list(session.scalars(
            select(ReviewAudit)
            .where(ReviewAudit.thread_id == thread_id)
            .order_by(ReviewAudit.audit_id)
        ).all())


def _excel_rows_by_sku(out_path: str) -> dict[str, dict]:
    """读输出 Excel，返回 {SKU: {"net":…, "gross":…, "factory":…}}。"""
    wb = load_workbook(out_path)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    col_sku = header.index("SHOHIN_CD") + 1
    col_factory = header.index("MAKER_MEI_KJ") + 1
    col_net = header.index("净重") + 1
    col_gross = header.index("毛重") + 1
    rows = {}
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_sku).value
        if v is None:
            continue
        rows[str(v)] = {
            "factory": ws.cell(row=row, column=col_factory).value,
            "net": ws.cell(row=row, column=col_net).value,
            "gross": ws.cell(row=row, column=col_gross).value,
        }
    return rows


# ---------------------------------------------------------------------------
# 1. completed 批次：跳过工厂被拉回续跑（三工厂 A approve / B 跳过 / C approve）
# ---------------------------------------------------------------------------

def test_completed_batch_skip_then_add(monkeypatch):
    _force_mock_extraction(monkeypatch)
    F_A, F_B, F_C = "补入厂甲", "补入厂乙", "补入厂丙"
    SKU_A, SKU_B, SKU_C = "4900000000101", "4900000000102", "4900000000103"
    THREAD = "ADDF-E2E-SKIP"

    xlsx = _write_manifest(TMP / "addf_skip.xlsx", [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
        (F_C, SKU_C, "ITEM-C", 30),
    ])
    upstream = _make_upstream("skip", [F_A, F_B, F_C])
    _write_batch_config(THREAD, xlsx, upstream)

    r1 = service.run_until_interrupt(THREAD, downstream_file_path=xlsx,
                                     upstream_root=upstream)
    assert r1["status"] == "pending_human_review"
    assert r1["review_data"]["factory_name"] == F_A
    r2 = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r2["review_data"]["factory_name"] == F_B
    r3 = service.resume_order(THREAD, {"approved": False, "skipped": True, "items": []})
    assert r3["review_data"]["factory_name"] == F_C
    r4 = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r4["status"] == "success", r4

    # 批次已 completed（next 为空）
    assert not service.get_order_state(THREAD)["next_nodes"]

    # ---- 补充工厂：B（被跳过）应被拉回并续跑到 Node5 挂起 ----
    r5 = service.add_factories_to_batch(THREAD)
    assert r5["added"] == 1 and r5["factories"] == [F_B], r5
    assert r5["status"] == "pending_human_review", r5
    assert r5["review_data"]["factory_name"] == F_B, r5["review_data"]

    # approve B → 批次再次 completed
    r6 = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r6["status"] == "success", r6

    state = service.get_order_state(THREAD)
    outputs = state["values"].get("factory_outputs") or {}
    assert set(outputs) == {F_A, F_B, F_C}, sorted(outputs)

    # Excel 含三家：A/C 首次写入，B 补充后写入（净重 = 5.0 × 发注数）
    rows = _excel_rows_by_sku(r6["final_output_path"])
    assert rows[SKU_A]["net"] == 50.0 and rows[SKU_A]["gross"] == 53.0
    assert rows[SKU_B]["net"] == 100.0 and rows[SKU_B]["gross"] == 106.0
    assert rows[SKU_C]["net"] == 150.0 and rows[SKU_C]["gross"] == 159.0
    print("[断言通过] completed 批次补充跳过工厂：B 拉回→挂起→approve→三家齐写入")

    # 审计：B 先 factory_skipped，后追加 approved 行；批次详情 B 角色回到 done
    audits = [r for r in _audit_rows(THREAD) if r.factory_name == F_B]
    assert audits[0].result_status == "factory_skipped" and audits[0].approved is False
    assert audits[-1].approved is True
    detail = service.get_batch_detail(THREAD)
    roles = {f["factory"]: f["role"] for f in detail["factories"]}
    assert roles[F_B] == "done", roles
    print("[断言通过] 审计与详情页：B 由 skipped 回到 done（最新一条 approved）")


# ---------------------------------------------------------------------------
# 2. 装箱单尾部追加新工厂 D（completed 批次 + 输出副本行同步）
# ---------------------------------------------------------------------------

def test_add_new_factory_from_appended_manifest(monkeypatch):
    _force_mock_extraction(monkeypatch)
    F_A, F_B, F_D = "追加厂甲", "追加厂乙", "追加厂丁"
    SKU_A, SKU_B, SKU_D = "4900000000201", "4900000000202", "4900000000204"
    THREAD = "ADDF-E2E-NEW"

    xlsx_path = TMP / "addf_new.xlsx"
    xlsx = _write_manifest(xlsx_path, [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
    ])
    upstream = _make_upstream("new", [F_A, F_B, F_D])
    _write_batch_config(THREAD, xlsx, upstream)

    r = service.run_until_interrupt(THREAD, downstream_file_path=xlsx,
                                    upstream_root=upstream)
    assert r["review_data"]["factory_name"] == F_A
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["review_data"]["factory_name"] == F_B
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["status"] == "success", r
    out_path = r["final_output_path"]

    # 用户在装箱单尾部追加一行新工厂 D
    xlsx = _write_manifest(xlsx_path, [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
        (F_D, SKU_D, "ITEM-D", 40),
    ])

    r = service.add_factories_to_batch(THREAD)
    assert r["added"] == 1 and r["factories"] == [F_D], r
    assert r["status"] == "pending_human_review"
    assert r["review_data"]["factory_name"] == F_D

    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["status"] == "success", r

    state = service.get_order_state(THREAD)
    outputs = state["values"].get("factory_outputs") or {}
    assert set(outputs) == {F_A, F_B, F_D}, sorted(outputs)

    # 输出副本已含 D 行且写入正确（行同步生效，非幽灵数据）；
    # A/B 行保持首次写入值，未被重复写坏
    rows = _excel_rows_by_sku(out_path)
    assert len(rows) == 3, rows
    assert rows[SKU_D]["factory"] == F_D
    assert rows[SKU_D]["net"] == round(MOCK_UNIT_NET * 40, 2)
    assert rows[SKU_D]["gross"] == round(MOCK_UNIT_GROSS * 40, 2)
    assert rows[SKU_A]["net"] == 50.0 and rows[SKU_B]["net"] == 100.0
    print("[断言通过] 装箱单追加新工厂：D 进入处理，输出副本行同步正确")


# ---------------------------------------------------------------------------
# 3. 无待补充：全部 approve 后调用 → added==0 + message，state 不动
# ---------------------------------------------------------------------------

def test_nothing_to_add(monkeypatch):
    _force_mock_extraction(monkeypatch)
    F_A = "无补厂甲"
    THREAD = "ADDF-E2E-NONE"

    xlsx = _write_manifest(TMP / "addf_none.xlsx", [
        (F_A, "4900000000301", "ITEM-A", 10),
    ])
    upstream = _make_upstream("none", [F_A])
    _write_batch_config(THREAD, xlsx, upstream)

    r = service.run_until_interrupt(THREAD, downstream_file_path=xlsx,
                                    upstream_root=upstream)
    assert r["review_data"]["factory_name"] == F_A
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["status"] == "success", r

    r = service.add_factories_to_batch(THREAD)
    assert r["added"] == 0 and r["factories"] == [], r
    assert r["message"] == "没有待补充的工厂", r

    # state 未被触碰：仍 completed
    state = service.get_order_state(THREAD)
    assert not state["next_nodes"], state["next_nodes"]
    outputs = state["values"].get("factory_outputs") or {}
    assert set(outputs) == {F_A}
    print("[断言通过] 无待补充：added==0 + message，state 不动")


# ---------------------------------------------------------------------------
# 4. 已 approve 工厂不被重复处理（factory_outputs 键不重复、Excel 不重复写）
# ---------------------------------------------------------------------------

def test_approved_factories_not_reprocessed(monkeypatch):
    _force_mock_extraction(monkeypatch)
    F_A, F_B = "防重厂甲", "防重厂乙"
    SKU_A, SKU_B = "4900000000401", "4900000000402"
    THREAD = "ADDF-E2E-DUP"

    xlsx = _write_manifest(TMP / "addf_dup.xlsx", [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
    ])
    upstream = _make_upstream("dup", [F_A, F_B])
    _write_batch_config(THREAD, xlsx, upstream)

    r = service.run_until_interrupt(THREAD, downstream_file_path=xlsx,
                                    upstream_root=upstream)
    assert r["review_data"]["factory_name"] == F_A
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["review_data"]["factory_name"] == F_B
    # B 跳过 → completed；此时 A 已写入、B 待补充
    r = service.resume_order(THREAD, {"approved": False, "skipped": True, "items": []})
    assert r["status"] == "success", r
    out_path = r["final_output_path"]

    # 补充 B 并 approve
    r = service.add_factories_to_batch(THREAD)
    assert r["added"] == 1 and r["factories"] == [F_B], r
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["status"] == "success", r

    # factory_outputs 恰为 {A, B}（dict 键天然不重复，但确认无多余键）
    outputs = service.get_order_state(THREAD)["values"].get("factory_outputs") or {}
    assert set(outputs) == {F_A, F_B}, sorted(outputs)

    # Excel 仍是 2 行数据（A 未被重复写/追加行），A 的值未被改写
    rows = _excel_rows_by_sku(out_path)
    assert len(rows) == 2, rows
    assert rows[SKU_A]["net"] == 50.0 and rows[SKU_A]["gross"] == 53.0
    assert rows[SKU_B]["net"] == 100.0 and rows[SKU_B]["gross"] == 106.0
    print("[断言通过] 已 approve 工厂不重复处理：factory_outputs 与 Excel 均无重复")


# ---------------------------------------------------------------------------
# 5. pending_review 状态调用：挂起现场不丢（Node5 重挂起），pending 已合并
# ---------------------------------------------------------------------------

def test_pending_review_merge_and_resuspend(monkeypatch):
    _force_mock_extraction(monkeypatch)
    F_A, F_B, F_C = "挂起厂甲", "挂起厂乙", "挂起厂丙"
    SKU_A, SKU_B, SKU_C = "4900000000501", "4900000000502", "4900000000503"
    THREAD = "ADDF-E2E-PENDING"

    xlsx_path = TMP / "addf_pending.xlsx"
    xlsx = _write_manifest(xlsx_path, [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
    ])
    upstream = _make_upstream("pending", [F_A, F_B, F_C])
    _write_batch_config(THREAD, xlsx, upstream)

    r = service.run_until_interrupt(THREAD, downstream_file_path=xlsx,
                                    upstream_root=upstream)
    assert r["review_data"]["factory_name"] == F_A
    orig_items = r["review_data"]["items"]

    # 挂起中，装箱单尾部追加 C
    xlsx = _write_manifest(xlsx_path, [
        (F_A, SKU_A, "ITEM-A", 10),
        (F_B, SKU_B, "ITEM-B", 20),
        (F_C, SKU_C, "ITEM-C", 30),
    ])

    r = service.add_factories_to_batch(THREAD)
    assert r["added"] == 1 and r["factories"] == [F_C], r
    assert r["status"] == "pending_human_review", r
    # 重挂起：仍是当前工厂 A，payload 与挂起前等价（current_factory_data 未变）
    assert r["review_data"]["factory_name"] == F_A, r["review_data"]
    assert len(r["review_data"]["items"]) == len(orig_items)
    # get_review_payload 恢复现场可用
    payload = service.get_review_payload(THREAD)
    assert payload and payload["factory_name"] == F_A

    # pending 已合并（B 原有 + C 新增）
    state = service.get_order_state(THREAD)
    assert state["next_nodes"] == ["node5_human_review"], state["next_nodes"]
    assert set(state["values"].get("pending_factories") or []) == {F_B, F_C}
    print("[断言通过] pending_review 补充：Node5 重挂起现场不丢，pending 已合并")

    # 审核链走完：A → B → C → completed，三家全部写入
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["review_data"]["factory_name"] == F_B
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["review_data"]["factory_name"] == F_C
    r = service.resume_order(THREAD, {"approved": True, "items": []})
    assert r["status"] == "success", r

    outputs = service.get_order_state(THREAD)["values"].get("factory_outputs") or {}
    assert set(outputs) == {F_A, F_B, F_C}, sorted(outputs)
    rows = _excel_rows_by_sku(r["final_output_path"])
    assert rows[SKU_C]["net"] == 150.0 and rows[SKU_C]["gross"] == 159.0
    print("[断言通过] 挂起中补充的工厂随审核链走完：三家全部落 factory_outputs")


def main():
    import pytest  # 脚本模式下提供 MonkeyPatch（pytest 运行时是注入的 fixture）
    with pytest.MonkeyPatch.context() as mp:
        test_completed_batch_skip_then_add(mp)
    with pytest.MonkeyPatch.context() as mp:
        test_add_new_factory_from_appended_manifest(mp)
    with pytest.MonkeyPatch.context() as mp:
        test_nothing_to_add(mp)
    with pytest.MonkeyPatch.context() as mp:
        test_approved_factories_not_reprocessed(mp)
    with pytest.MonkeyPatch.context() as mp:
        test_pending_review_merge_and_resuspend(mp)
    print("\nadd_factories_test: PASS")


if __name__ == "__main__":
    main()
