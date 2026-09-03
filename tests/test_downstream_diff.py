# -*- coding: utf-8 -*-
"""下游装箱单差异对比器测试。

运行：
    cd app && PYTHONPATH=. python3 -m pytest tests/test_downstream_diff.py -v
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))
sys.path.insert(0, str(APP_ROOT / "validation"))

os.environ["EXTRACTION_MOCK"] = "1"
os.environ["DISPATCHER_MOCK"] = "1"

from app.config import get_settings  # noqa: E402
from app.orchestrator.downstream_diff import compare_downstream  # noqa: E402

from _test_isolation import isolate_to_tmp  # noqa: E402
from openpyxl import Workbook  # noqa: E402

TMP = isolate_to_tmp("yamato_downstream_diff_test_")


def _make_content(path: Path, factories: list[tuple[str, str]]):
    """构造最小 Content 表：[(工厂, SKU)]"""
    wb = Workbook()
    ws = wb.active
    settings = get_settings()
    ws.append([settings.col_factory, settings.col_sku, "M3"])
    for factory, sku in factories:
        ws.append([factory, sku, "0.1"])
    wb.save(path)


def test_no_change():
    old = TMP / "old1.xlsx"
    new = TMP / "new1.xlsx"
    _make_content(old, [("中地", "1234567890123")])
    _make_content(new, [("中地", "1234567890123")])

    result = compare_downstream(str(old), str(new))
    assert result["recommendation"] == "none"
    assert result["changed_factories"] == []


def test_factory_added():
    old = TMP / "old2.xlsx"
    new = TMP / "new2.xlsx"
    _make_content(old, [("中地", "1234567890123")])
    _make_content(new, [("中地", "1234567890123"), ("正達", "1234567890124")])

    result = compare_downstream(str(old), str(new))
    assert result["recommendation"] == "diff"
    assert result["added_factories"] == ["正達"]


def test_factory_removed():
    old = TMP / "old3.xlsx"
    new = TMP / "new3.xlsx"
    _make_content(old, [("中地", "1234567890123"), ("正達", "1234567890124")])
    _make_content(new, [("中地", "1234567890123")])

    result = compare_downstream(str(old), str(new))
    assert result["recommendation"] == "diff"
    assert result["removed_factories"] == ["正達"]


def test_content_changed():
    old = TMP / "old4.xlsx"
    new = TMP / "new4.xlsx"
    _make_content(old, [("中地", "1234567890123")])

    wb = Workbook()
    ws = wb.active
    settings = get_settings()
    ws.append([settings.col_factory, settings.col_sku, "M3"])
    ws.append(["中地", "1234567890123", "0.2"])
    wb.save(new)

    result = compare_downstream(str(old), str(new))
    assert result["recommendation"] == "full"
    assert result["changed_factories"] == ["中地"]


def test_structure_changed():
    old = TMP / "old5.xlsx"
    new = TMP / "new5.xlsx"
    wb1 = Workbook()
    ws1 = wb1.active
    settings = get_settings()
    ws1.append([settings.col_factory, settings.col_sku])
    ws1.append(["中地", "1234567890123"])
    wb1.save(old)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append([settings.col_factory, settings.col_sku, "新列"])
    ws2.append(["中地", "1234567890123", "x"])
    wb2.save(new)

    result = compare_downstream(str(old), str(new))
    assert result["recommendation"] == "full"
    assert result["structure_changed"]


if __name__ == "__main__":
    import pytest
    pytest.main([__file__, "-v"])
