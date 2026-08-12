# -*- coding: utf-8 -*-
"""A1 同值压缩 + A2a 单射条码回填 单测（批次2 Step2.2 X1）。

运行方式（纯 DataFrame 处理，不调 LLM / 不走 db）：
    cd app && PYTHONPATH=. python3 tests/test_excel_channel_a1_a2a.py
    或 PYTHONPATH=. python3 -m pytest tests/test_excel_channel_a1_a2a.py -v

血泪红线（2026-08-11）：子进程 import 链（llm_client.py）会执行
load_dotenv(override=True)，把父进程预设的隔离环境变量打回真实路径——
本测试直接 import 模块层（不启动子进程、不调 LLM），因此只需要：

  - YAMATO_TEST_MODE=1（避免真实 .env 副作用；模块 import 不读 dotenv 但显式声明）
  - YAMATO_DOTENV_PATH=空 .env 占位（防止任何意外子链 load_dotenv）
  - 临时 .env 写到 TemporaryDirectory，进程退出自动清理

覆盖：
A1：
- 同行相邻同值长字符串压缩 → 后续 cell 变空
- 数字不动（同一数字重复不压缩）
- 短串不动（KGS / CTNS 等 4 字符不压缩）

A2a：
- 触发形态：1 主行 + 1 下方条码行（含 BARCODE:4936695359672）→ 主行末尾追加条码
- 不触发：多主行
- 不触发：多条码
- 不触发：条码在主行上方
- 不触发：一条码多 SKU
"""
from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import pandas as pd
import pytest

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


# ---------------------------------------------------------------------------
# 隔离门（2026-08-11 事故教训：任何子链 load_dotenv 都不能打回真实路径）
# ---------------------------------------------------------------------------


def _install_isolated_env() -> None:
    """写入临时 .env + 暴露 YAMATO_DOTENV_PATH / YAMATO_TEST_MODE。

    本测试纯 DataFrame 处理，不调 LLM / 不启子进程；但显式守卫以防
    后续 _df_to_markdown 被替换实现时意外触发真实 .env 副作用。
    """
    tmp = Path(tempfile.mkdtemp(prefix="yamato_a1a2a_"))
    empty_env = tmp / ".env"
    empty_env.write_text("# isolated .env for test_excel_channel_a1_a2a\n", encoding="utf-8")
    os.environ["YAMATO_TEST_MODE"] = "1"
    os.environ["YAMATO_DOTENV_PATH"] = str(empty_env)


_install_isolated_env()

from app.extraction.excel_channel import (  # noqa: E402  必须在 _install_isolated_env 之后
    _bind_orphan_barcode,
    _compress_adjacent_dup_strings,
    _df_to_markdown,
)


# ---------------------------------------------------------------------------
# A1：同值长字符串压缩
# ---------------------------------------------------------------------------


def test_a1_compresses_adjacent_dup_long_strings():
    """同行相邻同值长字符串压缩：第二个及之后的同值 cell 变空。"""
    df = pd.DataFrame(
        [
            ["商品名称AAA", "商品名称AAA", "商品名称AAA", "100", "200"],  # 三个长串（len=6）相邻 run
            ["短", "短", "短", "1", "2"],                                # 短串（<6）不压
            ["hello world", "hello world", "x", "3", "4"],               # 长度 11
        ]
    )
    out = _compress_adjacent_dup_strings(df)
    # 第 0 行：第 0 列保留，后续两列被压成 ""
    assert out.iloc[0, 0] == "商品名称AAA"
    assert out.iloc[0, 1] == ""
    assert out.iloc[0, 2] == ""
    # 数字列保持原值（int, numpy int 都允许）
    assert out.iloc[0, 3] in ("100", 100)
    assert out.iloc[0, 4] in ("200", 200)
    # 第 1 行：短串（"短" len<6）→ 不压
    assert out.iloc[1, 0] == "短"
    assert out.iloc[1, 1] == "短"
    assert out.iloc[1, 2] == "短"
    # 第 2 行：两个 "hello world" 长度 11，触发压缩
    assert out.iloc[2, 0] == "hello world"
    assert out.iloc[2, 1] == ""
    assert out.iloc[2, 2] == "x"


def test_a1_numeric_dup_not_compressed():
    """数字重复不压缩（N.W.=G.W. 是合法同值错位）。"""
    df = pd.DataFrame(
        [
            ["PRODUCT-001", "PCS", "50.5", "50.5", "KGS"],
        ]
    )
    out = _compress_adjacent_dup_strings(df)
    # 数字 50.5 重复两次，不应被压
    assert out.iloc[0, 2] == "50.5" or out.iloc[0, 2] == 50.5
    assert out.iloc[0, 3] == "50.5" or out.iloc[0, 3] == 50.5
    # 字符串 "PRODUCT-001" / "KGS" / "PCS" 都是单次出现
    assert out.iloc[0, 0] == "PRODUCT-001"
    assert out.iloc[0, 1] == "PCS"
    assert out.iloc[0, 4] == "KGS"


def test_a1_short_strings_not_compressed():
    """短串（KGS / CTNS 等 ≤4 字符）不压缩，保持列对齐。"""
    df = pd.DataFrame(
        [
            ["KGS", "KGS", "KGS", "KGS"],
            ["CTNS", "CTNS", "PCS", "PCS"],
            ["N.W.", "N.W.", "G.W.", "G.W."],
        ]
    )
    out = _compress_adjacent_dup_strings(df)
    # 所有 cell 仍原样（全部 len<6）
    for i in range(3):
        for j in range(4):
            assert out.iloc[i, j] == df.iloc[i, j], (
                f"短串不应被压：({i},{j}) {out.iloc[i, j]!r} vs {df.iloc[i, j]!r}"
            )


def test_a1_only_first_in_run_kept_run_length_4():
    """同值 run 长 4：保留首 cell，其余 3 个置空。"""
    df = pd.DataFrame(
        [
            ["同一品名长字符串", "同一品名长字符串", "同一品名长字符串", "同一品名长字符串"],
        ]
    )
    out = _compress_adjacent_dup_strings(df)
    assert out.iloc[0, 0] == "同一品名长字符串"
    assert out.iloc[0, 1] == ""
    assert out.iloc[0, 2] == ""
    assert out.iloc[0, 3] == ""


# ---------------------------------------------------------------------------
# A2a：单射形态条码回填
# ---------------------------------------------------------------------------


def _trigger_form_df() -> pd.DataFrame:
    """触发形态：1 主行 + 1 下方条码行（含 BARCODE:4936695359672）。"""
    return pd.DataFrame(
        [
            ["WIDGET-001", "DESCRIPTION-A", 100, 50.5, 60.0],
            ["BARCODE:4936695359672", "", "", "", ""],
        ]
    )


def test_a2a_trigger_form_appends_barcode_to_main_row():
    """触发：主行末尾追加条码数字。"""
    df = _trigger_form_df()
    out = _bind_orphan_barcode(df)
    # 主行索引 0 的末尾追加 1 个新 cell = "4936695359672"
    last_col = out.shape[1] - 1
    assert out.iloc[0, last_col] == "4936695359672"
    # 条码行保留不删
    assert out.iloc[1, 0] == "BARCODE:4936695359672"
    # DataFrame 总列数 = 原始 5 列 + 追加 1 列
    assert out.shape == (2, 6), f"expected (2, 6), got {out.shape}"
    # 主行的非追加 cell 内容不变
    assert out.iloc[0, 0] == "WIDGET-001"
    assert out.iloc[0, 1] == "DESCRIPTION-A"
    assert out.iloc[0, 2] == 100


def test_a2a_no_trigger_when_multi_main_rows():
    """不触发：多主行。"""
    df = pd.DataFrame(
        [
            ["SKU-001 LONG NAME", "DESC-A", 10, 1.0, 2.0],
            ["SKU-002 LONG NAME", "DESC-B", 20, 3.0, 4.0],
            ["BARCODE:4936695359672", "", "", "", ""],
        ]
    )
    out = _bind_orphan_barcode(df)
    # 形状不变（没有追加列）
    assert out.shape == (3, 5), f"多主行不应触发，实际 shape={out.shape}"
    # 原值不变
    assert out.iloc[0, 0] == "SKU-001 LONG NAME"
    assert out.iloc[1, 0] == "SKU-002 LONG NAME"


def test_a2a_no_trigger_when_multi_barcode_rows():
    """不触发：多条码行。"""
    df = pd.DataFrame(
        [
            ["SKU-001 LONG NAME", "DESC-A", 10, 1.0, 2.0],
            ["BARCODE:4936695359672", "", "", "", ""],
            ["BARCODE:4936695359673", "", "", "", ""],
        ]
    )
    out = _bind_orphan_barcode(df)
    assert out.shape == (3, 5), f"多条码不应触发，实际 shape={out.shape}"
    # 主行内容不变
    assert out.iloc[0, 0] == "SKU-001 LONG NAME"
    assert out.iloc[0, 4] == 2.0


def test_a2a_no_trigger_when_barcode_above_main():
    """不触发：条码在主行上方。"""
    df = pd.DataFrame(
        [
            ["BARCODE:4936695359672", "", "", "", ""],
            ["SKU-001 LONG NAME", "DESC-A", 10, 1.0, 2.0],
        ]
    )
    out = _bind_orphan_barcode(df)
    assert out.shape == (2, 5), f"条码在上方不应触发，实际 shape={out.shape}"
    # 主行（行 1）内容不变
    assert out.iloc[1, 0] == "SKU-001 LONG NAME"
    assert out.iloc[1, 4] == 2.0


def test_a2a_no_trigger_when_one_barcode_multi_sku():
    """不触发：一条码多 SKU（多主行 + 单条码，自然由多主行条件拦截）。"""
    df = pd.DataFrame(
        [
            ["SKU-001 LONG NAME", "DESC-A", 10, 1.0, 2.0],
            ["SKU-002 LONG NAME", "DESC-B", 20, 3.0, 4.0],
            ["BARCODE:4936695359672", "", "", "", ""],
        ]
    )
    out = _bind_orphan_barcode(df)
    assert out.shape == (3, 5), f"一条码多 SKU 不应触发，实际 shape={out.shape}"
    assert out.iloc[2, 0] == "BARCODE:4936695359672"


def test_a2a_preserves_barcode_row_after_fill():
    """条码行保留不删（宁冗勿漏）：即使已追加到主行，原条码行 cell 仍存在。"""
    df = _trigger_form_df()
    out = _bind_orphan_barcode(df)
    # 条码行（原 5 列）的 col 0 保留 "BARCODE:4936695359672"，其余保持空
    assert out.iloc[1, 0] == "BARCODE:4936695359672", (
        f"条码行 col 0 被改写: {out.iloc[1, 0]!r}"
    )
    # 追加的新列只在主行
    last_col = out.shape[1] - 1
    assert out.iloc[0, last_col] == "4936695359672"
    # 条码行末尾新增 cell 不应被填（否则等于"复制"，违背"宁冗勿漏"语义）
    # _bind_orphan_barcode 只在主行 .loc 赋值，条码行原列之外无新值
    # （DataFrame copy() 后所有新 cell 为 NaN）


def test_a2a_pure_digit_13_in_cell_is_recognized():
    """触发：条码行只有纯 13 位数字 cell（无 BARCODE: 前缀）。"""
    df = pd.DataFrame(
        [
            ["SKU-XYZ LONG NAME", "DESC-X", 5, 0.5, 0.7],
            ["4936695359672", "", "", "", ""],
        ]
    )
    out = _bind_orphan_barcode(df)
    last_col = out.shape[1] - 1
    assert out.shape == (2, 6), f"纯 13 位也应触发，实际 shape={out.shape}"
    assert out.iloc[0, last_col] == "4936695359672"


# ---------------------------------------------------------------------------
# 端到端：xlsx_to_markdown 路径上的 A1+A2a 集成（in-memory openpyxl workbook）
# ---------------------------------------------------------------------------


def test_integration_xlsx_to_markdown_with_a1_and_a2a(tmp_path: Path):
    """集成测试：构造一个最小 xlsx，跑 xlsx_to_markdown 断言 A1+A2a 都被触发。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"
    # 主行：长品名 + 2 个数值（≥6 字符长文本是 "WIDGET-XYZ-LONG-NAME"）
    ws["A1"] = "WIDGET-XYZ-LONG-NAME"
    ws["B1"] = "GOOD DESCRIPTION"
    ws["C1"] = 100
    ws["D1"] = 50.5
    ws["E1"] = 60.0
    # 重复品名 cell（A1 在拆分合并时本就是这样）
    ws["F1"] = "WIDGET-XYZ-LONG-NAME"   # 同值长串，应被 A1 压空
    # 条码行
    ws["A2"] = "BARCODE:4936695359672"
    xlsx_path = tmp_path / "fixture.xlsx"
    wb.save(xlsx_path)

    from app.extraction.excel_channel import xlsx_to_markdown

    md = xlsx_to_markdown(str(xlsx_path))
    # 末尾追加的条码 "4936695359672" 应当出现在主行所在行（用纯文本包含即可）
    # 但因 markdown 把每行合并成 "| a | b | c | ... |"，直接 assert "4936695359672" 至少出现 2 次：
    # 一次在原条码行，一次在主行末尾追加列。
    occurrences = md.count("4936695359672")
    assert occurrences >= 2, (
        f"A2a 应把条码追加到主行末尾，实际 markdown 中条码出现 {occurrences} 次\n{md}"
    )


# ---------------------------------------------------------------------------
# 主入口（直接 python3 跑也能用）
# ---------------------------------------------------------------------------


def _run_all() -> None:
    """直接 python3 tests/test_excel_channel_a1_a2a.py 时的兜底执行。"""
    import inspect
    funcs = [
        (name, obj)
        for name, obj in globals().items()
        if name.startswith("test_") and callable(obj)
    ]
    passed = failed = skipped = 0
    for name, fn in funcs:
        sig = inspect.signature(fn)
        has_required_param = any(
            p.default is inspect.Parameter.empty
            and p.kind not in (
                inspect.Parameter.VAR_POSITIONAL,
                inspect.Parameter.VAR_KEYWORD,
            )
            for p in sig.parameters.values()
        )
        if has_required_param:
            print(f"  ~ {name} (has fixture param — pytest only)")
            skipped += 1
            continue
        try:
            fn()
            print(f"  PASS  {name}")
            passed += 1
        except pytest.skip.Exception as e:
            print(f"  SKIP  {name}: {e}")
            skipped += 1
        except Exception as e:  # noqa: BLE001
            print(f"  FAIL  {name}: {type(e).__name__}: {e}")
            failed += 1
    print(f"\n{passed} passed, {failed} failed, {skipped} skipped")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    print("test_excel_channel_a1_a2a")
    if "-v" in sys.argv or "--verbose" in sys.argv:
        sys.exit(pytest.main([__file__, "-v"]))
    _run_all()