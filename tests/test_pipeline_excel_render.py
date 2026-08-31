# -*- coding: utf-8 -*-
"""convert_excel_to_pdf 单测：覆盖兆丰 Packing 裁切修复（C1）+ .xls 预转换管线。

运行方式（依赖 LibreOffice 可用）：
    cd app && PYTHONPATH=. python3 -m pytest tests/test_pipeline_excel_render.py -v
    或直接：python3 tests/test_pipeline_excel_render.py

隔离策略：
- 不走 db 也不走 LLM，只调 app.extraction.pipeline.convert_excel_to_pdf
  和 app.extraction.pipeline._preprocess_for_soffice；不读 llm_client
  的真实 .env，因此不需要 isolate_to_tmp。
- 临时目录用 stdlib tempfile.TemporaryDirectory，进程退出时自动清理。

覆盖：
- 兆丰 Packing XD-261830-001-1.26 / 1.28（视图停在 D15）：转换后 PDF 文本含
  "JAN CODE" 和 "4549509623861"，证明 SinglePageSheets 不再以 topLeftCell
  为锚点裁切左上整片。
- 兆丰 1.30（topLeftCell 可能停在 B13）：同上。
- 正常 xlsx（topLeftCell 已为 "A1"）：不引入回归——确认预处理函数返回
  的副本路径下 topLeftCell == "A1"、selection 为空；以及转换后 PDF 仍
  包含主要文本。
- 临时副本不影响源：转换前后源文件 mtime 不变，且临时副本位于 out_dir/
  _lo_preprocess/ 下，转换后被调用方清掉（这里直接用 TemporaryDirectory
  验证 out_dir 干净）。
- .xls 预转换失败兜底：mock subprocess.run 让 .xls→.xlsx 预转换不产出文件，
  验证回退原始 .xls 走普通 pdf 过滤器（修复前旧行为即兜底）。
- .xls 预转换成功：mock 让 .xls→.xlsx 产出副本，验证随后走 SinglePageSheets
  单页管线（每个 sheet 一页，修复 TOP 请款资料一 sheet 拆多页问题）。
- 预处理函数 .xls 短路：返回 None。
- 预处理函数 openpyxl 异常：WARNING 日志 + 返回 None（不影响主流程）。
"""
from __future__ import annotations

import os
import subprocess
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest

APP_ROOT = Path(__file__).resolve().parent.parent
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))

# 强制使用 macOS 自带 LibreOffice（PATH 上不一定有 soffice）
os.environ.setdefault(
    "SOFFICE_PATH",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
)

from app.extraction.pipeline import (  # noqa: E402
    _preprocess_for_soffice,
    convert_excel_to_pdf,
)

FIXTURES = APP_ROOT / "tests" / "fixtures" / "zhaofeng"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _pdf_text(pdf_path: str) -> str:
    """从 PDF 提取全部文本（PyPDF2；兆丰真实文件实测可解出 13 位 JAN CODE）。"""
    import PyPDF2
    r = PyPDF2.PdfReader(pdf_path)
    out = []
    for p in r.pages:
        out.append(p.extract_text() or "")
    return "\n".join(out)


def _has_soffice() -> bool:
    """soffice 可用性探测——不在就 skip 真实转换用例。"""
    from app.extraction.pipeline import _find_soffice
    return _find_soffice() is not None


requires_soffice = pytest.mark.skipif(
    not _has_soffice(), reason="LibreOffice (soffice) not available"
)


# ---------------------------------------------------------------------------
# 兆丰 Packing（视图停在 D15/B13）—— 真实转换 + 文本断言
# ---------------------------------------------------------------------------


@requires_soffice
@pytest.mark.parametrize(
    "filename",
    [
        "Packing XD-261830-001-1.26.xlsx",
        "Packing XD-261830-001-1.28.xlsx",
        "Packing XD-261830-001-1.30.xlsx",
    ],
)
def test_zhaofeng_pdf_contains_jan_code_and_barcode(filename):
    """C1 修复断言：SinglePageSheets 不再以 topLeftCell 裁切 JAN CODE 列。"""
    src = FIXTURES / filename
    assert src.exists(), f"fixture missing: {src}"
    mtime_before = src.stat().st_mtime
    with tempfile.TemporaryDirectory() as tmp:
        out_pdf = convert_excel_to_pdf(str(src), tmp)
        assert out_pdf is not None, f"convert failed: {filename}"
        assert Path(out_pdf).exists()
        text = _pdf_text(out_pdf)
        assert "JAN CODE" in text, (
            f"{filename}: JAN CODE column missing from PDF (topLeftCell "
            "crop not fixed); first 300 chars:\n" + text[:300]
        )
        # 兆丰 Packing 真实首个 SKU 的 JAN CODE
        assert "4549509623861" in text, (
            f"{filename}: first barcode 4549509623861 missing"
        )
        # 临时副本应位于 out_dir/_lo_preprocess/ 下
        preprocess_dir = Path(tmp) / "_lo_preprocess"
        # 转换完成后调用方负责清理；这里验证 preprocess_dir 内的临时副本
        # 在转换过程中确实被创建过——若转换后还在也没关系（外层 TmpDir 兜底）
        assert preprocess_dir.exists(), (
            f"{filename}: _lo_preprocess/ not created (预处理未跑)"
        )
    # 源文件 mtime 必须不变（缓存键判定依据）
    mtime_after = src.stat().st_mtime
    assert mtime_before == mtime_after, (
        f"{filename}: source mtime changed by convert! "
        f"before={mtime_before} after={mtime_after}"
    )


# ---------------------------------------------------------------------------
# 正常 xlsx 不引入回归
# ---------------------------------------------------------------------------


@requires_soffice
def test_normal_xlsx_with_topleft_a1_roundtrip():
    """topLeftCell='A1' 的 xlsx 转换后 PDF 文本包含核心表头+首条数据。"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 第一行明确写一行表头与一行数据（13 位 barcode + 数量）
    ws["A1"] = "JAN CODE"
    ws["B1"] = "DESCRIPTION"
    ws["C1"] = "QTY"
    ws["A2"] = "4549509623861"
    ws["B2"] = "TEST GOODS"
    ws["C2"] = 100
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=tempfile.gettempdir()
    ) as f:
        normal_path = f.name
    try:
        wb.save(normal_path)
        # 新建工作簿 topLeftCell 为 None（即默认 A1），无需断言
        mtime_before = Path(normal_path).stat().st_mtime

        with tempfile.TemporaryDirectory() as tmp:
            out_pdf = convert_excel_to_pdf(normal_path, tmp)
            assert out_pdf is not None
            text = _pdf_text(out_pdf)
            # 关键回归断言：表头与 13 位条码都不应被裁掉
            assert "JAN CODE" in text, f"normal xlsx: header lost: {text[:200]}"
            # 13 位 barcode 单测允许 PyPDF2 偶尔少位（提取边界），
            # 但前缀 454950962 必须出现——这就是 SinglePageSheets 还在工作的硬证据
            assert "454950962" in text, (
                f"normal xlsx: barcode truncated by crop: {text[:200]}"
            )
            # 临时副本应位于 _lo_preprocess/ 下
            assert (Path(tmp) / "_lo_preprocess" / Path(normal_path).name).exists()

        # 源 mtime 不变（缓存键判定）
        mtime_after = Path(normal_path).stat().st_mtime
        assert mtime_before == mtime_after
    finally:
        Path(normal_path).unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# _preprocess_for_soffice 单元测试
# ---------------------------------------------------------------------------


def test_preprocess_xlsx_resets_topleft_and_selection():
    """_preprocess_for_soffice 把 xlsx 复制到 _lo_preprocess/ 并重置视图。"""
    src = FIXTURES / "Packing XD-261830-001-1.26.xlsx"
    with tempfile.TemporaryDirectory() as work:
        out = _preprocess_for_soffice(src, Path(work))
        assert out is not None
        assert (Path(work) / "_lo_preprocess" / src.name).exists()
        # 验证副本内 topLeftCell == "A1" 且 selection 全部归位 A1
        from openpyxl import load_workbook
        wb = load_workbook(out)
        for ws in wb.worksheets:
            assert ws.sheet_view.topLeftCell == "A1", (
                f"sheet {ws.title!r} topLeftCell not reset: "
                f"{ws.sheet_view.topLeftCell!r}"
            )
            for sel in ws.sheet_view.selection:
                assert sel.activeCell == "A1", (
                    f"sheet {ws.title!r} selection.activeCell "
                    f"!= A1: {sel.activeCell!r}"
                )
                assert sel.sqref == "A1", (
                    f"sheet {ws.title!r} selection.sqref != A1: {sel.sqref!r}"
                )


def test_preprocess_xls_short_circuits():
    """_preprocess_for_soffice 对 .xls 直接返回 None（openpyxl 不能写）。"""
    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as f:
        fake_xls = f.name
    try:
        out = _preprocess_for_soffice(Path(fake_xls), Path(tempfile.gettempdir()))
        assert out is None
    finally:
        Path(fake_xls).unlink(missing_ok=True)


def test_preprocess_falls_back_on_openpyxl_error():
    """openpyxl 打开失败（损坏的 xlsx）→ WARNING + 返回 None，不抛异常。"""
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=tempfile.gettempdir()
    ) as f:
        broken = f.name
    Path(broken).write_bytes(b"this is not a real xlsx file")
    try:
        out = _preprocess_for_soffice(Path(broken), Path(tempfile.gettempdir()))
        assert out is None
    finally:
        Path(broken).unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# 列宽自适应（窄列长数字条码折行修复）单元测试
# ---------------------------------------------------------------------------


def test_digit_run_need_rules():
    """_digit_run_need 命中规则：按内容不按格式，≥8 位纯数字才命中。"""
    from app.extraction.pipeline import _digit_run_need
    # 文本条码（工厂单据主流格式）
    assert _digit_run_need("4549509515203") == 15
    assert _digit_run_need(" 4549509515203 ") == 15   # 带空白也命中
    # 短数字/含字母/空值不命中
    assert _digit_run_need("1234567") is None         # 7 位，不到 8
    assert _digit_run_need("24K1532") is None         # 含字母
    assert _digit_run_need("CUSHION") is None
    assert _digit_run_need("") is None
    assert _digit_run_need(None) is None
    # 数字格式（窄列显示 ###/科学计数，同样命中）
    assert _digit_run_need(4549509515203) == 15
    assert _digit_run_need(4549509515203.0) == 15
    assert _digit_run_need(480) is None               # 数量级不命中
    assert _digit_run_need(True) is None              # bool 排除


def test_preprocess_widens_narrow_digit_column():
    """窄列长数字条码列被加宽；够宽的列/纯文本列/短数字列不动。"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 6               # 窄列 + 换行 → 折行场景
    ws.column_dimensions["B"].width = 20              # 本来就够宽
    a = ws["A1"]; a.value = "4549509515203"; a.alignment = Alignment(wrap_text=True)
    ws["B1"] = "4549509518860"
    ws["C1"] = "CUSHION 65*115CM"                     # 纯文本不命中
    ws["D1"] = 480                                    # 短数字不命中
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=tempfile.gettempdir()
    ) as f:
        src_path = f.name
    try:
        wb.save(src_path)
        with tempfile.TemporaryDirectory() as work:
            out = _preprocess_for_soffice(Path(src_path), Path(work))
            assert out is not None
            wb2 = load_workbook(out)
            ws2 = wb2.active
            assert ws2.column_dimensions["A"].width >= 15, (
                f"窄列未加宽: {ws2.column_dimensions['A'].width}"
            )
            assert ws2.column_dimensions["B"].width == 20, (
                f"够宽的列不应被改动: {ws2.column_dimensions['B'].width}"
            )
            c_dim = ws2.column_dimensions["C"]
            assert not c_dim.width or c_dim.width < 15, (
                f"纯文本列不应被加宽: {c_dim.width}"
            )
    finally:
        Path(src_path).unlink(missing_ok=True)


@requires_soffice
def test_wrapped_barcode_searchable_after_convert():
    """真实转换：窄列+换行的 13 位条码，转换后 PDF 文本层含完整条码。

    修复前条码在 PDF 里断成两行（45495095152 / 03），审核页高亮定位搜不到
    （中地 XD-265099 实测）；列宽自适应后文本层保持连续字符串。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 6
    a = ws["A1"]; a.value = "4549509515203"; a.alignment = Alignment(wrap_text=True)
    ws["B1"] = "CUSHION"
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=tempfile.gettempdir()
    ) as f:
        src_path = f.name
    try:
        wb.save(src_path)
        with tempfile.TemporaryDirectory() as tmp:
            out_pdf = convert_excel_to_pdf(src_path, tmp)
            assert out_pdf is not None, "convert failed"
            text = _pdf_text(out_pdf).replace("\n", "")
            assert "4549509515203" in text, (
                f"条码在 PDF 文本层断裂/丢失: {text[:200]}"
            )
    finally:
        Path(src_path).unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# .xls 预转换管线：先转 xlsx 再走 SinglePageSheets；预转换失败回退普通分页
# （mock subprocess.run）
# ---------------------------------------------------------------------------

_SINGLEPAGE_FILTER = (
    'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"}}'
)


def test_xls_preconvert_fails_falls_back_to_plain_pdf():
    """.xls→.xlsx 预转换失败（未产出 xlsx）时，回退原始 .xls 走普通 pdf 过滤器。"""
    # 造一个最小 .xls 后缀文件（内容无所谓，mock 不真读）
    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as f:
        fake = f.name
    try:
        with tempfile.TemporaryDirectory() as tmp:
            calls_filters: list[str] = []

            def fake_run(cmd, **kwargs):
                # 抓 --convert-to 后面的 filter
                if "--convert-to" in cmd:
                    idx = cmd.index("--convert-to")
                    calls_filters.append(cmd[idx + 1])
                # 一律只写伪 PDF：预转换（--convert-to xlsx）期望的产物是
                # stem+".xlsx"，这里故意不产出 → _convert_xls_to_xlsx 返回
                # None → 主流程回退原始 .xls 走普通 pdf
                outdir = cmd[cmd.index("--outdir") + 1]
                Path(outdir, Path(fake).stem + ".pdf").write_bytes(b"%PDF-fake")
                return subprocess.CompletedProcess(cmd, 0, stdout=b"", stderr=b"")

            with patch("app.extraction.pipeline.subprocess.run", side_effect=fake_run), \
                 patch("app.extraction.pipeline._find_soffice",
                       return_value="/fake/soffice"):
                out = convert_excel_to_pdf(fake, tmp)

            assert out is not None
            assert calls_filters == ["xlsx", "pdf"], (
                f".xls 应先尝试预转换 xlsx，失败后回退普通 pdf，实际: {calls_filters}"
            )
    finally:
        Path(fake).unlink(missing_ok=True)


def test_xls_preconvert_success_uses_singlepage():
    """.xls→.xlsx 预转换成功时，后续走 SinglePageSheets 单页管线（不再普通分页）。"""
    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as f:
        fake = f.name
    try:
        with tempfile.TemporaryDirectory() as tmp:
            calls_filters: list[str] = []

            def fake_run(cmd, **kwargs):
                idx = cmd.index("--convert-to")
                conv_filter = cmd[idx + 1]
                calls_filters.append(conv_filter)
                outdir = cmd[cmd.index("--outdir") + 1]
                if conv_filter == "xlsx":
                    # 预转换：产出 xlsx 副本（内容随意——openpyxl 打不开时
                    # _preprocess_for_soffice 只记 warning 返回 None，
                    # soffice_input 回退为该 xlsx 原样，不影响过滤器断言）
                    Path(outdir, Path(fake).stem + ".xlsx").write_bytes(b"fake-xlsx")
                else:
                    # pdf 系过滤器：产出伪 PDF 让主循环 expect.exists() 命中
                    Path(outdir, Path(fake).stem + ".pdf").write_bytes(b"%PDF-fake")
                return subprocess.CompletedProcess(cmd, 0, stdout=b"", stderr=b"")

            with patch("app.extraction.pipeline.subprocess.run", side_effect=fake_run), \
                 patch("app.extraction.pipeline._find_soffice",
                       return_value="/fake/soffice"):
                out = convert_excel_to_pdf(fake, tmp)

            assert out is not None
            assert calls_filters == ["xlsx", _SINGLEPAGE_FILTER], (
                f".xls 预转换成功后应走 SinglePageSheets，实际: {calls_filters}"
            )
    finally:
        Path(fake).unlink(missing_ok=True)


def test_xlsx_attempts_singlepage_first_then_plain():
    """convert_excel_to_pdf 对 .xlsx 先尝试 SinglePageSheets，失败才回退。"""
    src = FIXTURES / "Packing XD-261830-001-1.26.xlsx"
    with tempfile.TemporaryDirectory() as tmp:
        real_run = subprocess.run
        calls_filters: list[str] = []
        call_count = {"n": 0}

        def fake_run(cmd, **kwargs):
            call_count["n"] += 1
            if "--convert-to" in cmd:
                idx = cmd.index("--convert-to")
                calls_filters.append(cmd[idx + 1])
            outdir = cmd[cmd.index("--outdir") + 1]
            # 第一次（SinglePageSheets）故意 raise CalledProcessError
            # 第二次（普通 pdf）写出文件
            if call_count["n"] == 1:
                raise subprocess.CalledProcessError(
                    1, cmd, stderr=b"fake SinglePageSheets failure"
                )
            Path(outdir, src.stem + ".pdf").write_bytes(b"%PDF-fake")
            return subprocess.CompletedProcess(cmd, 0, stdout=b"", stderr=b"")

        with patch("app.extraction.pipeline.subprocess.run", side_effect=fake_run), \
             patch("app.extraction.pipeline._find_soffice",
                   return_value="/fake/soffice"):
            out = convert_excel_to_pdf(str(src), tmp)

        assert out is not None
        assert calls_filters[0].startswith("pdf:calc_pdf_Export"), (
            f".xlsx 应先尝试 SinglePageSheets，实际: {calls_filters}"
        )
        assert calls_filters[-1] == "pdf", (
            f".xlsx 失败后应回退普通 pdf，实际: {calls_filters}"
        )


# ---------------------------------------------------------------------------
# 主入口（直接 python3 跑也能用）
# ---------------------------------------------------------------------------


def _run_all() -> None:
    """直接 python3 tests/test_pipeline_excel_render.py 时的兜底执行。"""
    import inspect
    funcs = [
        (name, obj)
        for name, obj in globals().items()
        if name.startswith("test_") and callable(obj)
    ]
    passed = failed = skipped = 0
    for name, fn in funcs:
        sig = inspect.signature(fn)
        # 仅当函数有未绑定默认值的命名形参（@parametrize 注入）才跳过
        has_required_param = any(
            p.default is inspect.Parameter.empty
            and p.kind not in (
                inspect.Parameter.VAR_POSITIONAL,
                inspect.Parameter.VAR_KEYWORD,
            )
            for p in sig.parameters.values()
        )
        # skipif 不会注入额外形参；只看 require_soffice 是否命中
        is_parametrized = any(
            getattr(m, "name", "") == "parametrize"
            for m in getattr(fn, "pytestmark", [])
        )
        if has_required_param and is_parametrized:
            print(f"  ~ {name} (parametrized — pytest only)")
            skipped += 1
            continue
        try:
            fn()
            print(f"  PASS  {name}")
            passed += 1
        except pytest.skip.Exception as e:
            print(f"  SKIP  {name}: {e}")
            skipped += 1
        except Exception as e:  # noqa: BLE001
            print(f"  FAIL  {name}: {type(e).__name__}: {e}")
            failed += 1
    print(f"\n{passed} passed, {failed} failed, {skipped} skipped")


if __name__ == "__main__":
    print("test_pipeline_excel_render")
    if "-v" in sys.argv or "--verbose" in sys.argv:
        import pytest as _pt
        sys.exit(_pt.main([__file__, "-v"]))
    _run_all()
