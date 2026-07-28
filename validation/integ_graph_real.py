# -*- coding: utf-8 -*-
"""Node3 真实联调：去 mock，真实 qwen3.7-plus 驱动山東中地端到端。

流程：
  Node1 解析真实下游表 → Node2 别名命中「中地」文件夹
  → Node3 FactorySession 增量模式真实提取（LLM 只跑真箱单）
  → Node4 计算+查库 → Node5 interrupt 挂起
  → 校验 payload 与 GT 逐格一致（14 SKU）
  → resume approved（不改动数值，新 SKU 补录合规字段）
  → Node6 写 Excel 副本+落库 → Node7 → END

用法（在 app/ 目录下）：
  LLM_ENABLE_THINKING=0 python3 validation/integ_graph_real.py --reset

--reset：删除 master.db / checkpoints.db / 输出副本（此前均为 mock 冒烟数据）。

下游表使用_原文件（无 中文品名/净重/毛重 三列的客户原始文件），按 2026-07-28
用户定的写入规则断言：
- Node6 首次写入时在 SHOHIN_MEI_E 后插入三列（32/33/34 位），已存在则跳过；
- 净重/毛重 = 单重 × SOTOBAKO_D_HACCHU_SU，2 位小数；
- 中文品名 = 主数据 name_cn；其余列与原文件逐格一致（格式严格不变）。
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

APP_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_ROOT))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.api import service  # noqa: E402
from app.config import get_settings  # noqa: E402
from app.db.models import FactorySKU  # noqa: E402
from app.db.session import get_session  # noqa: E402
from ground_truth import get_factory_ground_truth  # noqa: E402
from sqlalchemy import select  # noqa: E402

THREAD_ID = "INTEG-ZD3-REAL"
TARGET_FACTORY_JP = "山東中地"   # 下游表 MAKER_MEI_KJ 原文
TARGET_FACTORY_LOCAL = "中地"    # 本地文件夹名（alias_map 命中）
# 生产下游表为_原文件（无三列的客户原始文件）；可用环境变量 INTEG_DOWNSTREAM_FILE 覆盖
DOWNSTREAM_FILE = os.environ.get(
    "INTEG_DOWNSTREAM_FILE",
    "/Users/nz/Downloads/yamato/96/"
    "ContentsOfTheContainer_202624_青島XD_20260708_原文件.xlsx",
)
TOL = 0.01


def reset_env() -> None:
    settings = get_settings()
    for p in (settings.master_db_abs, settings.checkpoint_db_abs):
        if p.exists():
            p.unlink()
            print(f"[reset] 已删除 {p}")
    out_copy = settings.output_dir_abs / (Path(DOWNSTREAM_FILE).stem + "_filled.xlsx")
    if out_copy.exists():
        out_copy.unlink()
        print(f"[reset] 已删除 {out_copy}")


def check_payload_against_gt(review_data: dict) -> list[str]:
    """把 interrupt payload 的提取值与 GT 逐格对照，返回错误描述列表。"""
    gt = get_factory_ground_truth(TARGET_FACTORY_LOCAL)
    errors: list[str] = []
    items_by_sku = {i["sku"]: i for i in review_data["items"]}

    for sku, g in gt.items():
        it = items_by_sku.get(sku)
        if it is None:
            errors.append(f"SKU {sku} 未出现在提取结果中")
            continue
        ext = it["extracted_data"]
        if ext.get("total_quantity") != g["total_quantity"]:
            errors.append(f"{sku} 件数 {ext.get('total_quantity')} != GT {g['total_quantity']}")
        for k, gk in (("total_net_weight", "total_net_weight"),
                      ("total_gross_weight", "total_gross_weight")):
            v = ext.get(k)
            if v is None or abs(v - g[gk]) >= TOL:
                errors.append(f"{sku} {k} {v} != GT {g[gk]}")
    extra = sorted(set(items_by_sku) - set(gt))
    if extra:
        errors.append(f"提取结果含 GT 外 SKU: {extra}")
    if review_data.get("missing_skus"):
        errors.append(f"missing_skus 非空: {review_data['missing_skus']}")
    return errors


def check_output_excel(out_path: Path, review_data: dict) -> None:
    """按 2026-07-28 写入规则校验输出副本：

    1. 中文品名/净重/毛重 三列已插入在 SHOHIN_MEI_E 之后（32/33/34 位）；
    2. 中地各行：净重/毛重 = 单重 × SOTOBAKO_D_HACCHU_SU（2 位小数），
       中文品名 = 审核提交的 name_cn；
    3. 其余 57 列与原文件逐格一致（格式严格不变：未插入的列零改动）。
    """
    from openpyxl import load_workbook

    gt = get_factory_ground_truth(TARGET_FACTORY_LOCAL)
    name_by_sku = {}
    unit_by_sku = {}
    for i, item in enumerate(review_data["items"]):
        calc = item.get("calculation") or {}
        unit_by_sku[item["sku"]] = (calc.get("calculated_unit_net"),
                                    calc.get("calculated_unit_gross"))
        ext = item.get("extracted_data") or {}
        # 与 resume 提交逻辑一致的 name_cn 期望值
        name_by_sku[item["sku"]] = (
            ext.get("sku_name") or f"待补品名-{i + 1}"
        ) if item.get("is_new_sku") else (item.get("db_record") or {}).get("name_cn")

    wb = load_workbook(out_path)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    assert header[31:34] == ["中文品名", "净重", "毛重"], \
        f"三列未按预期插入 32-34 位: {header[28:38]}"
    assert header[31] == "中文品名" and header[34] == "D_HACCHU_SU"
    i_sku = header.index("SHOHIN_CD") + 1
    i_factory = header.index("MAKER_MEI_KJ") + 1
    i_qty = header.index("SOTOBAKO_D_HACCHU_SU") + 1
    i_cn, i_net, i_gross = 32, 33, 34

    n_checked = 0
    for row in ws.iter_rows(min_row=2):
        if (row[i_factory - 1].value or "").strip() != TARGET_FACTORY_JP:
            continue
        sku = str(row[i_sku - 1].value).strip()
        qty = float(row[i_qty - 1].value or 0)
        unit_net, unit_gross = unit_by_sku[sku]
        assert abs(row[i_net - 1].value - round(unit_net * qty, 2)) < 1e-9, \
            f"{sku} 行{row[0].row} 净重 {row[i_net - 1].value} != {round(unit_net * qty, 2)}"
        assert abs(row[i_gross - 1].value - round(unit_gross * qty, 2)) < 1e-9, \
            f"{sku} 行{row[0].row} 毛重不符"
        assert row[i_cn - 1].value == name_by_sku[sku], \
            f"{sku} 行{row[0].row} 中文品名 {row[i_cn - 1].value!r} != {name_by_sku[sku]!r}"
        assert sku in gt
        n_checked += 1
    assert n_checked > 0
    print(f"[断言通过] 三列已插入 32-34 位；中地 {n_checked} 行 "
          f"净重/毛重=单重×SOTOBAKO_D_HACCHU_SU(2位小数)、中文品名=name_cn 全部正确")
    wb.close()

    # 未插入的 57 列与原文件逐格一致
    wb_o = load_workbook(DOWNSTREAM_FILE, read_only=True)
    ws_o = wb_o[wb_o.sheetnames[0]]
    wb_n = load_workbook(out_path, read_only=True)
    ws_n = wb_n[wb_n.sheetnames[0]]
    diff = 0
    for r, (ro, rn) in enumerate(zip(ws_o.iter_rows(values_only=True),
                                     ws_n.iter_rows(values_only=True)), 1):
        mapped = list(rn[:31]) + list(rn[34:60])  # 去掉新增三列后应等于原行
        if list(ro) != mapped:
            diff += 1
            if diff <= 3:
                print(f"  ✗ 行{r} 非插入列不一致")
    assert diff == 0, f"{diff} 行非插入列被改动"
    print(f"[断言通过] 其余 57 列与原文件逐格一致（格式严格不变）")
    wb_o.close()
    wb_n.close()

    # 幂等：三列已存在时再次写入不得重复插入（多工厂循环复用同一副本）
    from app.nodes.writer import _ensure_three_columns
    wb2 = load_workbook(out_path)
    ws2 = wb2[wb2.sheetnames[0]]
    n_cols_before = ws2.max_column
    _ensure_three_columns(ws2)
    assert ws2.max_column == n_cols_before, "三列已存在却被重复插入"
    wb2.close()
    print(f"[断言通过] 三列已存在时跳过插入（幂等，多工厂循环安全）")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()
    if args.reset:
        reset_env()

    # ---- A：启动，跑到 Node5 挂起 ----
    print(f"\n===== 启动真实流程 thread_id={THREAD_ID}（LLM 真实调用）=====")
    result = service.run_until_interrupt(
        THREAD_ID, downstream_file_path=DOWNSTREAM_FILE,
        factory_filter=[TARGET_FACTORY_JP],
    )
    assert result["status"] == "pending_human_review", f"未挂起: {result}"
    rd = result["review_data"]
    print(f"\n===== 挂起！审核 payload 摘要 =====")
    print(f"工厂: {rd['factory_name']}  文件夹: {rd['folder_path']}")
    print(f"源单据: {len(rd['source_documents'])} 个  SKU: {len(rd['items'])} 个  "
          f"缺失: {rd['missing_skus']}")
    cov = rd.get("extraction_coverage") or {}
    print(f"覆盖率: {cov.get('extracted')}/{cov.get('expected')}  "
          f"extra: {cov.get('extra')}")
    for issue in rd.get("extraction_issues") or []:
        tag = {"blocking": "🔴", "warning": "🟡", "info": "ℹ️"}.get(issue["level"], "?")
        print(f"  {tag} {issue['type']}: {issue['message'][:80]}")
    st = {s: sum(1 for i in rd["items"] if i["status"] == s)
          for s in {i["status"] for i in rd["items"]}}
    print(f"状态分布: {st}")

    # ---- B：payload 对 GT 逐格校验 ----
    print(f"\n===== 提取值 vs GT 逐格校验 =====")
    errors = check_payload_against_gt(rd)
    if errors:
        for e in errors:
            print(f"  ✗ {e}")
        print(f"\n❌ 联调失败：{len(errors)} 处不一致")
        return 1
    print(f"  ✓ {len(rd['items'])} 个 SKU 全字段一致（件数/净重/毛重）")

    # ---- C：resume approved（不动数值，新 SKU 补合规字段）----
    print(f"\n===== resume approved =====")
    human_items = []
    for i, item in enumerate(rd["items"]):
        h = {"sku": item["sku"],
             "extracted_data": dict(item["extracted_data"]),
             "calculation": dict(item["calculation"] or {})}
        if item.get("is_new_sku"):
            h["name_cn"] = (item["extracted_data"].get("sku_name")
                            or f"待补品名-{i + 1}")
            h["hs_code"] = "9404909000"
            h["inspection_required"] = False
        human_items.append(h)
    resume_result = service.resume_order(THREAD_ID, {"approved": True, "items": human_items})
    print(f"resume 结果: {json.dumps(resume_result, ensure_ascii=False)[:300]}")

    # ---- D：断言 ----
    print(f"\n===== 断言 =====")
    out_path = Path(resume_result.get("final_output_path") or "")
    assert out_path.exists(), f"输出 Excel 不存在: {out_path}"
    print(f"[断言通过] 输出 Excel: {out_path}")
    check_output_excel(out_path, rd)

    with get_session() as s:
        rows = s.scalars(select(FactorySKU)).all()
        assert len(rows) == len(rd["items"]), \
            f"factory_skus 行数 {len(rows)} != 提取 SKU 数 {len(rd['items'])}"
        sample = rows[0]
        print(f"[断言通过] factory_skus {len(rows)} 行，示例: sku={sample.sku_code} "
              f"name_cn={sample.name_cn} unit_net={sample.unit_net_weight}")
        # 落库单重 = GT 总重/件数（纯 Python 除法结果）
        gt = get_factory_ground_truth(TARGET_FACTORY_LOCAL)
        for r in rows:
            g = gt.get(r.sku_code)
            if g and g["total_quantity"]:
                expect_net = g["total_net_weight"] / g["total_quantity"]
                actual = float(r.unit_net_weight) if r.unit_net_weight is not None else None
                assert actual is not None and abs(actual - expect_net) < TOL, \
                    f"{r.sku_code} 落库单净重 {actual} != {expect_net}"
        print(f"[断言通过] 全部落库单重 = GT 总重/件数")

    state = service.get_order_state(THREAD_ID)
    assert state["values"].get("validation_status") == "Approved"
    assert resume_result["status"] == "success"
    print(f"[断言通过] validation_status == Approved，流程走到 END")

    print(f"\n🎉 Node3 真实联调端到端全部通过！")
    return 0


if __name__ == "__main__":
    sys.exit(main())
